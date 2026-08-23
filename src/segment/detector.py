"""Region / repeated-block detection (설계문서 §4.3, src.segment).

Strategy (문서 §4.3 알고리즘을 일반화):
1. 셀·병합·스타일·이미지 anchor를 좌표계에 올린다 (SheetGrid).
2. 범례(legend) 영역을 먼저 찾아 로컬 색 의미(style semantics)를 만든다 (§10.2).
3. 반복되는 제목 스타일 시그니처(병합 폭 + fill + bold + 시작 열)로 동일 template
   block들을 묶는다 — 제목 텍스트가 아닌 레이아웃 반복성 기반이므로 행 간격이
   달라도 동작한다 (§4.2 layout fingerprint).
4. 각 block 내부를 행 단위 상태기계로 TABLE / KEY_VALUE / PROFILE / SUMMARY /
   NOTE 로 분류하고, 병합 span 트리로 계층형 header_path를 만든다.
5. Block 단위 record 후보를 만들고 field/image를 귀속시킨다.
"""
from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field as dc_field

from openpyxl.utils import get_column_letter, range_boundaries

from src.common.models import (
    BlockInfo,
    CellInfo,
    FieldInfo,
    ImageInfo,
    Region,
    SheetSegmentation,
    SheetStructure,
)

LEGEND_KEYWORDS = ("범례", "의미", "구분", "legend")

# style_role inference from legend meaning text (문서 §10.2: 로컬 범례 우선)
_ROLE_PATTERNS = [
    (re.compile(r"입력|수기"), "input"),
    (re.compile(r"PLC|자동측정", re.I), "measured"),
    (re.compile(r"계산|수식|자동|참조"), "calculated"),
    (re.compile(r"이상|이탈|오류|error|불량|규격", re.I), "error"),
    (re.compile(r"확인|주의|warn", re.I), "warning"),
    (re.compile(r"정상|합격|ok", re.I), "ok"),
]

# 인라인 범례: "파랑=PLC" / "노란색=수기입력 / 파란색=PLC" (한 셀에 여러 쌍 가능)
_INLINE_LEGEND_PAIR_RE = re.compile(r"(파랑|파란색|파란글씨|노랑|노란색|회색|빨강|빨간색|빨간글씨|녹색|초록)\s*=\s*([^/=]+)")

_COLOR_CLASS = {"파랑": "blue", "파란색": "blue", "파란글씨": "blue",
                "노랑": "yellow", "노란색": "yellow",
                "회색": "gray", "빨강": "red", "빨간색": "red", "빨간글씨": "red",
                "녹색": "green", "초록": "green"}

# 스펙/범위 토큰 ("≤80", "110~190", "47~52") — 헤더 경로에서 제약으로 분리
_SPEC_TOKEN_RE = re.compile(r"^\s*[≤≥<>]?\s*\d+(\.\d+)?\s*([~\-]\s*\d+(\.\d+)?)?\s*$")

# units 미제공 시의 최소 단위 토큰 집합
_FALLBACK_UNITS = {"℃", "°C", "degC", "K", "°F", "bar", "kPa", "MPa", "Pa", "kg", "g",
                   "t", "ton", "metric ton", "mm", "cm", "%", "wt%", "mass%", "ppm",
                   "cP", "Pa·s", "Pa.s", "mPa·s", "kWh", "MWh", "rpm", "EA", "pcs",
                   "mm/s", "0~1", "mass fraction", "h", "min", "text", "enum", "timestamp"}


def classify_hue(rgb: str) -> str | None:
    """'FFDDEBF7' → blue/yellow/gray/red/green 대략 분류 (인라인 범례 매칭용)."""
    h = rgb[-6:]
    try:
        r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    except ValueError:
        return None
    mx, mn = max(r, g, b), min(r, g, b)
    if mx - mn <= 12:
        return "gray" if mx < 245 else None   # 흰색 계열은 의미 없음
    if r >= g >= b and (r - b) > 20 and (r - g) < 40:
        return "yellow"
    if r > g and r > b:
        return "red"
    if b > r and b >= g:
        return "blue"
    if g > r and g > b:
        return "green"
    return None

_UNIT_IN_LABEL_RE = re.compile(r"^(.*?)[\s]*\(([^()]{1,10})\)\s*$")
_UNIT_TOKEN_RE = re.compile(r"^[A-Za-z%℃°/·\^0-9\.\-]+$")

# column-header semantics for row-oriented tables (개념=행 방향)
_LIMIT_RE = re.compile(r"기준|상한|하한|한계|규격|spec", re.I)
_MEASURED_RE = re.compile(r"실측|측정|값|value", re.I)
_JUDGE_RE = re.compile(r"판정|결과|합부|result", re.I)
_UNIT_COL_RE = re.compile(r"^\s*단위\s*$|^\s*units?\s*$", re.I)
# 첫 열 헤더가 '이름' 성격이면 행은 개념이 아니라 인스턴스다 (예: 재료명/품명)
_INSTANCE_NAME_COL_RE = re.compile(r"명\s*(\(|$)|이름|name", re.I)
_SUMMARY_LABEL_RE = re.compile(r"종합|최종|overall|final", re.I)


def split_label_unit(label: str) -> tuple[str, str | None]:
    """'투입량(kg)' -> ('투입량', 'kg'); '치수 검사 (mm)' -> ('치수 검사', 'mm')."""
    m = _UNIT_IN_LABEL_RE.match(label or "")
    if m and _UNIT_TOKEN_RE.match(m.group(2).strip()):
        return m.group(1).strip(), m.group(2).strip()
    return (label or "").strip(), None


class SheetGrid:
    """Random-access view over a SheetStructure with merge resolution."""

    def __init__(self, sheet: SheetStructure):
        self.sheet = sheet
        self.by_addr: dict[str, CellInfo] = {c.address: c for c in sheet.cells}
        self.by_rc: dict[tuple[int, int], CellInfo] = {(c.row, c.col): c for c in sheet.cells}
        # merge master lookup for covered coordinates
        self.cover_master: dict[tuple[int, int], CellInfo] = {}
        for c in sheet.cells:
            if c.merged_range:
                mn_c, mn_r, mx_c, mx_r = range_boundaries(c.merged_range)
                for r in range(mn_r, mx_r + 1):
                    for col in range(mn_c, mx_c + 1):
                        self.cover_master[(r, col)] = c

    def cell(self, row: int, col: int) -> CellInfo | None:
        return self.by_rc.get((row, col))

    def resolved(self, row: int, col: int) -> CellInfo | None:
        """Cell at (row,col), following merges to the master cell."""
        c = self.by_rc.get((row, col))
        if c is not None and c.value is not None:
            return c
        return self.cover_master.get((row, col))

    def row_values(self, row: int, min_col: int, max_col: int) -> list[CellInfo]:
        out = []
        for col in range(min_col, max_col + 1):
            c = self.by_rc.get((row, col))
            if c is not None and c.value is not None:
                out.append(c)
        return out

    def merge_width(self, c: CellInfo) -> int:
        if not c.merged_range:
            return 1
        mn_c, _, mx_c, _ = range_boundaries(c.merged_range)
        return mx_c - mn_c + 1


def _is_text(c: CellInfo) -> bool:
    return isinstance(c.value, str) and not c.is_formula and c.value.strip() != ""


def _fingerprint(parts: list[str]) -> str:
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:16]


@dataclass
class _RowClass:
    kind: str                      # BLANK/TITLE/HEADER/KV/DATA/SUMMARY/NOTE
    cells: list[CellInfo] = dc_field(default_factory=list)


class RegionDetector:
    def __init__(self, parser_rules: dict | None = None, units=None):
        self.parser_rules = parser_rules or {}
        self.units = units                 # UnitRegistry (optional)
        self._legend_rows: set[int] = set()

    # ------------------------------------------------------------- tokens ----
    # 단위행에 섞이는 스키마 서술 토큰 — 경로에서 제거하되 단위로는 쓰지 않는다
    PSEUDO_UNIT_TOKENS = {"text", "enum", "timestamp", "h", "min", "unit", "memo",
                          "name", "qty", "visual", "rule", "manual", "flag"}

    def _is_unit_token(self, text: str) -> bool:
        t = (text or "").strip()
        if not t or len(t) > 14:
            return False
        if t.lower() in self.PSEUDO_UNIT_TOKENS:
            return True
        if self.units is not None:
            norm = self.units.normalize_unit(t)
            return bool(self.units.dimensions_of(norm))
        return t in _FALLBACK_UNITS

    def _split_label_unit(self, label: str) -> tuple[str, str | None]:
        """split_label_unit + 단위 검증: '(KR)'/'(EN)' 같은 비단위 괄호는
        단위로 취급하지 않는다 (registry가 있으면 등록 단위만 인정)."""
        base, unit = split_label_unit(label)
        if unit is None:
            return base, None
        if self.units is not None:
            if self.units.dimensions_of(unit):
                return base, unit
            return (label or "").strip(), None
        return (base, unit) if unit in _FALLBACK_UNITS or not unit.isalpha() or len(unit) > 2 \
            else ((label or "").strip(), None)

    # ------------------------------------------------------------ legend ----
    def detect_inline_legend(self, grid: SheetGrid) -> tuple[dict[str, str], set[int]]:
        """셀 텍스트 안의 '파랑=PLC' 형태 범례를 찾아 hue 매칭으로 fill→의미를 만든다."""
        pairs_by_row: dict[int, dict[str, str]] = {}
        for c in grid.sheet.cells:
            if not _is_text(c):
                continue
            found = _INLINE_LEGEND_PAIR_RE.findall(str(c.value))
            if found:
                row_pairs = pairs_by_row.setdefault(c.row, {})
                for name, meaning in found:
                    row_pairs[_COLOR_CLASS[name]] = meaning.strip()
        semantics: dict[str, str] = {}
        legend_rows: set[int] = set()
        for row, class_map in pairs_by_row.items():
            if len(class_map) < 2:
                continue
            legend_rows.add(row)
            fills = {c.fill_rgb for c in grid.sheet.cells if c.fill_rgb}
            for rgb in fills:
                cls = classify_hue(rgb)
                if cls in class_map:
                    semantics.setdefault(rgb, class_map[cls])
        return semantics, legend_rows

    def detect_legend(self, grid: SheetGrid) -> tuple[str | None, dict[str, str]]:
        for c in grid.sheet.cells:
            if _is_text(c) and any(k in str(c.value).lower() or k in str(c.value) for k in LEGEND_KEYWORDS):
                semantics: dict[str, str] = {}
                max_r = c.row
                for r in range(c.row + 1, c.row + 12):
                    swatch = grid.cell(r, c.col)
                    neighbor = grid.cell(r, c.col + 1)
                    if swatch is None or neighbor is None or not _is_text(neighbor):
                        break
                    if swatch.fill_rgb:
                        # 문서마다 의미 텍스트 위치가 다르다: 색 셀 안('정상') 또는
                        # 옆 셀('검사자 입력'). role 패턴에 맞는 쪽을 의미로 채택한다.
                        neighbor_txt = str(neighbor.value).lstrip("# ").strip()
                        swatch_txt = str(swatch.value).strip() if _is_text(swatch) else ""
                        meaning = neighbor_txt
                        if not any(p.search(neighbor_txt) for p, _ in _ROLE_PATTERNS):
                            if swatch_txt and any(p.search(swatch_txt) for p, _ in _ROLE_PATTERNS):
                                meaning = swatch_txt
                        semantics[swatch.fill_rgb] = meaning
                        max_r = r
                if semantics:
                    bbox = f"{get_column_letter(c.col)}{c.row}:{get_column_letter(c.col + 1)}{max_r}"
                    return bbox, semantics
        return None, {}

    @staticmethod
    def style_roles_from_semantics(semantics: dict[str, str]) -> dict[str, str]:
        roles: dict[str, str] = {}
        for rgb, meaning in semantics.items():
            for pat, role in _ROLE_PATTERNS:
                if pat.search(meaning):
                    roles[rgb] = role
                    break
        return roles

    # ------------------------------------------------------- block titles ----
    def detect_block_headers(self, grid: SheetGrid, legend_bbox: str | None) -> list[CellInfo]:
        legend_cols: set[int] = set()
        if legend_bbox:
            mn_c, _, mx_c, _ = range_boundaries(legend_bbox)
            legend_cols = set(range(mn_c, mx_c + 1))

        candidates: list[CellInfo] = []
        for c in grid.sheet.cells:
            if not _is_text(c) or c.col in legend_cols:
                continue
            width = grid.merge_width(c)
            if width >= 3 and (c.bold or c.fill_rgb):
                candidates.append(c)

        groups: dict[tuple, list[CellInfo]] = {}
        for c in candidates:
            sig = (c.col, grid.merge_width(c), c.fill_rgb, c.bold)
            groups.setdefault(sig, []).append(c)

        # 반복 그룹이 여럿이면(예: block 제목 + 내부 소제목 + 특이사항 merge가 모두
        # 반복) 최상위 파티션인 그룹 하나만 block 경계로 쓴다. 최상위 그룹은 가장
        # 먼저 등장하는 반복 그룹이다 — 내부 구조는 항상 자기 block 제목 아래에 온다.
        repeated_groups = [cells for cells in groups.values() if len(cells) >= 2]
        if repeated_groups:
            top = min(repeated_groups, key=lambda cells: min(c.row for c in cells))
            return sorted(top, key=lambda c: c.row)
        return []

    # -------------------------------------------------------- row classes ----
    def _classify_row(self, grid: SheetGrid, row: int, min_col: int, max_col: int,
                      block_width: int, in_table: bool, table_cols: list[int]) -> _RowClass:
        if row in self._legend_rows:
            return _RowClass("BLANK")
        cells = grid.row_values(row, min_col, max_col)
        if not cells:
            return _RowClass("BLANK")

        # NOTE: one wide merged text cell
        if len(cells) == 1 and cells[0].merged_range and _is_text(cells[0]):
            if grid.merge_width(cells[0]) >= max(3, block_width - 3):
                return _RowClass("NOTE", cells)

        # HEADER band: >=2 bold cells sharing one fill
        bold_fills = [c.fill_rgb for c in cells if c.bold and c.fill_rgb and _is_text(c)]
        if len(bold_fills) >= 2 and len(set(bold_fills)) == 1:
            return _RowClass("HEADER", cells)

        if in_table and table_cols:
            hits = sum(1 for c in cells if c.col in table_cols)
            if hits >= max(2, int(len(table_cols) * 0.6)):
                return _RowClass("DATA", cells)

        # KV: >=2 (text label, value) adjacent pairs
        pairs = 0
        by_col = {c.col: c for c in cells}
        for c in cells:
            if _is_text(c) and (by_col.get(c.col + 1) or by_col.get(c.col + 2)):
                nxt = by_col.get(c.col + 1) or by_col.get(c.col + 2)
                if nxt is not None and nxt.col > c.col and not (_is_text(nxt) and grid.merge_width(nxt) > 2):
                    pairs += 1
        if pairs >= 2:
            return _RowClass("KV", cells)

        # SUMMARY: label + one value (label first, textual, value often formula)
        if 1 <= len(cells) <= 3 and _is_text(cells[0]):
            others = cells[1:]
            if others and any((not _is_text(c)) or c.is_formula for c in others):
                return _RowClass("SUMMARY", cells)
            if not others and cells[0].merged_range is None:
                return _RowClass("SUMMARY", cells)

        if pairs == 1:
            return _RowClass("KV", cells)
        return _RowClass("NOTE", cells) if all(_is_text(c) for c in cells) else _RowClass("DATA", cells)

    # ---------------------------------------------------------- segments ----
    def segment_sheet(self, sheet: SheetStructure) -> SheetSegmentation:
        grid = SheetGrid(sheet)
        legend_bbox, semantics = self.detect_legend(grid)
        inline_semantics, legend_rows = self.detect_inline_legend(grid)
        self._legend_rows = legend_rows
        for rgb, meaning in inline_semantics.items():
            semantics.setdefault(rgb, meaning)
        style_roles = self.style_roles_from_semantics(semantics)
        # approved per-document rules can extend/override inferred roles (§10.2)
        for rgb, role in (self.parser_rules.get("style_roles") or {}).items():
            style_roles[rgb] = role

        seg = SheetSegmentation(
            sheet_name=sheet.sheet_name,
            sheet_index=sheet.sheet_index,
            style_semantics=semantics,
            legend_bbox=legend_bbox,
        )

        headers = self.detect_block_headers(grid, legend_bbox)
        if headers:
            # 반복 제목 기반 block 분할 (기존 경로)
            for i, h in enumerate(headers):
                top = h.row
                bottom = headers[i + 1].row - 1 if i + 1 < len(headers) else sheet.max_row
                mn_c, _, mx_c, _ = range_boundaries(h.merged_range) if h.merged_range else (h.col, h.row, h.col, h.row)
                block = self._build_block(grid, str(h.value).strip(), h.address,
                                          top, top + 1, bottom, mn_c, mx_c, style_roles,
                                          f"{h.col}w{grid.merge_width(h)}")
                seg.blocks.append(block)
        else:
            # 반복 구조가 없는 시트: 열 밴드 × 빈 행 기반 섹션 분할 (다영역/병렬 지원)
            seg.blocks.extend(self._segment_sections(grid, sheet, style_roles))

        # image attribution: anchor row belongs to the covering block, else nearest
        for img in sheet.images:
            anchor_row = img.anchor_row + 1  # anchors are 0-based
            target = None
            for b in seg.blocks:
                if b.min_row <= anchor_row <= b.max_row:
                    target = b
                    break
            if target is None and seg.blocks:
                target = min(seg.blocks, key=lambda b: min(abs(b.min_row - anchor_row), abs(b.max_row - anchor_row)))
            if target is not None:
                target.images.append(img)
        return seg

    def _segment_sections(self, grid: SheetGrid, sheet: SheetStructure,
                          style_roles: dict[str, str]) -> list[BlockInfo]:
        """반복 제목이 없는 시트: 빈 열로 밴드를 나누고 빈 행으로 섹션을 나눈다.

        좌우 병렬 블록([Block A]|[Block B], AREA-1|AREA-3)과 세로 다영역
        (메타 KV / 메인 표 / 부속 로그)을 모두 독립 block으로 분리한다.
        """
        value_cells = [c for c in sheet.cells
                       if c.value is not None and c.row not in self._legend_rows]
        if not value_cells:
            return []

        # 열 밴드: 값이 있는 열의 연속 구간
        cols_used = sorted({c.col for c in value_cells})
        bands: list[tuple[int, int]] = []
        start = prev = cols_used[0]
        for col in cols_used[1:]:
            if col > prev + 1:
                bands.append((start, prev))
                start = col
            prev = col
        bands.append((start, prev))

        blocks: list[BlockInfo] = []
        for mn_c, mx_c in bands:
            band_cells = [c for c in value_cells if mn_c <= c.col <= mx_c]
            rows_used = sorted({c.row for c in band_cells})
            if not rows_used:
                continue
            sections: list[list[int]] = [[rows_used[0]]]
            for r in rows_used[1:]:
                if r > sections[-1][-1] + 1:
                    sections.append([])
                sections[-1].append(r)
            for sec_rows in sections:
                top, bottom = sec_rows[0], sec_rows[-1]
                # 섹션 첫 행이 단독 wide 제목이면 제목으로 소비
                first = [c for c in band_cells if c.row == top]
                title, title_addr, content_top = grid.sheet.sheet_name, None, top
                if len(first) == 1 and _is_text(first[0]) and (
                        grid.merge_width(first[0]) >= 3 or bottom == top):
                    title = str(first[0].value).strip()
                    title_addr = first[0].address
                    content_top = top + 1
                if content_top > bottom:
                    continue  # 제목뿐인 섹션 (시트 타이틀 등)
                block = self._build_block(grid, title, title_addr or f"R{top}",
                                          top, content_top, bottom, mn_c, mx_c,
                                          style_roles, f"band{mn_c}-{mx_c}")
                if any(r.fields for r in block.regions) or block.images:
                    blocks.append(block)
        return blocks

    def _build_block(self, grid: SheetGrid, title: str, title_address: str,
                     top: int, content_top: int, bottom: int,
                     min_col: int, max_col: int, style_roles: dict[str, str],
                     width_sig: str) -> BlockInfo:
        block = BlockInfo(
            block_id=f"{grid.sheet.sheet_name}!R{top}",
            sheet_name=grid.sheet.sheet_name,
            title=title,
            title_address=title_address,
            min_row=top,
            max_row=bottom,
        )
        block_width = max_col - min_col + 1

        rows: list[tuple[int, _RowClass]] = []
        in_table = False
        table_cols: list[int] = []
        for r in range(content_top, bottom + 1):
            rc = self._classify_row(grid, r, min_col, max_col, block_width, in_table, table_cols)
            if rc.kind == "HEADER":
                if not in_table:
                    table_cols = []
                in_table = True
                table_cols = sorted(set(table_cols) | {c.col for c in rc.cells})
            elif rc.kind in ("BLANK", "NOTE", "KV", "SUMMARY"):
                in_table = False
                if rc.kind != "DATA":
                    table_cols = [] if rc.kind == "BLANK" else table_cols
            rows.append((r, rc))

        # group consecutive rows into regions
        i = 0
        ridx = 0
        while i < len(rows):
            r, rc = rows[i]
            if rc.kind == "BLANK":
                i += 1
                continue
            if rc.kind == "HEADER":
                hdr_rows = [rows[i]]
                j = i + 1
                while j < len(rows) and rows[j][1].kind == "HEADER":
                    hdr_rows.append(rows[j])
                    j += 1
                data_rows = []
                while j < len(rows) and rows[j][1].kind == "DATA":
                    data_rows.append(rows[j])
                    j += 1
                region = self._build_table_region(grid, hdr_rows, data_rows, min_col, max_col, style_roles, ridx, block.block_id)
                block.regions.append(region)
                ridx += 1
                i = j
                continue
            same = [rows[i]]
            j = i + 1
            while j < len(rows) and rows[j][1].kind == rc.kind:
                same.append(rows[j])
                j += 1
            if rc.kind == "KV":
                block.regions.append(self._build_kv_region(grid, same, style_roles, ridx, block.block_id))
            elif rc.kind == "SUMMARY":
                block.regions.append(self._build_summary_region(grid, same, style_roles, ridx, block.block_id))
            else:  # NOTE
                block.regions.append(self._build_note_region(grid, same, ridx, block.block_id))
            ridx += 1
            i = j

        parts = [f"{reg.region_type}:{len(reg.fields)}" for reg in block.regions]
        block.layout_fingerprint = _fingerprint([width_sig] + parts)
        return block

    # ---------------------------------------------------- region builders ----
    def _bbox(self, rows: list[tuple[int, _RowClass]]) -> tuple[str, int, int, int, int]:
        min_r = min(r for r, _ in rows)
        max_r = max(r for r, _ in rows)
        cols = [c.col for _, rc in rows for c in rc.cells]
        min_c, max_c = (min(cols), max(cols)) if cols else (1, 1)
        return (
            f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}",
            min_r, max_r, min_c, max_c,
        )

    def _mk_field(self, block_id: str, label_cell: CellInfo | None, value_cell: CellInfo,
                  raw_label: str, header_path: list[str], unit: str | None,
                  style_roles: dict[str, str], row_key: str | None = None) -> FieldInfo:
        style_role = style_roles.get(value_cell.fill_rgb or "", "unknown")
        return FieldInfo(
            field_id=f"{block_id}/{value_cell.address}",
            address=value_cell.address,
            label_address=label_cell.address if label_cell else None,
            raw_label=raw_label,
            header_path=header_path,
            raw_value=value_cell.value,
            cached_value=value_cell.cached_value,
            is_formula=value_cell.is_formula,
            formula=value_cell.formula,
            formula_refs=value_cell.formula_refs,
            raw_unit=unit,
            style_role="calculated" if value_cell.is_formula and style_role == "unknown" else style_role,
            fill_rgb=value_cell.fill_rgb,
            row_key=row_key,
        )

    def _build_kv_region(self, grid, rows, style_roles, ridx, block_id) -> Region:
        bbox, mn_r, mx_r, mn_c, mx_c = self._bbox(rows)
        region = Region(region_id=f"{block_id}/r{ridx}", region_type="KEY_VALUE",
                        bbox=bbox, min_row=mn_r, max_row=mx_r, min_col=mn_c, max_col=mx_c)
        for r, rc in rows:
            # 순차 소비: (label, value), (label, value)... 값 셀이 다음 쌍의 label로
            # 재사용되지 않도록 쌍 단위로 스캔한다.
            cells = sorted(rc.cells, key=lambda c: c.col)
            i = 0
            while i < len(cells) - 1:
                c, v = cells[i], cells[i + 1]
                if _is_text(c) and v.col - c.col <= 2 + grid.merge_width(c) - 1:
                    label, unit = self._split_label_unit(str(c.value))
                    region.fields.append(self._mk_field(block_id, c, v, label, [label], unit, style_roles))
                    i += 2
                else:
                    i += 1
        region.layout_fingerprint = _fingerprint(["KV"] + sorted(f.raw_label for f in region.fields))
        return region

    def _header_paths(self, grid: SheetGrid, hdr_rows, data_cols: list[int]
                      ) -> tuple[dict[int, list[str]], dict[int, str], dict[int, str]]:
        """헤더 경로 + 열별 단위/규격.

        다층 헤더 안에 단위 행("degC", "kg")이나 스펙 행("110~190", "≤80")이
        섞여 있으면 header_path에서 분리해 열 단위/제약으로 보존한다.
        """
        paths: dict[int, list[str]] = {}
        col_units: dict[int, str] = {}
        col_specs: dict[int, str] = {}
        for col in data_cols:
            raw_path: list[str] = []
            for r, _ in hdr_rows:
                cell = grid.resolved(r, col)
                if cell is not None and _is_text(cell):
                    txt = str(cell.value).strip()
                    if not raw_path or raw_path[-1] != txt:
                        raw_path.append(txt)
            cleaned: list[str] = []
            for el in raw_path:
                if self._is_unit_token(el):
                    # 'text'/'unit'/'memo' 같은 서술 토큰은 경로에서만 제거하고
                    # 실제 단위(차원 있음)만 열 단위로 기록한다
                    pseudo = el.strip().lower() in self.PSEUDO_UNIT_TOKENS
                    if not pseudo and (self.units is None or self.units.dimensions_of(el)):
                        col_units[col] = el
                elif _SPEC_TOKEN_RE.match(el):
                    col_specs[col] = el
                else:
                    cleaned.append(el)
            paths[col] = cleaned if cleaned else raw_path
        return paths, col_units, col_specs

    def _build_table_region(self, grid, hdr_rows, data_rows, min_col, max_col,
                            style_roles, ridx, block_id) -> Region:
        header_cols = sorted({c.col for _, rc in hdr_rows for c in rc.cells})
        # columns covered by merged headers count too
        all_cols = set(header_cols)
        for r, rc in hdr_rows:
            for c in rc.cells:
                if c.merged_range:
                    a, _, b, _ = range_boundaries(c.merged_range)
                    all_cols.update(range(a, b + 1))
        for _, rc in data_rows:
            all_cols.update(c.col for c in rc.cells)
        data_cols = sorted(all_cols)

        paths, col_units, col_specs = self._header_paths(grid, hdr_rows, data_cols)
        leaf = {col: (paths[col][-1] if paths[col] else "") for col in data_cols}

        rows_all = hdr_rows + data_rows
        bbox, mn_r, mx_r, mn_c, mx_c = self._bbox(rows_all)

        unit_cols = [c for c in data_cols if _UNIT_COL_RE.match(leaf[c] or "")]
        orientation = "row_concept" if unit_cols else "col_concept"
        # 단위 열이 있어도 첫 열이 '재료명' 같은 인스턴스 이름 열이면
        # 행=인스턴스(col_concept)다 — 행 라벨을 개념으로 오인하지 않는다
        if orientation == "row_concept" and data_cols:
            first_leaf = leaf.get(data_cols[0]) or ""
            if _INSTANCE_NAME_COL_RE.search(first_leaf):
                orientation = "col_concept"

        # row-key column: 앞쪽 열부터 탐색해 전부 텍스트인 첫 열을 채택한다.
        # 세로 병합 그룹 열(예: 반복 2행에 걸친 Time/부재료)은 건너뛰며 계속 보고,
        # 일반 값 열을 만나면 중단해 중간의 메모/판정 열을 키로 오인하지 않는다.
        key_col = None
        for col in data_cols:
            vals = [grid.cell(r, col) for r, _ in data_rows]
            vals = [v for v in vals if v is not None and v.value is not None]
            if vals and all(_is_text(v) for v in vals):
                key_col = col
                break
            has_vmerge = any(
                v.merged_range and (lambda b: b[0] == b[2] and b[3] > b[1])(
                    range_boundaries(v.merged_range))
                for v in vals)
            if not has_vmerge:
                break

        region_type = "TABLE"
        if orientation == "col_concept" and len(data_rows) == 1:
            # single measurement row + group title carrying a unit → horizontal profile
            group_units = [self._split_label_unit(p)[1] for path in paths.values() for p in path]
            if any(u for u in group_units):
                region_type = "PROFILE"

        region = Region(region_id=f"{block_id}/r{ridx}", region_type=region_type,
                        bbox=bbox, min_row=mn_r, max_row=mx_r, min_col=mn_c, max_col=mx_c,
                        orientation=orientation)

        if orientation == "row_concept":
            self._extract_row_concept(grid, region, data_rows, data_cols, leaf, paths,
                                      unit_cols[0], key_col, style_roles, block_id)
        else:
            self._extract_col_concept(grid, region, data_rows, data_cols, leaf, paths,
                                      key_col, style_roles, block_id, col_units)

        region.layout_fingerprint = _fingerprint(
            [region_type, orientation] + [leaf[c] for c in data_cols])
        return region

    def _extract_row_concept(self, grid, region, data_rows, data_cols, leaf, paths,
                             unit_col, key_col, style_roles, block_id):
        """Each data row is a concept (e.g. 진동/토출압력); columns give roles."""
        label_col = key_col if key_col is not None else data_cols[0]
        for r, rc in data_rows:
            by_col = {c.col: c for c in rc.cells}
            label_cell = by_col.get(label_col)
            if label_cell is None or not _is_text(label_cell):
                continue
            concept_label, label_unit = self._split_label_unit(str(label_cell.value))
            unit_cell = by_col.get(unit_col)
            row_unit = str(unit_cell.value).strip() if unit_cell is not None and unit_cell.value is not None else label_unit
            for col in data_cols:
                if col in (label_col, unit_col):
                    continue
                v = by_col.get(col)
                if v is None or v.value is None:
                    continue
                col_hdr = leaf[col]
                f = self._mk_field(block_id, label_cell, v, concept_label,
                                   [*paths[label_col][:-1], concept_label, col_hdr] if paths[label_col] else [concept_label, col_hdr],
                                   row_unit, style_roles, row_key=col_hdr)
                if _JUDGE_RE.search(col_hdr):
                    f.style_role = "result"
                    f.raw_unit = None
                elif _LIMIT_RE.search(col_hdr):
                    f.style_role = f.style_role if f.style_role not in ("unknown",) else "input"
                    f.row_key = col_hdr
                region.fields.append(f)

    def _extract_col_concept(self, grid, region, data_rows, data_cols, leaf, paths,
                             key_col, style_roles, block_id, col_units=None):
        """Each column leaf header is a concept (e.g. 외경/길이); rows are instances."""
        col_units = col_units or {}
        # '온도 프로파일 (℃)' 같은 단위를 가진 그룹 제목이 key 열 헤더에 있으면
        # 나머지 열의 문맥/단위로 상속한다 (가로형 profile, §2.3).
        group_path: list[str] = []
        group_unit: str | None = None
        # 그룹 제목 상속은 key 열이 있을 때만 — 첫 데이터 열을 그룹으로 오인 금지
        anchor_col = key_col
        if anchor_col is not None and paths.get(anchor_col):
            _, u = self._split_label_unit(paths[anchor_col][-1])
            if u:
                group_path = paths[anchor_col]
                group_unit = u
        for r, rc in data_rows:
            by_col = {c.col: c for c in rc.cells}
            key_cell = by_col.get(key_col) if key_col is not None else None
            if key_cell is None and key_col is not None:
                # 세로 병합된 key 셀(예: 부재료가 반복 2행에 걸쳐 병합) 값 상속
                master = grid.resolved(r, key_col)
                if master is not None and master.merged_range:
                    a, b_, c_, d_ = range_boundaries(master.merged_range)
                    if a == c_ and b_ <= r <= d_:
                        key_cell = master
            row_key = str(key_cell.value).strip() if key_cell is not None and key_cell.value is not None else f"row{r}"
            for col in data_cols:
                if col == key_col:
                    continue
                v = by_col.get(col)
                if v is None or v.value is None:
                    continue
                path = paths[col]
                if not path:
                    continue
                label, unit = self._split_label_unit(path[-1])
                if unit is None:
                    for p in path[:-1]:
                        _, u = self._split_label_unit(p)
                        if u:
                            unit = u
                            break
                unit = unit or col_units.get(col)   # 단위 헤더 행에서 온 열 단위
                full_path = path
                if group_path and group_path[-1] not in path:
                    full_path = [*group_path, *path]
                    unit = unit or group_unit
                f = self._mk_field(block_id, key_cell, v, label, full_path, unit, style_roles, row_key=row_key)
                if _JUDGE_RE.search(label):
                    f.style_role = "result"
                    f.raw_unit = None
                region.fields.append(f)

    def _build_summary_region(self, grid, rows, style_roles, ridx, block_id) -> Region:
        bbox, mn_r, mx_r, mn_c, mx_c = self._bbox(rows)
        region = Region(region_id=f"{block_id}/r{ridx}", region_type="SUMMARY",
                        bbox=bbox, min_row=mn_r, max_row=mx_r, min_col=mn_c, max_col=mx_c)
        for r, rc in rows:
            label_cell = rc.cells[0]
            label, unit = self._split_label_unit(str(label_cell.value))
            for v in rc.cells[1:]:
                f = self._mk_field(block_id, label_cell, v, label, [label], unit, style_roles)
                f.style_role = "result" if (_JUDGE_RE.search(label) or _SUMMARY_LABEL_RE.search(label)) else f.style_role
                region.fields.append(f)
        region.layout_fingerprint = _fingerprint(["SUMMARY"] + sorted(f.raw_label for f in region.fields))
        return region

    def _build_note_region(self, grid, rows, ridx, block_id) -> Region:
        bbox, mn_r, mx_r, mn_c, mx_c = self._bbox(rows)
        text = " ".join(str(c.value).strip() for _, rc in rows for c in rc.cells if _is_text(c))
        region = Region(region_id=f"{block_id}/r{ridx}", region_type="NOTE",
                        bbox=bbox, min_row=mn_r, max_row=mx_r, min_col=mn_c, max_col=mx_c,
                        note_text=text)
        region.layout_fingerprint = _fingerprint(["NOTE"])
        return region


def segment_workbook(structure, parser_rules: dict | None = None, units=None,
                     skip_sheets: set[str] | None = None) -> list[SheetSegmentation]:
    """Segment every sheet of an inspected workbook (다중 시트 기본 전제, §2).

    skip_sheets: 문서 내장 사전 시트 등 레코드로 만들지 않을 시트 이름.
    """
    rules = parser_rules or {}
    skip = skip_sheets or set()
    out = []
    for sheet in structure.sheets:
        if sheet.sheet_name in skip:
            continue
        doc_rules = rules.get(structure.file_name, {}) if rules else {}
        det = RegionDetector(parser_rules=doc_rules, units=units)
        out.append(det.segment_sheet(sheet))
    return out
