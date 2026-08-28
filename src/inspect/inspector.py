"""WorkbookInspector — structural extraction of XLSX (설계문서 §4, §13 src.inspect).

Reads every sheet of a workbook and captures cell values, formulas with cached
values, fills, bold flags, merged ranges and image anchors. Nothing is thrown
away: downstream stages decide meaning (설계문서 §1 "시각적 요소는 장식이 아니다").
"""
from __future__ import annotations

import datetime
import json
import os
import re
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from src.common.hashing import sha256_bytes, sha256_file
from src.common.models import CellInfo, ImageInfo, SheetStructure, WorkbookStructure

PARSER_VERSION = "2.0.0"

# A1-style references inside formulas, including ranges ("B10:D12") and
# sheet-qualified refs ("Sheet1!A1" — sheet part captured separately).
_REF_RE = re.compile(r"(?:'[^']+'|[A-Za-z0-9_가-힣]+)?!?\$?([A-Z]{1,3})\$?([0-9]{1,7})(?::\$?([A-Z]{1,3})\$?([0-9]{1,7}))?")


def extract_formula_refs(formula: str) -> list[str]:
    """수식 참조 추출 — 시트 한정 참조('시트'!W15)는 시트명까지 보존한다
    (cross-sheet lineage, 설계문서 §10.3)."""
    refs: list[str] = []
    for m in _REF_RE.finditer(formula):
        refs.append(m.group(0).replace("$", ""))
    return refs


def _fill_rgb(cell) -> str | None:
    f = cell.fill
    if f is None or f.patternType is None:
        return None
    fg = f.fgColor
    if fg is not None and fg.type == "rgb" and fg.rgb and fg.rgb not in ("00000000",):
        return str(fg.rgb)
    return None


def _jsonable_value(v):
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return v.isoformat()
    return v


_BAD_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")


def _repair_sheet_names(path: Path) -> Path:
    """workbook.xml의 시트명에서 금지 문자를 '_'로 바꾼 임시 사본을 만든다."""
    import tempfile
    import zipfile

    tmp = Path(tempfile.mkstemp(suffix=".xlsx")[1])
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/workbook.xml":
                xml = data.decode("utf-8")
                xml = re.sub(
                    r'name="([^"]*)"',
                    lambda m: 'name="' + _BAD_SHEET_CHARS.sub("_", m.group(1)) + '"',
                    xml,
                )
                data = xml.encode("utf-8")
            zout.writestr(item, data)
    return tmp


def _is_drm_file(path: Path) -> bool:
    """Detect NASCA DRM-encrypted xlsx by file header (<## NASC)."""
    try:
        with open(path, 'rb') as f:
            return f.read(8).startswith(b'<## NASC')
    except Exception:
        return False


def _extract_render(ws, excel_app, cap_row: int, cap_col: int) -> dict | None:
    """COM 세션 내에서 시트 렌더 데이터를 추출한다.

    전략: ws.Copy() → 임시 xlsx 저장 → openpyxl로 완전 렌더 읽기.
    Copy가 막히면(일부 DRM 시트) COM row-batch로 폴백.
    """
    import tempfile, os
    import openpyxl
    from openpyxl.utils import get_column_letter, range_boundaries

    used = ws.UsedRange
    max_row = used.Rows.Count if used else 1
    max_col = used.Columns.Count if used else 1

    # ── Primary: Copy → SaveAs → openpyxl (완전 충실, ~3s/sheet) ──
    try:
        ws.Copy()
        new_wb = excel_app.ActiveWorkbook
        tmp = os.path.join(tempfile.gettempdir(), f"_drm_render_{id(ws)}.xlsx")
        new_wb.SaveAs(tmp, 51)  # xlOpenXMLWorkbook
        new_wb.Close(SaveChanges=False)

        try:
            owb = openpyxl.load_workbook(tmp, data_only=True)
            ows = owb[owb.sheetnames[0]]
            render = _render_from_openpyxl(ows, cap_row, cap_col)
            owb.close()
            return render
        finally:
            try:
                os.unlink(tmp)
            except OSError:
                pass
    except Exception:
        pass  # Copy 실패 → 폴백

    # ── Fallback: COM row-batch (병합 제외, 값/스타일/열폭/행높이) ──
    return _extract_render_com_fallback(ws, used, max_row, max_col, cap_row, cap_col)


# ── 렌더 헬퍼 (openpyxl 스타일 → JSON 변환) ──
_EMU_PX = 9525.0
_BORDER_W = {"thin": 1, "hair": 1, "dotted": 1, "dashed": 1, "mediumDashed": 2,
             "medium": 2, "thick": 3, "double": 3}
_BORDER_STYLE = {"dotted": "dotted", "dashed": "dashed", "mediumDashed": "dashed",
                 "double": "double"}


def _render_rgb(color) -> str | None:
    if color is None or getattr(color, "type", None) != "rgb":
        return None
    v = color.rgb
    if not isinstance(v, str) or len(v) != 8 or v == "00000000":
        return None
    return "#" + v[-6:]


def _render_border_css(side) -> str | None:
    if side is None or side.style is None:
        return None
    w = _BORDER_W.get(side.style, 1)
    st = _BORDER_STYLE.get(side.style, "solid")
    return f"{w}px {st} {_render_rgb(side.color) or '#8d97a5'}"


def _render_fmt_value(v, fmt: str) -> tuple[str, bool]:
    """(표시 문자열, 숫자 여부) — Excel number_format 근사."""
    import datetime as _dt
    if v is None:
        return "", False
    if isinstance(v, bool):
        return str(v), False
    if isinstance(v, _dt.datetime):
        return (v.strftime("%Y-%m-%d") if (v.hour, v.minute, v.second) == (0, 0, 0)
                else v.strftime("%Y-%m-%d %H:%M")), False
    if isinstance(v, (_dt.date, _dt.time)):
        return str(v), False
    if isinstance(v, (int, float)):
        f = fmt or "General"
        dec = 0
        if "." in f:
            dec = sum(1 for ch in f.split(".", 1)[1] if ch in "0#")
        try:
            if "%" in f:
                return f"{v * 100:.{dec}f}%", True
            if "#,##" in f or "#,#" in f:
                return f"{v:,.{dec}f}", True
            if f != "General" and any(ch in f for ch in "0#"):
                return f"{v:.{dec}f}", True
        except (ValueError, OverflowError):
            pass
        if isinstance(v, float):
            s = f"{v:.6f}".rstrip("0").rstrip(".")
            return (s if s else "0"), True
        return str(v), True
    return str(v), False


def _render_from_openpyxl(ows, cap_row: int, cap_col: int) -> dict:
    """openpyxl 워크시트 → _render_sheet()와 동일한 렌더 JSON."""
    import base64
    from openpyxl.utils import get_column_letter, range_boundaries

    max_r = min(ows.max_row or 1, cap_row)
    max_c = min(ows.max_column or 1, cap_col)

    # 열 너비 / 행 높이
    col_px = []
    for c in range(1, max_c + 1):
        dim = ows.column_dimensions.get(get_column_letter(c))
        w = dim.width if dim is not None and dim.width else 8.43
        col_px.append(max(14, round(w * 7 + 5)))
    row_px = []
    for r in range(1, max_r + 1):
        dim = ows.row_dimensions.get(r)
        h = dim.height if dim is not None and dim.height else 15.0
        row_px.append(max(12, round(h * 4 / 3)))

    # 병합 셀
    spans: dict[str, tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for m in ows.merged_cells.ranges:
        a, b, c2, d = range_boundaries(str(m))
        spans[f"{b},{a}"] = (min(d, max_r) - b + 1, min(c2, max_c) - a + 1)
        for rr in range(b, min(d, max_r) + 1):
            for cc in range(a, min(c2, max_c) + 1):
                if (rr, cc) != (b, a):
                    covered.add((rr, cc))

    # 셀 데이터
    cells = []
    for row in ows.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
        for cell in row:
            key = (cell.row, cell.column)
            if key in covered:
                continue
            v, is_num = _render_fmt_value(cell.value, cell.number_format)
            fill = None
            if cell.fill is not None and cell.fill.patternType == "solid":
                fill = _render_rgb(cell.fill.fgColor)
            fnt = cell.font
            bd = {}
            for side, name in ((cell.border.top, "t"), (cell.border.right, "r"),
                               (cell.border.bottom, "b"), (cell.border.left, "l")):
                css = _render_border_css(side)
                if css:
                    bd[name] = css
            al = cell.alignment
            if v == "" and not fill and not bd and key not in \
                    {tuple(map(int, k.split(","))) for k in spans}:
                continue
            rs, cs = spans.get(f"{cell.row},{cell.column}", (1, 1))
            d: dict = {"r": cell.row, "c": cell.column, "v": v}
            if rs > 1 or cs > 1:
                d["rs"], d["cs"] = rs, cs
            if is_num:
                d["n"] = 1
            if fill:
                d["f"] = fill
            if fnt is not None:
                if fnt.bold:
                    d["b"] = 1
                if fnt.italic:
                    d["i"] = 1
                if fnt.size and abs(float(fnt.size) - 11.0) > 0.1:
                    d["sz"] = round(float(fnt.size) * 4 / 3)
                fc = _render_rgb(fnt.color)
                if fc and fc != "#000000":
                    d["fc"] = fc
            if bd:
                d["bd"] = bd
            if al is not None:
                if al.horizontal in ("center", "right", "left"):
                    d["ha"] = al.horizontal[0]
                if al.wrap_text:
                    d["wr"] = 1
            cells.append(d)

    # 이미지
    _EMU_PX = 9525.0
    cum_x = [0]
    for w in col_px:
        cum_x.append(cum_x[-1] + w)
    cum_y = [0]
    for h in row_px:
        cum_y.append(cum_y[-1] + h)
    images = []
    total_img = 0
    for img in getattr(ows, "_images", []):
        try:
            data = img._data()
        except Exception:
            continue
        if len(data) > 1_500_000 or total_img + len(data) > 4_000_000:
            continue
        frm = getattr(img.anchor, "_from", None)
        if frm is None:
            continue
        col0 = min(frm.col, max_c - 1)
        row0 = min(frm.row, max_r - 1)
        x = cum_x[col0] + frm.colOff / _EMU_PX
        y = cum_y[row0] + frm.rowOff / _EMU_PX
        to = getattr(img.anchor, "to", None)
        ext = getattr(img.anchor, "ext", None)
        if to is not None:
            w = cum_x[min(to.col, max_c)] + to.colOff / _EMU_PX - x
            h = cum_y[min(to.row, max_r)] + to.rowOff / _EMU_PX - y
        elif ext is not None:
            w, h = ext.cx / _EMU_PX, ext.cy / _EMU_PX
        else:
            w, h = 160, 120
        if w <= 4 or h <= 4:
            continue
        total_img += len(data)
        mt = (getattr(img, "format", None) or "png").lower()
        images.append({"x": round(x), "y": round(y), "w": round(w), "h": round(h),
                       "src": f"data:image/{mt};base64,{base64.b64encode(data).decode()}"})

    return {
        "max_row": max_r, "max_col": max_c,
        "cols": col_px, "rows": row_px,
        "cells": cells, "images": images,
        "gridlines": bool(ows.sheet_view.showGridLines),
        "truncated": (ows.max_row or 1) > cap_row or (ows.max_column or 1) > cap_col,
    }


def _extract_render_com_fallback(ws, used, max_row: int, max_col: int,
                                  cap_row: int, cap_col: int) -> dict:
    """Copy 불가 시트용 COM 폴백 — 값/열폭/행높이/bold/fill (병합 생략)."""
    # 열 너비
    col_px = []
    for c in range(1, cap_col + 1):
        try:
            w = ws.Columns(c).ColumnWidth
            col_px.append(max(14, round((w or 8.43) * 7 + 5)))
        except Exception:
            col_px.append(64)

    # 행 높이
    row_px = []
    for r in range(1, cap_row + 1):
        try:
            h = ws.Rows(r).RowHeight
            row_px.append(max(12, round((h or 15.0) * 4 / 3)))
        except Exception:
            row_px.append(20)

    # 값 bulk read
    cells = []
    if cap_row > 0 and cap_col > 0:
        try:
            data_range = ws.Range(ws.Cells(1, 1), ws.Cells(cap_row, cap_col))
            values = data_range.Value
        except Exception:
            values = None
        if values:
            for r_idx, row_data in enumerate(values, 1):
                if row_data is None: continue
                for c_idx, val in enumerate(row_data, 1):
                    if val is None: continue
                    v_str, is_num = _fmt_value_com(val)
                    if v_str == "": continue
                    d: dict = {"r": r_idx, "c": c_idx, "v": v_str}
                    if is_num: d["n"] = 1
                    cells.append(d)

    total_rows = used.Rows.Count if used else 1
    total_cols = used.Columns.Count if used else 1
    return {
        "max_row": cap_row, "max_col": cap_col,
        "cols": col_px, "rows": row_px,
        "cells": cells, "images": [],
        "gridlines": True,
        "truncated": total_rows > cap_row or total_cols > cap_col,
    }


def _fmt_value_com(v) -> tuple[str, bool]:
    """COM Value → (표시 문자열, 숫자 여부)."""
    import datetime as _dt
    if v is None:
        return "", False
    if isinstance(v, bool):
        return str(v), False
    if isinstance(v, _dt.datetime):
        return (v.strftime("%Y-%m-%d") if (v.hour, v.minute, v.second) == (0, 0, 0)
                else v.strftime("%Y-%m-%d %H:%M")), False
    if isinstance(v, (_dt.date, _dt.time)):
        return str(v), False
    if isinstance(v, (int, float)):
        if isinstance(v, float):
            s = f"{v:.6f}".rstrip("0").rstrip(".")
            return (s if s else "0"), True
        return str(v), True
    return str(v), False


def _inspect_via_com(path: Path, relative_to: Path | None,
                    parser_version: str) -> WorkbookStructure:
    """Read workbook via Excel COM automation (for DRM-encrypted files).

    원본 파일은 읽기만 하고 전혀 수정하지 않는다.
    COM이 열려있는 동안 렌더 데이터(병합/스타일/열폭/행높이)까지 추출해
    sheet_render 캐시에 저장한다 — 웹에서 파일 재오픈 없이 즉시 응답.
    """
    import win32com.client
    import pythoncom

    try:
        pythoncom.CoInitializeEx(pythoncom.COINIT_APARTMENTTHREADED)
    except Exception:
        pass

    excel = win32com.client.DispatchEx('Excel.Application')
    try:
        excel.Visible = False
    except Exception:
        pass
    try:
        excel.DisplayAlerts = False
    except Exception:
        pass

    try:
        wb = excel.Workbooks.Open(str(path.resolve()), 0, True)
        rel = str(path.relative_to(relative_to)) if relative_to else path.name
        stat = path.stat()
        file_hash = sha256_file(path)
        structure = WorkbookStructure(
            file_name=path.name,
            relative_path=rel,
            sha256=file_hash,
            file_size=stat.st_size,
            modified_time=stat.st_mtime,
            parser_version=parser_version,
        )

        # document_id 계산 (kg.store.stable_id와 동일 로직)
        import hashlib as _hl
        logical_path = rel if relative_to else str(path)
        document_id = _hl.sha256(logical_path.encode("utf-8")).hexdigest()[:16]

        # 렌더 캐시 저장용
        render_cache_list: list[tuple[str, str]] = []  # [(sheet_name, render_json)]

        for idx in range(1, wb.Sheets.Count + 1):
            ws = wb.Sheets(idx)
            sheet_name = ws.Name
            used = ws.UsedRange
            max_row = used.Rows.Count if used else 0
            max_col = used.Columns.Count if used else 0

            sheet = SheetStructure(
                sheet_name=sheet_name,
                sheet_index=idx - 1,
                max_row=max_row,
                max_col=max_col,
            )

            cap_row = min(max_row, 2000)
            cap_col = min(max_col, 100)
            render_cap_row = min(max_row, 300)  # _SHEET_CAP_ROWS
            render_cap_col = min(max_col, 40)   # _SHEET_CAP_COLS

            # ── Bulk read values ──
            values = None
            if cap_row > 0 and cap_col > 0:
                try:
                    data_range = ws.Range(ws.Cells(1, 1), ws.Cells(cap_row, cap_col))
                    values = data_range.Value
                except Exception:
                    pass

            if values:
                for r_idx, row_data in enumerate(values, 1):
                    if row_data is None:
                        continue
                    for c_idx, val in enumerate(row_data, 1):
                        if val is None:
                            continue
                        addr = f"{get_column_letter(c_idx)}{r_idx}"
                        info = CellInfo(
                            address=addr, row=r_idx, col=c_idx,
                            value=_jsonable_value(val),
                        )
                        sheet.cells.append(info)

            structure.sheets.append(sheet)

            # ── Extract render data (same COM session) ──
            try:
                render = _extract_render(ws, excel, render_cap_row, render_cap_col)
                if render:
                    render["sheet"] = sheet_name
                    render_cache_list.append((sheet_name, json.dumps(render, ensure_ascii=False)))
            except Exception:
                pass  # 렌더 실패해도 파서는 계속

        wb.Close(SaveChanges=False)

        # ── Save render cache to DB ──
        if render_cache_list:
            try:
                from kg.store import KgStore
                db_path = Path(path.parent)
                # 워크스페이스 DB 경로 탐색
                for candidate in [path.parent.parent / "data" / "kg" / "kg.db",
                                  path.parent / "data" / "kg" / "kg.db"]:
                    if candidate.exists():
                        db_path = candidate
                        break
                else:
                    # run_ingest.py에서 주입한 경로 사용
                    db_path = Path(os.environ.get("_KG_DB_PATH", ""))
                    if not db_path.exists():
                        db_path = None
                if db_path and db_path.exists():
                    store = KgStore(db_path)
                    for sname, rjson in render_cache_list:
                        store.save_render(document_id, sname, rjson, file_hash)
                    store.commit()
                    store.close()
            except Exception:
                pass  # DB 저장 실패해도 파서는 계속

        return structure
    finally:
        try:
            excel.Quit()
        except Exception:
            pass


class WorkbookInspector:
    """Extracts a WorkbookStructure from one xlsx file (all sheets)."""

    def __init__(self, parser_version: str = PARSER_VERSION):
        self.parser_version = parser_version

    def inspect(self, path: Path, relative_to: Path | None = None) -> WorkbookStructure:
        path = Path(path)

        # DRM-encrypted files must be read via Excel COM
        if _is_drm_file(path):
            return _inspect_via_com(path, relative_to, self.parser_version)

        try:
            wb_f = openpyxl.load_workbook(path, data_only=False)
            wb_v = openpyxl.load_workbook(path, data_only=True)
        except ValueError:
            # 시트명에 금지 문자('210?' 등)가 있으면 openpyxl이 로드를 거부한다.
            # 원본은 불변(§10) — 시트명만 정화한 임시 사본으로 재시도한다.
            repaired = _repair_sheet_names(path)
            try:
                wb_f = openpyxl.load_workbook(repaired, data_only=False)
                wb_v = openpyxl.load_workbook(repaired, data_only=True)
            finally:
                repaired.unlink(missing_ok=True)

        rel = str(path.relative_to(relative_to)) if relative_to else path.name
        stat = path.stat()
        structure = WorkbookStructure(
            file_name=path.name,
            relative_path=rel,
            sha256=sha256_file(path),
            file_size=stat.st_size,
            modified_time=stat.st_mtime,
            parser_version=self.parser_version,
        )

        for idx, sheet_name in enumerate(wb_f.sheetnames):
            ws = wb_f[sheet_name]
            wsv = wb_v[sheet_name]

            merged = [str(m) for m in ws.merged_cells.ranges]
            covered: dict[str, str] = {}
            masters: dict[str, str] = {}
            for m in ws.merged_cells.ranges:
                rng = str(m)
                for row in range(m.min_row, m.max_row + 1):
                    for col in range(m.min_col, m.max_col + 1):
                        addr = f"{get_column_letter(col)}{row}"
                        if row == m.min_row and col == m.min_col:
                            masters[addr] = rng
                        else:
                            covered[addr] = rng

            sheet = SheetStructure(
                sheet_name=sheet_name,
                sheet_index=idx,
                max_row=ws.max_row or 0,
                max_col=ws.max_column or 0,
            )
            sheet.merged_ranges = sorted(merged)

            for row in ws.iter_rows():
                for cell in row:
                    fill = _fill_rgb(cell)
                    if cell.value is None and fill is None and cell.coordinate not in masters:
                        continue
                    is_formula = isinstance(cell.value, str) and cell.value.startswith("=")
                    cached = None
                    if is_formula:
                        cached = _jsonable_value(wsv[cell.coordinate].value)
                    info = CellInfo(
                        address=cell.coordinate,
                        row=cell.row,
                        col=cell.column,
                        value=_jsonable_value(cell.value),
                        cached_value=cached,
                        is_formula=is_formula,
                        formula=cell.value if is_formula else None,
                        formula_refs=extract_formula_refs(cell.value) if is_formula else [],
                        fill_rgb=fill,
                        bold=bool(cell.font and cell.font.bold),
                        number_format=cell.number_format if cell.number_format != "General" else None,
                        merged_range=masters.get(cell.coordinate),
                        merged_into=covered.get(cell.coordinate),
                    )
                    sheet.cells.append(info)

            for img in getattr(ws, "_images", []):
                try:
                    data = img._data()
                except Exception:
                    continue
                anchor = img.anchor
                frm = getattr(anchor, "_from", None)
                a_row = frm.row if frm is not None else 0
                a_col = frm.col if frm is not None else 0
                sheet.images.append(
                    ImageInfo(
                        image_hash=sha256_bytes(data),
                        anchor_row=a_row,
                        anchor_col=a_col,
                        media_path=getattr(img, "path", ""),
                        ext=getattr(img, "format", "png") or "png",
                    )
                )

            structure.sheets.append(sheet)
        return structure
