"""CanonicalWorkbookExporter — 고정 5-sheet 출력 계약 (설계문서 §7).

원본 양식이 무엇이든 동일한 시트/컬럼 계약을 유지한다. 저장 스키마(long)와
분석 스키마(wide)는 분리 — wide는 build_wide_view()가 concept pivot으로 만든다.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

from src.loader.versioned_loader import VersionedLoader

HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
HEADER_FONT = Font(bold=True)

SHEETS = {
    "01_Record_Index": ["record_key", "record_type", "business_key", "event_time",
                        "overall_status", "note", "source_sheet", "version", "semantic_hash"],
    "02_Observations": ["record_key", "concept_id", "raw_label", "value", "unit",
                        "value_role", "status_code", "row_key"],
    "03_Source_Lineage": ["record_key", "observation_key", "source_sheet", "source_address",
                          "raw_label", "header_path", "raw_value", "raw_unit",
                          "source_document_version_id"],
    "04_Attachments": ["record_key", "image_hash", "source_anchor", "uri"],
    "05_Mapping_Log": ["raw_label", "context", "concept_id", "confidence", "decision",
                       "approved_by", "mapping_version"],
}


class CanonicalWorkbookExporter:
    def __init__(self, loader: VersionedLoader):
        self.loader = loader

    def export(self, out_path: Path) -> Path:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        conn = self.loader.conn

        ws = self._sheet(wb, "01_Record_Index")
        for r in self.loader.current_records():
            ws.append([r["record_key"], r["record_type"], r["business_key"], r["event_time"],
                       r["overall_status"], r["note"], r["source_sheet"], r["version"],
                       r["semantic_hash"][:12]])

        ws = self._sheet(wb, "02_Observations")
        for o in self.loader.current_observations():
            value = o["normalized_value_num"] if o["normalized_value_num"] is not None else o["normalized_value_text"]
            ws.append([o["record_key"], o["concept_id"], o["raw_label"], value,
                       o["canonical_unit"], o["value_role"], o["status_code"], o["row_key"]])

        ws = self._sheet(wb, "03_Source_Lineage")
        for o in self.loader.current_observations():
            raw = o["raw_value_num"] if o["raw_value_num"] is not None else o["raw_value_text"]
            ws.append([o["record_key"], o["observation_key"], o["source_sheet"], o["source_address"],
                       o["raw_label"], o["header_path"], raw, o["raw_unit"],
                       o["source_document_version_id"]])

        ws = self._sheet(wb, "04_Attachments")
        for a in conn.execute("SELECT * FROM attachment WHERE is_current=1 ORDER BY record_key"):
            ws.append([a["record_key"], a["image_hash"], a["source_anchor"], a["uri"]])

        ws = self._sheet(wb, "05_Mapping_Log")
        for m in conn.execute("SELECT * FROM mapping_decision ORDER BY field_signature"):
            ws.append([m["raw_label"], m["context"], m["concept_id"], m["confidence"],
                       m["decision"], m["approved_by"], m["mapping_version"]])

        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        return out_path

    def _sheet(self, wb, name: str):
        ws = wb.create_sheet(name)
        ws.append(SHEETS[name])
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        ws.freeze_panes = "A2"
        return ws

    # ------------------------------------------------------------ analysis ----
    def build_wide_view(self, concept_ids: list[str]) -> list[dict]:
        """선택 concept만 pivot한 분석용 wide dataset (§6.1, §11.3)."""
        rows: dict[str, dict] = {}
        for o in self.loader.current_observations():
            if o["concept_id"] not in concept_ids:
                continue
            rec = rows.setdefault(o["record_key"], {"record_key": o["record_key"]})
            value = o["normalized_value_num"] if o["normalized_value_num"] is not None else o["normalized_value_text"]
            col = o["concept_id"] if not o["row_key"] else f'{o["concept_id"]}[{o["row_key"]}]'
            rec[col] = value
        return list(rows.values())
