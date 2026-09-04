"""Immutable XLSX registration and replaceable read-only Viewer backend.

The original workbook is the source of truth. LibreOffice only reads it and
writes a disposable PDF into a SHA-256 keyed cache. Source locations remain
logical Excel coordinates and never contain renderer pixel geometry.
"""
from __future__ import annotations

import hashlib
import json
import os
import shutil
import stat
import subprocess
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from kg.store import KgStore, now_iso

ALLOWED_EXTENSION = ".xlsx"
MIN_XLSX_BYTES = 100


class ViewerError(ValueError):
    """Safe, user-facing Viewer validation or rendering error."""


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def validate_unlocked_xlsx(path: Path) -> dict:
    """Validate actual workbook readability, not an unlock service response."""
    path = Path(path)
    if not path.is_file():
        raise ViewerError("FILE_MISSING: unlocked file does not exist")
    if path.suffix.lower() != ALLOWED_EXTENSION:
        raise ViewerError("UNSUPPORTED_FORMAT: only .xlsx is supported")
    if path.stat().st_size < MIN_XLSX_BYTES:
        raise ViewerError("FILE_CORRUPT: workbook is unexpectedly small")
    if not zipfile.is_zipfile(path):
        raise ViewerError("DRM_PROTECTED: file is not a readable XLSX container")
    try:
        workbook = load_workbook(path, read_only=False, data_only=False)
        if not workbook.sheetnames:
            raise ViewerError("FILE_CORRUPT: workbook has no sheets")
        sheets = []
        for index, sheet in enumerate(workbook.worksheets):
            images = []
            for image_index, image in enumerate(getattr(sheet, "_images", [])):
                anchor = getattr(image, "anchor", None)
                marker = getattr(anchor, "_from", None)
                images.append({
                    "image_id": f"image-{image_index + 1}",
                    "anchor": ({"row": marker.row + 1, "col": marker.col + 1}
                               if marker is not None else None),
                })
            sheets.append({
                "sheet_index": index,
                "sheet_name": sheet.title,
                "state": sheet.sheet_state,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "merged_ranges": [str(value) for value in sheet.merged_cells.ranges],
                "images": images,
            })
        workbook.close()
    except ViewerError:
        raise
    except Exception as exc:
        raise ViewerError(f"WORKBOOK_OPEN_FAILED: {exc}") from exc
    return {"sha256": sha256_file(path), "size": path.stat().st_size, "sheets": sheets}


def _safe_source(unlock_staging_root: Path, candidate: Path) -> Path:
    root = Path(unlock_staging_root).resolve()
    source = Path(candidate).resolve()
    try:
        source.relative_to(root)
    except ValueError as exc:
        raise ViewerError("PATH_FORBIDDEN: source must be inside unlock staging root") from exc
    if source.is_symlink():
        raise ViewerError("PATH_FORBIDDEN: symlink sources are not allowed")
    return source


def register_unlocked(store: KgStore, document_id: str, document_version: str,
                      source: Path, unlock_staging_root: Path,
                      immutable_root: Path) -> dict:
    """Validate then copy an authorized unlock result into immutable storage."""
    docver = store.conn.execute(
        "SELECT document_id,file_hash FROM document_version WHERE version_id=?",
        (document_version,)).fetchone()
    if docver is None or docver["document_id"] != document_id:
        raise ViewerError("UNKNOWN_VERSION: document version does not belong to document")
    source = _safe_source(unlock_staging_root, source)
    metadata = validate_unlocked_xlsx(source)
    destination_dir = Path(immutable_root).resolve() / document_id
    destination_dir.mkdir(parents=True, exist_ok=True)
    destination = destination_dir / f"{metadata['sha256']}.xlsx"
    if not destination.exists():
        temporary = destination.with_suffix(".tmp")
        shutil.copyfile(source, temporary)
        if sha256_file(temporary) != metadata["sha256"]:
            temporary.unlink(missing_ok=True)
            raise ViewerError("COPY_FAILED: immutable copy checksum mismatch")
        os.replace(temporary, destination)
        destination.chmod(stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH)
    elif sha256_file(destination) != metadata["sha256"]:
        raise ViewerError("IMMUTABLE_SOURCE_CORRUPT: stored checksum mismatch")

    now = now_iso()
    store.conn.execute(
        """INSERT INTO viewer_document_version
             (document_id,document_version,sha256,unlocked_path,drm_status,
              drm_error,render_status,render_error,sheet_count,registered_at,rendered_at)
           VALUES (?,?,?,?,? ,NULL,'PENDING',NULL,?,?,NULL)
           ON CONFLICT(document_id,document_version) DO UPDATE SET
             sha256=excluded.sha256,unlocked_path=excluded.unlocked_path,
             drm_status='READY',drm_error=NULL,sheet_count=excluded.sheet_count""",
        (document_id, document_version, metadata["sha256"], str(destination),
         "READY", len(metadata["sheets"]), now))
    store.conn.execute(
        "DELETE FROM viewer_sheet WHERE document_id=? AND document_version=?",
        (document_id, document_version))
    for sheet in metadata["sheets"]:
        store.conn.execute(
            "INSERT INTO viewer_sheet VALUES (?,?,?,?,?,?,?,?,?)",
            (document_id, document_version, sheet["sheet_index"], sheet["sheet_name"],
             sheet["state"], sheet["max_row"], sheet["max_column"],
             json.dumps(sheet["merged_ranges"], ensure_ascii=False),
             json.dumps(sheet["images"], ensure_ascii=False)))
    return document_metadata(store, document_id, document_version)


def mark_validation_failed(store: KgStore, document_id: str, document_version: str,
                           error: str) -> None:
    """Persist unlock validation failure without touching parsing or KG data."""
    docver = store.conn.execute(
        "SELECT document_id FROM document_version WHERE version_id=?",
        (document_version,)).fetchone()
    if docver is None or docver["document_id"] != document_id:
        return
    store.conn.execute(
        """INSERT INTO viewer_document_version
             (document_id,document_version,sha256,unlocked_path,drm_status,drm_error,
              render_status,render_error,sheet_count,registered_at,rendered_at)
           VALUES (?,?,NULL,NULL,'FAILED',?,'PENDING',NULL,NULL,?,NULL)
           ON CONFLICT(document_id,document_version) DO UPDATE SET
             drm_status='FAILED',drm_error=excluded.drm_error""",
        (document_id, document_version, error[:1000], now_iso()))


def _viewer_row(store: KgStore, document_id: str, document_version: str):
    row = store.conn.execute(
        """SELECT v.*,d.filename,dv.parsed_at FROM viewer_document_version v
             JOIN document d ON d.document_id=v.document_id
             JOIN document_version dv ON dv.version_id=v.document_version
            WHERE v.document_id=? AND v.document_version=?""",
        (document_id, document_version)).fetchone()
    if row is None:
        raise ViewerError("VIEWER_SOURCE_NOT_REGISTERED")
    return row


def document_metadata(store: KgStore, document_id: str, document_version: str) -> dict:
    row = _viewer_row(store, document_id, document_version)
    return {key: row[key] for key in (
        "document_id", "document_version", "filename", "sha256", "drm_status",
        "drm_error", "render_status", "render_error", "sheet_count",
        "registered_at", "rendered_at", "parsed_at")}


def sheets(store: KgStore, document_id: str, document_version: str,
           include_hidden: bool = True) -> list[dict]:
    _viewer_row(store, document_id, document_version)
    query = "SELECT * FROM viewer_sheet WHERE document_id=? AND document_version=?"
    params = [document_id, document_version]
    if not include_hidden:
        query += " AND state='visible'"
    rows = store.conn.execute(query + " ORDER BY sheet_index", params).fetchall()
    output = []
    for row in rows:
        item = dict(row)
        item["merged_ranges"] = json.loads(item.pop("merged_ranges_json"))
        item["images"] = json.loads(item.pop("images_json"))
        output.append(item)
    return output


def cache_paths(cache_root: Path, digest: str) -> tuple[Path, Path, Path]:
    directory = Path(cache_root).resolve() / digest
    return directory / "workbook.pdf", directory / "sheets.json", directory / "render_metadata.json"


def prepare_render(store: KgStore, document_id: str, document_version: str,
                   cache_root: Path) -> dict:
    """DB stage 1 (call under the shared lock): validate and mark RUNNING.

    Returns everything the LibreOffice stage needs so that stage can run
    without touching the shared connection.
    """
    row = _viewer_row(store, document_id, document_version)
    if row["drm_status"] != "READY":
        raise ViewerError("DRM_NOT_READY")
    source = Path(row["unlocked_path"])
    digest = row["sha256"]
    pdf, sheet_json, render_json = cache_paths(cache_root, digest)
    if pdf.is_file() and pdf.stat().st_size > 0:
        return {"cached": True, "pdf": pdf}
    store.conn.execute(
        "UPDATE viewer_document_version SET render_status='RUNNING',render_error=NULL WHERE document_id=? AND document_version=?",
        (document_id, document_version))
    return {"cached": False, "source": source, "digest": digest, "pdf": pdf,
            "sheet_json": sheet_json, "render_json": render_json}


def execute_render(prep: dict, soffice: str | None = None) -> None:
    """Renderer stage (never hold the shared DB lock here): the LibreOffice
    subprocess can take up to two minutes and must not stall the web app.
    Verifies the immutable source hash before and after conversion."""
    source, digest, pdf = prep["source"], prep["digest"], prep["pdf"]
    before = sha256_file(source)
    if before != digest:
        raise ViewerError("IMMUTABLE_SOURCE_CORRUPT")
    executable = soffice or shutil.which("libreoffice") or shutil.which("soffice")
    if executable is None:
        raise ViewerError("RENDERER_UNAVAILABLE: LibreOffice is not installed")
    pdf.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="viewer-lo-") as profile:
        command = [executable, "--headless", "--nologo", "--nodefault", "--nolockcheck",
                   f"-env:UserInstallation=file://{profile}", "--convert-to", "pdf",
                   "--outdir", str(pdf.parent), str(source)]
        completed = subprocess.run(command, capture_output=True, text=True, timeout=120)
    generated = pdf.parent / f"{source.stem}.pdf"
    if completed.returncode != 0 or not generated.is_file():
        error = (completed.stderr or completed.stdout or "LibreOffice did not create a PDF").strip()
        raise ViewerError(f"RENDER_FAILED: {error}")
    if generated != pdf:
        os.replace(generated, pdf)
    after = sha256_file(source)
    if after != before:
        pdf.unlink(missing_ok=True)
        raise ViewerError("IMMUTABILITY_VIOLATION: source changed during rendering")


def mark_render_failed(store: KgStore, document_id: str, document_version: str,
                       error: str) -> None:
    """DB stage (under lock): persist a renderer failure."""
    store.conn.execute(
        "UPDATE viewer_document_version SET render_status='FAILED',render_error=? WHERE document_id=? AND document_version=?",
        (error[:1000], document_id, document_version))


def finalize_render(store: KgStore, document_id: str, document_version: str,
                    prep: dict) -> None:
    """DB stage 2 (under lock): cache metadata + SUCCESS status."""
    sheet_data = sheets(store, document_id, document_version)
    prep["sheet_json"].write_text(
        json.dumps(sheet_data, ensure_ascii=False, indent=2), encoding="utf-8")
    prep["render_json"].write_text(
        json.dumps({"sha256": prep["digest"], "engine": "libreoffice-pdf",
                    "read_only": True}, indent=2), encoding="utf-8")
    store.conn.execute(
        "UPDATE viewer_document_version SET render_status='SUCCESS',render_error=NULL,rendered_at=? WHERE document_id=? AND document_version=?",
        (now_iso(), document_id, document_version))


def render_preview(store: KgStore, document_id: str, document_version: str,
                   cache_root: Path, soffice: str | None = None) -> Path:
    """Single-caller convenience (tests/CLI): all three stages in sequence."""
    prep = prepare_render(store, document_id, document_version, cache_root)
    if prep["cached"]:
        return prep["pdf"]
    try:
        execute_render(prep, soffice)
    except ViewerError as exc:
        mark_render_failed(store, document_id, document_version, str(exc))
        raise
    finalize_render(store, document_id, document_version, prep)
    return prep["pdf"]


def normalize_a1_range(value: str) -> dict:
    """Return durable logical coordinates; reject sheet/path/pixel input."""
    if not value or "!" in value or "/" in value or "\\" in value:
        raise ViewerError("INVALID_RANGE: expected an A1 cell range")
    try:
        min_col, min_row, max_col, max_row = range_boundaries(value.upper())
    except (ValueError, TypeError) as exc:
        raise ViewerError("INVALID_RANGE: expected an A1 cell range") from exc
    from openpyxl.utils import get_column_letter
    a1 = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    return {"start_row": min_row, "start_col": min_col, "end_row": max_row,
            "end_col": max_col, "a1_range": a1}


def source_locator(store: KgStore, document_id: str, document_version: str,
                   sheet: str, a1_range: str, concept_id: str | None = None) -> dict:
    metadata = document_metadata(store, document_id, document_version)
    available = {item["sheet_name"] for item in sheets(store, document_id, document_version)}
    if sheet not in available:
        raise ViewerError("UNKNOWN_SHEET")
    coordinates = normalize_a1_range(a1_range)
    result = {"document_id": document_id, "document_version": document_version,
              "sheet": sheet, "range": {k: coordinates[k] for k in (
                  "start_row", "start_col", "end_row", "end_col")},
              "a1_range": coordinates["a1_range"], "concept_id": concept_id,
              "sha256": metadata["sha256"]}
    rows = store.conn.execute(
        "SELECT template_id,template_version,status FROM document_template_assignment WHERE document_id=? AND document_version=?",
        (document_id, document_version)).fetchall()
    # N:M — 배정 전체를 내려주되, 이 위치와 일치하는 매핑의 템플릿을 대표로 둔다
    result["parsing_templates"] = [dict(r) for r in rows]
    result["parsing_template"] = dict(rows[0]) if rows else None
    result["template_source"] = None
    result["effective_source"] = None
    result["mapping_source"] = None
    if rows:
        from kg.parsing import effective_mappings
        candidates = effective_mappings(store, document_version)
        mapping = next((item for item in candidates
                        if (concept_id is None or item["concept_id"] == concept_id)
                        and (item["effective_source"].get("sheet", sheet) == sheet)
                        and item["effective_source"].get("range") == coordinates["a1_range"]), None)
        if mapping:
            result["parsing_template"] = next(
                (dict(r) for r in rows if r["template_id"] == mapping["template_id"]),
                result["parsing_template"])
            result["template_source"] = mapping["template_source"]
            result["effective_source"] = mapping["effective_source"]
            result["mapping_source"] = mapping["mapping_source"]
    return result
