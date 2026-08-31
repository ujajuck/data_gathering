"""Versioned, declarative Excel parsing templates and document deltas.

This module deliberately returns extracted values and provenance only.  It does
not create KG nodes and it never lets hooks mutate the database.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from kg.store import KgStore, new_id, now_iso

LIFECYCLES = {"DRAFT", "ACTIVE", "DEPRECATED", "ARCHIVED"}
ASSIGNMENT_STATES = {"ASSIGNED", "PARSED", "REVIEW_REQUIRED", "OVERRIDDEN", "FAILED"}


class ParsingError(ValueError):
    """A template request is invalid or cannot be resolved."""


def _dump(value) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _load(value: str | None, default=None):
    return default if value is None else json.loads(value)


def _validate_spec(spec: dict) -> list[dict]:
    sheets = spec.get("sheet_templates")
    if not isinstance(sheets, list) or not sheets:
        raise ParsingError("sheet_templates must be a non-empty list")
    seen_sheets: set[str] = set()
    for sheet in sheets:
        name = sheet.get("name")
        if not isinstance(name, str) or not name or name in seen_sheets:
            raise ParsingError("sheet template names must be unique non-empty strings")
        seen_sheets.add(name)
        matcher = sheet.get("match", {})
        if not isinstance(matcher, dict) or not any(
                k in matcher for k in ("names", "name_regex", "headers")):
            raise ParsingError(f"sheet template {name} needs a matcher")
        if "name_regex" in matcher:
            try:
                re.compile(matcher["name_regex"])
            except (re.error, TypeError) as exc:
                raise ParsingError(f"invalid name_regex in {name}: {exc}") from exc
        mappings = sheet.get("mappings", [])
        if not isinstance(mappings, list):
            raise ParsingError(f"mappings in {name} must be a list")
        keys: set[str] = set()
        for mapping in mappings:
            key = mapping.get("key")
            if not isinstance(key, str) or not key or key in keys:
                raise ParsingError(f"mapping keys in {name} must be unique")
            keys.add(key)
            source = mapping.get("source")
            if not isinstance(source, dict) or not (
                    source.get("range") or source.get("key_search")):
                raise ParsingError(f"mapping {key} needs source.range or source.key_search")
    return sheets


def create_template(store: KgStore, template_id: str, name: str,
                    target_document_kg: str | None, lifecycle: str = "DRAFT") -> dict:
    if lifecycle not in LIFECYCLES:
        raise ParsingError("invalid template lifecycle")
    store.conn.execute(
        "INSERT INTO parsing_template VALUES (?,?,?,?,?)",
        (template_id, name, target_document_kg, lifecycle, now_iso()))
    return template_detail(store, template_id)


def add_version(store: KgStore, template_id: str, spec: dict,
                created_by: str | None = None) -> dict:
    sheets = _validate_spec(spec)
    template = store.conn.execute(
        "SELECT 1 FROM parsing_template WHERE template_id=?", (template_id,)).fetchone()
    if template is None:
        raise ParsingError("unknown template")
    version = store.conn.execute(
        "SELECT COALESCE(MAX(version),0)+1 FROM parsing_template_version WHERE template_id=?",
        (template_id,)).fetchone()[0]
    store.conn.execute(
        "INSERT INTO parsing_template_version VALUES (?,?,?,?,?)",
        (template_id, version, _dump(spec), now_iso(), created_by))
    for ordinal, sheet in enumerate(sheets):
        sid = new_id("SHT")
        store.conn.execute(
            "INSERT INTO sheet_template VALUES (?,?,?,?,?,?)",
            (sid, template_id, version, sheet["name"], _dump(sheet["match"]), ordinal))
        for mapping in sheet.get("mappings", []):
            store.conn.execute(
                "INSERT INTO template_mapping VALUES (?,?,?,?,?,?,?,?,?)",
                (new_id("TMAP"), sid, mapping["key"], mapping.get("concept_id"),
                 mapping.get("document_kg_node"), _dump(mapping["source"]),
                 mapping.get("type"), mapping.get("unit"),
                 _dump(mapping.get("normalization", {}))))
    return version_detail(store, template_id, version)


def template_detail(store: KgStore, template_id: str) -> dict:
    row = store.conn.execute(
        "SELECT * FROM parsing_template WHERE template_id=?", (template_id,)).fetchone()
    if row is None:
        raise ParsingError("unknown template")
    out = dict(row)
    out["versions"] = [r[0] for r in store.conn.execute(
        "SELECT version FROM parsing_template_version WHERE template_id=? ORDER BY version",
        (template_id,))]
    out["current_version"] = out["versions"][-1] if out["versions"] else None
    return out


def version_detail(store: KgStore, template_id: str, version: int) -> dict:
    row = store.conn.execute(
        "SELECT * FROM parsing_template_version WHERE template_id=? AND version=?",
        (template_id, version)).fetchone()
    if row is None:
        raise ParsingError("unknown template version")
    out = dict(row)
    out["spec"] = _load(out.pop("spec_json"))
    out["sheet_templates"] = []
    sheets = store.conn.execute(
        "SELECT * FROM sheet_template WHERE template_id=? AND template_version=? ORDER BY ordinal",
        (template_id, version)).fetchall()
    for sheet in sheets:
        item = dict(sheet)
        item["match"] = _load(item.pop("matcher_json"))
        item["mappings"] = []
        for mapping in store.conn.execute(
                "SELECT * FROM template_mapping WHERE sheet_template_id=? ORDER BY mapping_key",
                (sheet["sheet_template_id"],)):
            m = dict(mapping)
            m["source"] = _load(m.pop("source_json"))
            m["normalization"] = _load(m.pop("normalization_json"), {})
            item["mappings"].append(m)
        out["sheet_templates"].append(item)
    return out


def assign(store: KgStore, document_id: str, document_version: str,
           template_id: str, template_version: int) -> dict:
    docver = store.conn.execute(
        "SELECT document_id FROM document_version WHERE version_id=?", (document_version,)).fetchone()
    if docver is None or docver["document_id"] != document_id:
        raise ParsingError("document version does not belong to document")
    version_detail(store, template_id, template_version)
    previous = store.conn.execute(
        "SELECT template_id,template_version FROM document_template_assignment WHERE document_version=?",
        (document_version,)).fetchone()
    store.conn.execute(
        """INSERT INTO document_template_assignment VALUES (?,?,?,?,?,?)
           ON CONFLICT(document_id,document_version) DO UPDATE SET
             template_id=excluded.template_id, template_version=excluded.template_version,
             status='ASSIGNED', assigned_at=excluded.assigned_at""",
        (document_id, document_version, template_id, template_version, "ASSIGNED", now_iso()))
    if previous and (previous["template_id"], previous["template_version"]) != \
            (template_id, template_version):
        _audit_overrides(store, document_version, template_id, template_version)
    return dict(store.conn.execute(
        "SELECT * FROM document_template_assignment WHERE document_version=?",
        (document_version,)).fetchone())


def _audit_overrides(store: KgStore, document_version: str,
                     template_id: str, template_version: int) -> list[dict]:
    """Classify old-version deltas by stable sheet-template/mapping keys.

    A conflicting manual source remains effective until a user explicitly
    accepts the new template. It is never silently discarded on reassignment.
    """
    rows = store.conn.execute(
        """SELECT o.override_id,o.override_source_json,old_st.name sheet_name,
                  old_tm.mapping_key,new_tm.source_json new_source
             FROM document_override o
             JOIN template_mapping old_tm ON old_tm.mapping_id=o.template_mapping_id
             JOIN sheet_template old_st ON old_st.sheet_template_id=old_tm.sheet_template_id
             LEFT JOIN sheet_template new_st ON new_st.template_id=?
               AND new_st.template_version=? AND new_st.name=old_st.name
             LEFT JOIN template_mapping new_tm ON new_tm.sheet_template_id=new_st.sheet_template_id
               AND new_tm.mapping_key=old_tm.mapping_key
            WHERE o.document_version=? AND o.status IN ('APPROVED','CONFLICT','REDUNDANT')""",
        (template_id, template_version, document_version)).fetchall()
    audits = []
    conflicts = False
    for row in rows:
        override_source = _load(row["override_source_json"])
        new_source = _load(row["new_source"], None)
        status = "REDUNDANT" if new_source == override_source else "CONFLICT"
        conflicts |= status == "CONFLICT"
        store.conn.execute(
            "UPDATE document_override SET status=?,updated_at=? WHERE override_id=?",
            (status, now_iso(), row["override_id"]))
        audits.append({"override_id": row["override_id"], "sheet_template": row["sheet_name"],
                       "mapping_key": row["mapping_key"], "status": status,
                       "template_source": new_source, "manual_source": override_source})
    if conflicts:
        store.conn.execute(
            "UPDATE document_template_assignment SET status='REVIEW_REQUIRED' WHERE document_version=?",
            (document_version,))
    return audits


def save_override(store: KgStore, document_id: str, document_version: str,
                  template_mapping_id: str, source: dict, reason: str | None,
                  created_by: str | None, status: str = "APPROVED") -> dict:
    assignment = store.conn.execute(
        "SELECT * FROM document_template_assignment WHERE document_id=? AND document_version=?",
        (document_id, document_version)).fetchone()
    mapping = store.conn.execute(
        """SELECT tm.* FROM template_mapping tm JOIN sheet_template st
             ON st.sheet_template_id=tm.sheet_template_id
           WHERE tm.mapping_id=? AND st.template_id=? AND st.template_version=?""",
        (template_mapping_id, assignment["template_id"] if assignment else "",
         assignment["template_version"] if assignment else -1)).fetchone()
    if assignment is None or mapping is None:
        raise ParsingError("override mapping is not part of the assigned template version")
    if not isinstance(source, dict) or not source.get("range"):
        raise ParsingError("override_source.range is required")
    existing = store.conn.execute(
        "SELECT override_id,created_at FROM document_override WHERE document_version=? AND template_mapping_id=?",
        (document_version, template_mapping_id)).fetchone()
    oid = existing["override_id"] if existing else new_id("OVR")
    created = existing["created_at"] if existing else now_iso()
    store.conn.execute(
        """INSERT INTO document_override VALUES (?,?,?,?,?,?,?,?,?,?)
           ON CONFLICT(document_version,template_mapping_id) DO UPDATE SET
             override_source_json=excluded.override_source_json,status=excluded.status,
             reason=excluded.reason,created_by=excluded.created_by,updated_at=excluded.updated_at""",
        (oid, document_id, document_version, template_mapping_id, _dump(source), status,
         reason, created_by, created, now_iso()))
    store.conn.execute(
        "UPDATE document_template_assignment SET status='OVERRIDDEN' WHERE document_version=?",
        (document_version,))
    return override_detail(store, oid)


def override_detail(store: KgStore, override_id: str) -> dict:
    row = store.conn.execute(
        "SELECT * FROM document_override WHERE override_id=?", (override_id,)).fetchone()
    if row is None:
        raise ParsingError("unknown override")
    out = dict(row)
    out["override_source"] = _load(out.pop("override_source_json"))
    return out


def effective_mappings(store: KgStore, document_version: str) -> list[dict]:
    assignment = store.conn.execute(
        "SELECT * FROM document_template_assignment WHERE document_version=?", (document_version,)).fetchone()
    if assignment is None:
        raise ParsingError("document version has no template assignment")
    rows = store.conn.execute(
        """SELECT tm.*,st.name sheet_template,st.matcher_json,o.override_id,
                  o.override_source_json,o.status override_status,o.reason override_reason
             FROM sheet_template st JOIN template_mapping tm
               ON tm.sheet_template_id=st.sheet_template_id
             LEFT JOIN (
               SELECT * FROM (
                 SELECT o.*,old_tm.mapping_key,old_st.name sheet_name,
                        row_number() OVER (
                          PARTITION BY old_st.name,old_tm.mapping_key
                          ORDER BY o.updated_at DESC,o.override_id DESC) override_rank
                   FROM document_override o
                   JOIN template_mapping old_tm ON old_tm.mapping_id=o.template_mapping_id
                   JOIN sheet_template old_st ON old_st.sheet_template_id=old_tm.sheet_template_id
                  WHERE o.document_version=? AND o.status IN ('APPROVED','CONFLICT')
               ) WHERE override_rank=1
             ) o ON o.mapping_key=tm.mapping_key AND o.sheet_name=st.name
            WHERE st.template_id=? AND st.template_version=?
            ORDER BY st.ordinal,tm.mapping_key""",
        (document_version, assignment["template_id"], assignment["template_version"])).fetchall()
    result = []
    for row in rows:
        item = dict(row)
        item["matcher"] = _load(item.pop("matcher_json"))
        item["template_source"] = _load(item.pop("source_json"))
        override = _load(item.pop("override_source_json"), None)
        item["effective_source"] = override or item["template_source"]
        item["mapping_source"] = "MANUAL" if override else "TEMPLATE"
        item["normalization"] = _load(item.pop("normalization_json"), {})
        result.append(item)
    return result


def _route_sheet(workbook, matcher: dict) -> list:
    names = set(matcher.get("names", []))
    regex = matcher.get("name_regex")
    headers = {str(v).strip() for v in matcher.get("headers", [])}
    matched = []
    for sheet in workbook.worksheets:
        name_ok = sheet.title in names or (regex and re.search(regex, sheet.title))
        header_ok = False
        if headers:
            observed = {str(cell.value).strip() for row in sheet.iter_rows(
                min_row=1, max_row=min(sheet.max_row, 30),
                min_col=1, max_col=min(sheet.max_column, 30)) for cell in row
                if cell.value is not None}
            header_ok = bool(headers & observed)
        if name_ok or header_ok:
            matched.append(sheet)
    return matched


def _locate_key(sheet, source: dict) -> str | None:
    terms = {str(v).strip() for v in source.get("key_search", [])}
    offset = source.get("offset", {})
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() in terms:
                target = sheet.cell(cell.row + int(offset.get("row", 0)),
                                    cell.column + int(offset.get("col", 1)))
                return target.coordinate
    return None


def _read_range(sheet, address: str):
    min_col, min_row, max_col, max_row = range_boundaries(address)
    values = [[sheet.cell(row, col).value for col in range(min_col, max_col + 1)]
              for row in range(min_row, max_row + 1)]
    return values[0][0] if len(values) == 1 and len(values[0]) == 1 else values


def _convert_scalar(value, value_type: str | None, unit: str | None,
                    normalization: dict):
    if value is None:
        return None
    if value_type == "number":
        if isinstance(value, str):
            value = float(value.replace(",", "").strip())
        elif not isinstance(value, (int, float)):
            raise ValueError(f"expected number, got {type(value).__name__}")
    elif value_type == "text":
        value = str(value)
    target = normalization.get("target_unit")
    if target and unit and target != unit and isinstance(value, (int, float)):
        conversions = {("K", "C"): lambda x: x - 273.15,
                       ("C", "K"): lambda x: x + 273.15,
                       ("s", "min"): lambda x: x / 60,
                       ("min", "s"): lambda x: x * 60}
        converter = conversions.get((unit, target))
        if converter is None:
            raise ValueError(f"unsupported unit normalization: {unit} -> {target}")
        value = converter(value)
    return value


def _normalize_value(value, mapping: dict):
    if isinstance(value, list):
        return [[_convert_scalar(cell, mapping.get("value_type"), mapping.get("unit"),
                                 mapping.get("normalization", {})) for cell in row]
                for row in value]
    return _convert_scalar(value, mapping.get("value_type"), mapping.get("unit"),
                           mapping.get("normalization", {}))


def prepare_parse(store: KgStore, document_id: str, document_version: str) -> tuple[dict, list[dict]]:
    assignment = store.conn.execute(
        "SELECT * FROM document_template_assignment WHERE document_id=? AND document_version=?",
        (document_id, document_version)).fetchone()
    if assignment is None:
        raise ParsingError("document version has no template assignment")
    return dict(assignment), effective_mappings(store, document_version)


def extract_workbook(path: Path, mappings: list[dict]) -> list[dict]:
    """CPU/file work only: safe to run without the shared DB lock."""
    workbook = load_workbook(path, data_only=True)
    extracted = []
    for mapping in mappings:
        matched = _route_sheet(workbook, mapping["matcher"])
        explicit_sheet = mapping["effective_source"].get("sheet")
        if explicit_sheet:
            matched = [workbook[explicit_sheet]] if explicit_sheet in workbook.sheetnames else []
        if not matched:
            extracted.append({"mapping": mapping, "sheet_name": explicit_sheet,
                              "source_range": None, "value": None, "status": "MISSING",
                              "mapping_source": mapping["mapping_source"],
                              "effective_source": mapping["effective_source"],
                              "warning": "sheet matcher returned 0 sheets"})
            continue
        # Repeated matches are intentional: each physical sheet is a source.
        for sheet in matched:
            manual = mapping["mapping_source"] == "MANUAL"
            manual_sheet = mapping["effective_source"].get("sheet")
            use_manual = manual and (manual_sheet is None or manual_sheet == sheet.title)
            effective = mapping["effective_source"] if use_manual else mapping["template_source"]
            source_range = effective.get("range") or _locate_key(sheet, effective)
            try:
                if not source_range:
                    raise ValueError("source was not found")
                value = _normalize_value(_read_range(sheet, source_range), mapping)
                status, warning = ("MANUAL" if use_manual else "TEMPLATE"), None
            except (ValueError, IndexError, TypeError) as exc:
                value, status, warning = None, "MISSING", str(exc)
            extracted.append({"mapping": mapping, "sheet_name": sheet.title,
                              "source_range": source_range, "value": value, "status": status,
                              "mapping_source": "MANUAL" if use_manual else "TEMPLATE",
                              "effective_source": effective, "warning": warning})
    workbook.close()
    return extracted


def save_parse_run(store: KgStore, document_id: str, document_version: str,
                   assignment: dict, extracted: list[dict]) -> dict:
    run_id, started = new_id("RUN"), now_iso()
    warnings = sum(item["warning"] is not None for item in extracted)
    overrides = sum(item["mapping_source"] == "MANUAL" and not item["warning"]
                    for item in extracted)
    successes = len(extracted) - warnings
    store.conn.execute(
        "INSERT INTO parse_run VALUES (?,?,?,?,?,?,?,?,?,?,?)",
        (run_id, document_id, document_version, assignment["template_id"],
         assignment["template_version"], started, None, "RUNNING", 0, 0, 0))
    for item in extracted:
        mapping = item["mapping"]
        store.conn.execute(
            "INSERT INTO parsed_source VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
            (new_id("SRC"), run_id, mapping["mapping_id"], mapping["concept_id"],
             item["sheet_name"], item["source_range"], item["mapping_source"],
             _dump(mapping["template_source"]), _dump(item["effective_source"]),
             _dump(item["value"]), item["status"], item["warning"]))
    final = "SUCCESS" if warnings == 0 else ("REVIEW_REQUIRED" if successes else "FAILED")
    assignment_status = "OVERRIDDEN" if overrides and final == "SUCCESS" else final.replace("SUCCESS", "PARSED")
    store.conn.execute(
        """UPDATE parse_run SET finished_at=?,status=?,mapping_count=?,override_count=?,warning_count=?
           WHERE parse_run_id=?""",
        (now_iso(), final, len(extracted), overrides, warnings, run_id))
    store.conn.execute(
        "UPDATE document_template_assignment SET status=? WHERE document_version=?",
        (assignment_status, document_version))
    return parse_result(store, run_id)


def run_parse(store: KgStore, document_id: str, document_version: str,
              path: Path) -> dict:
    assignment, mappings = prepare_parse(store, document_id, document_version)
    return save_parse_run(store, document_id, document_version, assignment,
                          extract_workbook(path, mappings))


def parse_result(store: KgStore, run_id: str) -> dict:
    run = store.conn.execute("SELECT * FROM parse_run WHERE parse_run_id=?", (run_id,)).fetchone()
    if run is None:
        raise ParsingError("unknown parse run")
    result = dict(run)
    result["sources"] = []
    for row in store.conn.execute(
            "SELECT * FROM parsed_source WHERE parse_run_id=? ORDER BY parsed_source_id", (run_id,)):
        source = dict(row)
        for key in ("template_source_json", "effective_source_json", "value_json"):
            source[key.removesuffix("_json")] = _load(source.pop(key))
        result["sources"].append(source)
    return result


def grouped_documents(store: KgStore, target_document_kg: str) -> list[dict]:
    rows = store.conn.execute(
        """SELECT t.template_id,t.name,a.template_version,d.document_id,d.filename,a.status,
                  (SELECT count(*) FROM document_override o
                    WHERE o.document_version=a.document_version
                      AND o.status IN ('APPROVED','CONFLICT')) overrides
             FROM parsing_template t
             JOIN document_template_assignment a ON a.template_id=t.template_id
             JOIN document d ON d.document_id=a.document_id
            WHERE t.target_document_kg=? ORDER BY t.name,a.template_version,d.filename""",
        (target_document_kg,)).fetchall()
    groups: dict[tuple, dict] = {}
    for row in rows:
        key = (row["template_id"], row["template_version"])
        group = groups.setdefault(key, {"template_id": key[0], "template_name": row["name"],
                                        "version": key[1], "documents": [],
                                        "override_documents": 0, "review_required": 0,
                                        "failed": 0})
        group["documents"].append({"document_id": row["document_id"],
                                   "filename": row["filename"], "status": row["status"],
                                   "override_count": row["overrides"]})
        group["override_documents"] += row["overrides"] > 0
        group["review_required"] += row["status"] == "REVIEW_REQUIRED"
        group["failed"] += row["status"] == "FAILED"
    return list(groups.values())
