"""KG 간단 웹 UI — 단일 화면: 왼쪽 개념 검색·검수, 오른쪽 웹 xlsx 뷰어.

기존 7-뷰 UI와 달리 화면 하나만 둔다. 서버가 원본 workbook을 파싱해 그리드
JSON(병합/채움색/굵기 포함)을 내려주고, 프론트가 표로 그린다 — 외부 뷰어
라이브러리/CDN 없음. 역탐색 소스나 검수 항목을 클릭하면 뷰어가 해당 시트로
이동해 locator 범위를 하이라이트한다.

    python -m kg.webapp --ws domains/financier --port 8010
"""
from __future__ import annotations

import re
import threading
from pathlib import Path

import json

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, JSONResponse, PlainTextResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from src.inspect.inspector import WorkbookInspector
from src.mapping.concepts import normalize_label

from kg.acquisition import (create_request, list_requests, mark_ingested,
                            refresh_release_states, request_row,
                            sniff_container)
from kg.domain.loader import VALID_RELATIONS
from kg.groups import (clear_member_override, document_kgs, group_documents,
                       is_l1_concept, isa_roots, member_overrides,
                       set_member_override)
from kg.ingest import apply_parsed, document_id_for, parse_workbook
from kg.mapping.mapper import map_nodes_staged
from kg.mapping.recipe import (active_recipe, apply_recipe, preview_recipe,
                               rollback_recipe, snapshot_recipe, suggest_groups)
from kg.recrawl import MODES, recover_interrupted_runs, run_recrawl, start_run
from kg.search import concept_neighbors, reverse_lookup
from kg.store import KgStore, new_id, now_iso

WEB_DIR = Path(__file__).parent / "web_kg"
_SHEET_CAP_ROWS = 300
_SHEET_CAP_COLS = 40


class ReviewAction(BaseModel):
    mapping_id: str
    action: str            # approve / reject


class RemapAction(BaseModel):
    node_id: str
    concept_id: str


class ProposalReq(BaseModel):
    node_ids: list[str]


class BuildField(BaseModel):
    name: str
    concept: str
    unit: str | None = None
    type: str | None = None


class BuildReq(BaseModel):
    name: str
    fields: list[BuildField]
    include_nodes: dict[str, list[str]] = {}
    raw_node_ids: list[str] = []     # 양식별 '원값 유지' — 단위 변환 생략할 노드


class IngestReq(BaseModel):
    filename: str
    group_id: str | None = None      # DKG(=L1 root concept) 배정
    force: bool = False
    map: bool = True                 # False = 구조 분석+DKG 제안만 (매핑 보류)
                                     # — 배정 확정 후 레시피→judge 순서 보장용


class MemberReq(BaseModel):
    document_id: str
    state: str                       # INCLUDED / EXCLUDED


class RecipeReq(BaseModel):
    note: str = ""


class RecrawlReq(BaseModel):
    mode: str = "fill"               # fill / reset_auto
    document_ids: list[str] | None = None


class ConceptReq(BaseModel):
    concept_id: str | None = None
    canonical_name: str | None = None
    canonical_name_en: str | None = None
    description: str | None = None
    concept_type: str | None = None
    data_type: str | None = None
    domain_level: str | None = None
    canonical_unit: str | None = None
    unit_dimension: str | None = None


class AliasReq(BaseModel):
    concept_id: str
    alias: str


class DrmReq(BaseModel):
    filename: str
    note: str = ""


class RelationReq(BaseModel):
    source: str
    target: str
    type: str


class ParsingTemplateReq(BaseModel):
    template_id: str
    name: str
    target_document_kg: str | None = None
    lifecycle: str = "DRAFT"


class ParsingVersionReq(BaseModel):
    spec: dict
    created_by: str | None = None


class ParsingAssignReq(BaseModel):
    document_version: str
    template_id: str
    template_version: int


class ParsingOverrideReq(BaseModel):
    document_version: str
    template_mapping_id: str
    override_source: dict
    reason: str | None = None
    created_by: str | None = None


class ParsingOverridePatchReq(BaseModel):
    override_source: dict | None = None
    reason: str | None = None
    status: str | None = None


class ParsingParseReq(BaseModel):
    document_version: str


class ViewerRegisterReq(BaseModel):
    document_version: str
    staging_name: str


class ViewerRenderReq(BaseModel):
    document_version: str


class ViewerUnlockReq(BaseModel):
    document_version: str
    note: str = ""


_CID_RE = re.compile(r"^[A-Za-z0-9_\-]{1,64}$")
_DOCID_RE = re.compile(r"^[0-9a-f]{16}$")
_RCP_RE = re.compile(r"^RCP-[0-9a-f]{12}$")
_RCL_RE = re.compile(r"^RCL-[0-9a-f]{12}$")


def _node_role(node_meta: dict, data_type: str | None, concept_type: str | None) -> str:
    """Semantic Overlay 역할 (§4.2): KEY / VALUE / CONTEXT."""
    if concept_type in ("identifier", "temporal"):
        return "KEY"
    if (node_meta or {}).get("region_type") in ("KEY_VALUE", "SUMMARY", "NOTE"):
        return "CONTEXT"
    if data_type == "numeric":
        return "VALUE"
    return "CONTEXT"


# ------------------------------------------------------ 원본 충실 렌더 ----
# §10.1: 병합/색/테두리/글꼴/열폭/행고/이미지를 보존한 서버측 렌더 모델.
# 렌더와 값 추출(파서)은 분리 — 이 모델은 화면 전용이다.
_EMU_PX = 9525.0
_BORDER_W = {"thin": 1, "hair": 1, "dotted": 1, "dashed": 1, "mediumDashed": 2,
             "medium": 2, "thick": 3, "double": 3}
_BORDER_STYLE = {"dotted": "dotted", "dashed": "dashed", "mediumDashed": "dashed",
                 "double": "double"}


def _rgb(color) -> str | None:
    if color is None or getattr(color, "type", None) != "rgb":
        return None
    v = color.rgb
    if not isinstance(v, str) or len(v) != 8 or v == "00000000":
        return None
    return "#" + v[-6:]


def _border_css(side) -> str | None:
    if side is None or side.style is None:
        return None
    w = _BORDER_W.get(side.style, 1)
    st = _BORDER_STYLE.get(side.style, "solid")
    return f"{w}px {st} {_rgb(side.color) or '#8d97a5'}"


def _fmt_value(v, fmt: str) -> tuple[str, bool]:
    """(표시 문자열, 숫자 여부) — Excel number_format의 흔한 경우를 근사한다."""
    import datetime as _dt
    if v is None:
        return "", False
    if isinstance(v, bool):
        return str(v), False
    if isinstance(v, _dt.datetime):
        return (v.strftime("%Y-%m-%d") if (v.hour, v.minute, v.second) == (0, 0, 0)
                else v.strftime("%Y-%m-%d %H:%M")), False
    if isinstance(v, (_dt.date, _dt.time)):
        return str(v), False
    if isinstance(v, (int, float)):
        f = fmt or "General"
        dec = 0
        if "." in f:
            dec = sum(1 for ch in f.split(".", 1)[1] if ch in "0#")
        try:
            if "%" in f:
                return f"{v * 100:.{dec}f}%", True
            if "#,##" in f or "#,#" in f:
                return f"{v:,.{dec}f}", True
            if f != "General" and any(ch in f for ch in "0#"):
                return f"{v:.{dec}f}", True
        except (ValueError, OverflowError):
            pass
        if isinstance(v, float):
            s = f"{v:.6f}".rstrip("0").rstrip(".")
            return (s if s else "0"), True
        return str(v), True
    return str(v), False


def _is_drm_file(path: Path) -> bool:
    """Detect NASCA DRM-encrypted xlsx by file header."""
    try:
        with open(path, 'rb') as f:
            return f.read(8).startswith(b'<## NASC')
    except Exception:
        return False


def _load_render_wb(path: Path):
    import openpyxl

    from src.inspect.inspector import _repair_sheet_names
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except ValueError:
        repaired = _repair_sheet_names(path)
        try:
            return openpyxl.load_workbook(repaired, data_only=True)
        finally:
            repaired.unlink(missing_ok=True)


def _render_sheet_drm(path: Path, sheet_name: str | None) -> dict:
    """DRM(NASCA) 파일 렌더 — Excel COM 경유 (Windows + Excel 필요).

    원본 충실이 기본이다: COM으로 연 시트를 새 워크북으로 복사해 임시 xlsx로
    저장한 뒤, 일반 경로(_render_sheet — 병합/스타일/열폭/행높이/이미지/
    텍스트박스)로 그린다 (src.inspect.inspector의 검증된 Copy→SaveAs 패턴).
    DRM 정책이 시트 복사/다른 이름 저장을 막는 환경에서만 값-전용 렌더로
    폴백하며, 이때는 degraded 사유를 내려 화면에 저하 렌더임을 명시한다 —
    깨진 레이아웃을 '원본 충실'인 척 보여주지 않는다.
    """
    import os
    import tempfile

    import pythoncom
    import win32com.client

    try:
        pythoncom.CoInitializeEx(pythoncom.COINIT_APARTMENTTHREADED)
    except Exception:
        pass

    excel = win32com.client.DispatchEx('Excel.Application')
    try:
        excel.Visible = False
    except Exception:
        pass
    try:
        excel.DisplayAlerts = False
    except Exception:
        pass

    try:
        wb = excel.Workbooks.Open(str(path.resolve()), 0, True)
        names = [wb.Sheets(i).Name for i in range(1, wb.Sheets.Count + 1)]
        target = sheet_name or (names[0] if names else None)
        if target is None or target not in names:
            wb.Close(SaveChanges=False)
            raise HTTPException(404, f"sheet not found: {sheet_name}")

        ws = wb.Sheets(target)

        # ── 1) 원본 충실: 시트 복사 → 임시 xlsx → 일반 충실 렌더러 ──
        degraded_reason = None
        try:
            tmp = Path(tempfile.gettempdir()) / (
                f"_drm_render_{os.getpid()}_{abs(hash((str(path), target)))}.xlsx")
            ws.Copy()                          # 시트 1개짜리 새 워크북 생성
            new_wb = excel.ActiveWorkbook
            try:
                new_wb.SaveAs(str(tmp), 51)    # 51 = xlOpenXMLWorkbook
            finally:
                new_wb.Close(SaveChanges=False)
            try:
                out = _render_sheet(tmp, None)     # 복사본은 시트가 1개
            finally:
                tmp.unlink(missing_ok=True)
            out["sheet"] = target
            out["sheets"] = names                  # 시트 목록은 원본 기준
            wb.Close(SaveChanges=False)
            return out
        except HTTPException:
            raise
        except Exception as exc:                   # DRM 정책이 복사/저장 차단
            degraded_reason = str(exc).strip()[:160] or type(exc).__name__

        # ── 2) 폴백: 값-전용 (서식/병합 없음 — degraded로 명시) ──
        used = ws.UsedRange
        max_row = min(used.Rows.Count if used else 1, _SHEET_CAP_ROWS)
        max_col = min(used.Columns.Count if used else 1, _SHEET_CAP_COLS)

        # 실제 열 너비 / 행 높이 읽기
        col_px = []
        for c in range(1, max_col + 1):
            try:
                w = ws.Columns(c).ColumnWidth
                col_px.append(max(14, round((w or 8.43) * 7 + 5)))
            except Exception:
                col_px.append(64)
        row_px = []
        for r in range(1, max_row + 1):
            try:
                h = ws.Rows(r).RowHeight
                row_px.append(max(12, round((h or 15.0) * 4 / 3)))
            except Exception:
                row_px.append(20)

        # Bulk read values only — fast (~1s per sheet)
        cells = []
        if max_row > 0 and max_col > 0:
            try:
                data_range = ws.Range(ws.Cells(1, 1), ws.Cells(max_row, max_col))
                values = data_range.Value
            except Exception:
                values = None

            if values:
                for r_idx, row_data in enumerate(values, 1):
                    if row_data is None:
                        continue
                    for c_idx, val in enumerate(row_data, 1):
                        if val is None:
                            continue
                        v_str, is_num = _fmt_value(val, "General")
                        if v_str == "":
                            continue
                        d = {"r": r_idx, "c": c_idx, "v": v_str}
                        if is_num:
                            d["n"] = 1
                        cells.append(d)

        # 텍스트박스만이라도 건진다 (COM points 좌표 — 프런트가 96/72 환산)
        shapes = []
        try:
            for i in range(1, ws.Shapes.Count + 1):
                sh = ws.Shapes.Item(i)
                si: dict = {}
                try:
                    tf = sh.TextFrame2
                    if tf.HasText == -1:       # msoTrue
                        si["text"] = tf.TextRange.Text
                except Exception:
                    pass
                if "text" not in si:
                    try:
                        si["text"] = sh.TextFrame.Characters().Text
                    except Exception:
                        pass
                try:
                    si.update({"left": round(sh.Left), "top": round(sh.Top),
                               "width": round(sh.Width),
                               "height": round(sh.Height)})
                except Exception:
                    pass
                if si.get("text", "").strip():
                    shapes.append(si)
                if len(shapes) >= 60:
                    break
        except Exception:
            pass

        total_rows = used.Rows.Count if used else 1
        total_cols = used.Columns.Count if used else 1
        wb.Close(SaveChanges=False)
        return {
            "sheet": target, "sheets": names,
            "max_row": max_row, "max_col": max_col,
            "cols": col_px, "rows": row_px,
            "cells": cells, "images": [], "shapes": shapes,
            "gridlines": True,
            "truncated": total_rows > _SHEET_CAP_ROWS or total_cols > _SHEET_CAP_COLS,
            "degraded": "DRM 보호로 시트 복사가 차단되어 값만 표시합니다 — "
                        f"병합/서식 없음 ({degraded_reason})",
        }
    finally:
        try:
            excel.Quit()
        except Exception:
            pass


_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_DML = "http://schemas.openxmlformats.org/drawingml/2006/main"
_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
_DOC_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_SSML = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


_PROPS_CACHE: dict[str, tuple[float, dict]] = {}


def _doc_props(path: Path) -> dict:
    """xlsx core 속성(작성자/작성일/수정일) — docProps/core.xml만 가볍게 읽는다.

    파일 mtime 키 캐시로 목록 조회마다 재파싱하지 않는다. 작성일이 없으면
    파일 mtime으로 대체하고, 파일이 없으면(이동/CSV 등) 전부 None.
    """
    try:
        mtime = path.stat().st_mtime
    except OSError:
        return {"author": None, "created": None, "modified": None}
    hit = _PROPS_CACHE.get(str(path))
    if hit and hit[0] == mtime:
        return hit[1]
    props: dict = {"author": None, "created": None, "modified": None}
    try:
        import xml.etree.ElementTree as ET
        import zipfile
        with zipfile.ZipFile(path) as zf:
            core = ET.fromstring(zf.read("docProps/core.xml"))
        dc = "{http://purl.org/dc/elements/1.1/}"
        dct = "{http://purl.org/dc/terms/}"
        props["author"] = (core.findtext(f"{dc}creator") or "").strip() or None
        props["created"] = (core.findtext(f"{dct}created") or "").strip() or None
        props["modified"] = (core.findtext(f"{dct}modified") or "").strip() or None
    except Exception:
        pass                       # 잠긴/비정형 컨테이너 — 속성 없이 진행
    if not props["created"]:
        from datetime import datetime, timezone
        props["created"] = datetime.fromtimestamp(
            mtime, tz=timezone.utc).isoformat(timespec="seconds")
    _PROPS_CACHE[str(path)] = (mtime, props)
    return props


def _sheet_textboxes_raw(path: Path, sheet_name: str) -> list[dict]:
    """xlsx 드로잉 XML에서 텍스트박스/도형 텍스트를 앵커 원시값으로 추출한다.

    openpyxl은 도형(xdr:sp)을 노출하지 않으므로 zip에서 드로잉 XML을 직접
    읽는다. px 환산은 _textboxes_px가 그리드 확정 후에 수행한다 — 그리드
    범위를 드로잉까지 확장하는 계산에도 이 원시 앵커가 쓰인다.
    """
    import xml.etree.ElementTree as ET
    import zipfile

    def rels_map(zf: zipfile.ZipFile, rels_path: str) -> dict[str, str]:
        try:
            root = ET.fromstring(zf.read(rels_path))
        except (KeyError, ET.ParseError):
            return {}
        return {rel.get("Id"): rel.get("Target", "")
                for rel in root.findall(f"{{{_PKG_REL}}}Relationship")}

    def resolve(base_dir: str, target: str) -> str:
        if target.startswith("/"):
            return target.lstrip("/")
        parts: list[str] = base_dir.split("/") if base_dir else []
        for seg in target.split("/"):
            if seg == "..":
                if parts:
                    parts.pop()
            elif seg not in ("", "."):
                parts.append(seg)
        return "/".join(parts)

    boxes: list[dict] = []
    try:
        with zipfile.ZipFile(path) as zf:
            # 시트 이름 → 워크시트 파트 경로 (workbook.xml + rels)
            wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
            wb_rels = rels_map(zf, "xl/_rels/workbook.xml.rels")
            ws_part = None
            for sh in wb_root.findall(f"{{{_SSML}}}sheets/{{{_SSML}}}sheet"):
                if sh.get("name") == sheet_name:
                    ws_part = resolve("xl", wb_rels.get(
                        sh.get(f"{{{_DOC_REL}}}id", ""), ""))
                    break
            if not ws_part:
                return []
            # 워크시트 → 드로잉 파트
            ws_dir, ws_file = ws_part.rsplit("/", 1)
            ws_rels = rels_map(zf, f"{ws_dir}/_rels/{ws_file}.rels")
            ws_root = ET.fromstring(zf.read(ws_part))
            drawing_parts = []
            for dr in ws_root.findall(f"{{{_SSML}}}drawing"):
                target = ws_rels.get(dr.get(f"{{{_DOC_REL}}}id", ""), "")
                if target:
                    drawing_parts.append(resolve(ws_dir, target))
            def marker(m) -> tuple[int, int, int, int]:
                return (int(m.findtext(f"{{{_XDR}}}col", "0")),
                        int(m.findtext(f"{{{_XDR}}}colOff", "0")),
                        int(m.findtext(f"{{{_XDR}}}row", "0")),
                        int(m.findtext(f"{{{_XDR}}}rowOff", "0")))

            for part in drawing_parts:
                try:
                    droot = ET.fromstring(zf.read(part))
                except (KeyError, ET.ParseError):
                    continue
                for anchor in droot:
                    kind = anchor.tag.rsplit("}", 1)[-1]
                    if kind not in ("twoCellAnchor", "oneCellAnchor",
                                    "absoluteAnchor"):
                        continue
                    sp = anchor.find(f"{{{_XDR}}}sp")
                    if sp is None:
                        continue
                    tx = sp.find(f"{{{_XDR}}}txBody")
                    if tx is None:
                        continue
                    text = "\n".join(
                        "".join(t.text or ""
                                for t in p.findall(f".//{{{_DML}}}t"))
                        for p in tx.findall(f"{{{_DML}}}p"))
                    if not text.strip():
                        continue
                    frm = anchor.find(f"{{{_XDR}}}from")
                    to = anchor.find(f"{{{_XDR}}}to")
                    ext = anchor.find(f"{{{_XDR}}}ext")
                    pos = anchor.find(f"{{{_XDR}}}pos")
                    box: dict = {"text": text[:4000]}
                    if frm is not None:
                        box["frm"] = marker(frm)
                    elif pos is not None:
                        box["pos"] = (int(pos.get("x", "0")),
                                      int(pos.get("y", "0")))
                    else:
                        continue
                    if to is not None:
                        box["to"] = marker(to)
                    elif ext is not None:
                        box["ext"] = (int(ext.get("cx", "0")),
                                      int(ext.get("cy", "0")))
                    boxes.append(box)
                    if len(boxes) >= 60:
                        return boxes
    except Exception:
        return boxes           # 드로잉이 없거나 비정형 — 셀 렌더는 그대로 진행
    return boxes


def _textboxes_px(raw: list[dict], cum_x: list[float], cum_y: list[float],
                  max_r: int, max_c: int) -> list[dict]:
    """원시 앵커를 이미지와 같은 방식(누적 열폭/행높이 + EMU)으로 px 환산."""
    emu = _EMU_PX

    def mpx(m: tuple[int, int, int, int]) -> tuple[float, float]:
        col, coff, row, roff = m
        return (cum_x[min(col, max_c)] + coff / emu,
                cum_y[min(row, max_r)] + roff / emu)

    out = []
    for b in raw:
        if "frm" in b:
            x, y = mpx(b["frm"])
        else:
            x, y = b["pos"][0] / emu, b["pos"][1] / emu
        if "to" in b:
            x2, y2 = mpx(b["to"])
            w, h = x2 - x, y2 - y
        elif "ext" in b:
            w, h = b["ext"][0] / emu, b["ext"][1] / emu
        else:
            w, h = 160, 40
        if w <= 4 or h <= 4:
            continue
        out.append({"text": b["text"], "x": round(x), "y": round(y),
                    "w": round(w), "h": round(h)})
    return out


def _render_sheet(path: Path, sheet_name: str | None) -> dict:
    import base64

    from openpyxl.utils import get_column_letter, range_boundaries
    wb = _load_render_wb(path)
    names = wb.sheetnames
    target = sheet_name or (names[0] if names else None)
    if target is None or target not in names:
        raise HTTPException(404, f"sheet not found: {sheet_name}")
    ws = wb[target]
    raw_boxes = _sheet_textboxes_raw(path, target)

    # 데이터 범위 밖(아래/오른쪽)에 앵커된 이미지·텍스트박스도 원본 위치
    # 그대로 보이도록, 그리드를 드로잉 범위까지 확장한다. 범위 밖 행/열은
    # 기본 크기(64px/20px)라 ext 기반 앵커의 소요 칸 수 추정도 정확하다.
    need_r, need_c = ws.max_row or 1, ws.max_column or 1
    _DEF_W, _DEF_H = 64, 20

    def _extend(frm_cr: tuple[int, int] | None, to_cr: tuple[int, int] | None,
                ext_wh: tuple[float, float] | None) -> None:
        nonlocal need_r, need_c
        if frm_cr is None:
            return
        need_c = max(need_c, frm_cr[0] + 1)
        need_r = max(need_r, frm_cr[1] + 1)
        if to_cr is not None:
            need_c = max(need_c, to_cr[0])
            need_r = max(need_r, to_cr[1])
        elif ext_wh is not None:
            need_c = max(need_c, frm_cr[0] + 1 + int(ext_wh[0] // _DEF_W) + 1)
            need_r = max(need_r, frm_cr[1] + 1 + int(ext_wh[1] // _DEF_H) + 1)

    for img in getattr(ws, "_images", []):
        frm = getattr(img.anchor, "_from", None)
        to = getattr(img.anchor, "to", None)
        ext = getattr(img.anchor, "ext", None)
        _extend((frm.col, frm.row) if frm is not None else None,
                (to.col, to.row) if to is not None else None,
                (ext.cx / _EMU_PX, ext.cy / _EMU_PX) if ext is not None else None)
    for b in raw_boxes:
        frm, to, ext = b.get("frm"), b.get("to"), b.get("ext")
        _extend((frm[0], frm[2]) if frm else None,
                (to[0], to[2]) if to else None,
                (ext[0] / _EMU_PX, ext[1] / _EMU_PX) if ext else None)

    max_r = min(need_r, _SHEET_CAP_ROWS)
    max_c = min(need_c, _SHEET_CAP_COLS)

    # 열 너비(문자폭→px) / 행 높이(pt→px) — 원본 레이아웃의 뼈대
    col_px = []
    for c in range(1, max_c + 1):
        dim = ws.column_dimensions.get(get_column_letter(c))
        w = dim.width if dim is not None and dim.width else 8.43
        col_px.append(max(14, round(w * 7 + 5)))
    row_px = []
    for r in range(1, max_r + 1):
        dim = ws.row_dimensions.get(r)
        h = dim.height if dim is not None and dim.height else 15.0
        row_px.append(max(12, round(h * 4 / 3)))

    spans: dict[str, tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for m in ws.merged_cells.ranges:
        a, b, c2, d = range_boundaries(str(m))
        spans[f"{b},{a}"] = (min(d, max_r) - b + 1, min(c2, max_c) - a + 1)
        for rr in range(b, min(d, max_r) + 1):
            for cc in range(a, min(c2, max_c) + 1):
                if (rr, cc) != (b, a):
                    covered.add((rr, cc))

    cells = []
    for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
        for cell in row:
            key = (cell.row, cell.column)
            if key in covered:
                continue
            v, is_num = _fmt_value(cell.value, cell.number_format)
            fill = None
            if cell.fill is not None and cell.fill.patternType == "solid":
                fill = _rgb(cell.fill.fgColor)
            fnt = cell.font
            bd = {}
            for side, name in ((cell.border.top, "t"), (cell.border.right, "r"),
                               (cell.border.bottom, "b"), (cell.border.left, "l")):
                css = _border_css(side)
                if css:
                    bd[name] = css
            al = cell.alignment
            if v == "" and not fill and not bd and key not in \
                    {tuple(map(int, k.split(","))) for k in spans}:
                continue
            rs, cs = spans.get(f"{cell.row},{cell.column}", (1, 1))
            d = {"r": cell.row, "c": cell.column, "v": v}
            if rs > 1 or cs > 1:
                d["rs"], d["cs"] = rs, cs
            if is_num:
                d["n"] = 1
            if fill:
                d["f"] = fill
            if fnt is not None:
                if fnt.bold:
                    d["b"] = 1
                if fnt.italic:
                    d["i"] = 1
                if fnt.size and abs(float(fnt.size) - 11.0) > 0.1:
                    d["sz"] = round(float(fnt.size) * 4 / 3)
                fc = _rgb(fnt.color)
                if fc and fc != "#000000":
                    d["fc"] = fc
            if bd:
                d["bd"] = bd
            if al is not None:
                if al.horizontal in ("center", "right", "left"):
                    d["ha"] = al.horizontal[0]
                if al.wrap_text:
                    d["wr"] = 1
            cells.append(d)

    # 이미지 — 앵커 좌표(EMU)를 px로 환산해 그리드 위에 겹친다
    cum_x = [0]
    for w in col_px:
        cum_x.append(cum_x[-1] + w)
    cum_y = [0]
    for h in row_px:
        cum_y.append(cum_y[-1] + h)
    images = []
    total_img = 0
    for img in getattr(ws, "_images", []):
        try:
            data = img._data()
        except Exception:
            continue
        if len(data) > 1_500_000 or total_img + len(data) > 4_000_000:
            continue
        frm = getattr(img.anchor, "_from", None)
        if frm is None:
            continue
        col0 = min(frm.col, max_c - 1)
        row0 = min(frm.row, max_r - 1)
        x = cum_x[col0] + frm.colOff / _EMU_PX
        y = cum_y[row0] + frm.rowOff / _EMU_PX
        to = getattr(img.anchor, "to", None)
        ext = getattr(img.anchor, "ext", None)
        if to is not None:
            w = cum_x[min(to.col, max_c)] + to.colOff / _EMU_PX - x
            h = cum_y[min(to.row, max_r)] + to.rowOff / _EMU_PX - y
        elif ext is not None:
            w, h = ext.cx / _EMU_PX, ext.cy / _EMU_PX
        else:
            w, h = 160, 120
        if w <= 4 or h <= 4:
            continue
        total_img += len(data)
        mt = (getattr(img, "format", None) or "png").lower()
        images.append({"x": round(x), "y": round(y), "w": round(w), "h": round(h),
                       "src": f"data:image/{mt};base64,"
                              f"{base64.b64encode(data).decode()}"})

    return {"sheet": target, "sheets": names, "max_row": max_r, "max_col": max_c,
            "cols": col_px, "rows": row_px, "cells": cells, "images": images,
            "shapes": _textboxes_px(raw_boxes, cum_x, cum_y, max_r, max_c),
            "gridlines": bool(ws.sheet_view.showGridLines),
            "truncated": (ws.max_row or 1) > _SHEET_CAP_ROWS or
                         (ws.max_column or 1) > _SHEET_CAP_COLS}


def _grid_json(structure, sheet_name: str) -> dict:
    sheet = next((s for s in structure.sheets if s.sheet_name == sheet_name), None)
    if sheet is None:
        raise HTTPException(404, f"sheet not found: {sheet_name}")
    from openpyxl.utils import range_boundaries
    spans: dict[str, tuple[int, int]] = {}
    for rng in sheet.merged_ranges:
        a, b, c, d = range_boundaries(rng)
        spans[rng] = (d - b + 1, c - a + 1)          # (rowspan, colspan)
    cells = []
    max_r = max_c = 1
    for cell in sheet.cells:
        if cell.row > _SHEET_CAP_ROWS or cell.col > _SHEET_CAP_COLS:
            continue
        if cell.merged_into:                          # 병합 피복 셀은 마스터가 그린다
            max_r, max_c = max(max_r, cell.row), max(max_c, cell.col)
            continue
        rs, cs = spans.get(cell.merged_range or "", (1, 1))
        v = cell.cached_value if cell.is_formula and cell.cached_value is not None \
            else cell.value
        cells.append({
            "r": cell.row, "c": cell.col,
            "v": "" if v is None else str(v),
            "b": 1 if cell.bold else 0,
            "f": f"#{cell.fill_rgb[-6:]}" if cell.fill_rgb else None,
            "rs": rs, "cs": cs,
        })
        max_r = max(max_r, cell.row + rs - 1)
        max_c = max(max_c, cell.col + cs - 1)
    return {"sheet": sheet_name, "max_row": min(max_r, _SHEET_CAP_ROWS),
            "max_col": min(max_c, _SHEET_CAP_COLS), "cells": cells,
            "truncated": sheet.max_row > _SHEET_CAP_ROWS or
                         sheet.max_col > _SHEET_CAP_COLS}


def create_app(ws_root: str | Path) -> FastAPI:
    from kg.cli import Workspace
    from contextlib import contextmanager
    root = Path(ws_root).resolve()
    store = KgStore(root / "data" / "kg" / "kg.db", threadsafe=True)
    lock = threading.Lock()

    @contextmanager
    def wlock():
        """쓰기 구간용 lock — 예외 시 lock을 쥔 채 rollback 후 전파한다.

        with 블록을 예외로 탈출하면 lock이 먼저 풀리므로, 미커밋 부분 쓰기가
        경쟁 요청의 commit에 편승해 영구화되는 창이 생긴다(리뷰 확정 결함).
        모든 쓰기 엔드포인트는 lock 대신 이것을 쓴다.
        """
        with lock:
            try:
                yield
            except Exception:
                try:
                    store.conn.rollback()
                except Exception:
                    pass
                raise
    ws = Workspace(root, store=store)     # parser_rules/units/registry 공유 컨텍스트
    inspector = WorkbookInspector()
    struct_cache: dict[tuple[str, float], object] = {}
    recover_interrupted_runs(store)       # RUNNING 잔류 → FAILED (전역 409 해제)
    recrawl_state = {"busy": False, "run_id": None}

    app = FastAPI(title="KG viewer", docs_url=None, redoc_url=None)

    def _fresh_retriever():
        """생성 시점 캐시 — KG 편집 반영을 위해 매 작업마다 새로 만든다.
        (호출측이 lock을 잡고 부른다)"""
        from kg.mapping.retriever import DomainRetriever
        return DomainRetriever(store, units=ws.units)

    def _judge():
        from kg.mapping.judge import get_judge
        return get_judge()

    def _tree_node_table(document_id: str, sheet: str) -> dict | None:
        """tree_node에서 테이블 형태 JSON 생성 (CSV SECTION / DRM TABLE)."""
        with lock:
            # SECTION 기반 (CSV)
            sections = store.conn.execute("""
                SELECT node_id, node_name, tree_path FROM tree_node
                WHERE document_id=? AND status='ACTIVE' AND node_type='SECTION'
                  AND tree_path LIKE ?
                ORDER BY tree_path LIMIT 500
            """, (document_id, f"%/{sheet}/%")).fetchall()
            # TABLE 기반 (DRM 체크시트)
            if not sections:
                sections = store.conn.execute("""
                    SELECT node_id, node_name, tree_path FROM tree_node
                    WHERE document_id=? AND status='ACTIVE' AND node_type='TABLE'
                      AND tree_path LIKE ?
                    ORDER BY tree_path LIMIT 500
                """, (document_id, f"%/{sheet}/%")).fetchall()
            if not sections:
                return None
            cells = []
            col_map = {}
            max_col = 0
            for row_idx, sec in enumerate(sections):
                children = store.conn.execute("""
                    SELECT h.node_name as header, v.node_name as value,
                           json_extract(h.metadata, '$.concept_hint') as concept_hint
                    FROM tree_node h
                    LEFT JOIN tree_node v ON v.parent_node_id=h.node_id
                        AND v.node_type='VALUE' AND v.status='ACTIVE'
                    WHERE h.parent_node_id=? AND h.node_type='HEADER' AND h.status='ACTIVE'
                    ORDER BY h.tree_path
                """, (sec["node_id"],)).fetchall()
                for child in children:
                    hdr = child["header"]
                    if hdr not in col_map:
                        col_map[hdr] = max_col
                        cells.append({"r": 0, "c": max_col, "v": hdr, "b": 1,
                                     "f": "#D9D9D9", "bd": {"t": "1px solid #8d97a5",
                                     "r": "1px solid #8d97a5", "b": "1px solid #8d97a5",
                                     "l": "1px solid #8d97a5"}})
                        max_col += 1
                    col = col_map[hdr]
                    val = child["value"] or ""
                    cell = {"r": row_idx + 1, "c": col, "v": val,
                           "bd": {"t": "1px solid #ddd", "r": "1px solid #ddd",
                                  "b": "1px solid #ddd", "l": "1px solid #ddd"}}
                    if child["concept_hint"]:
                        cell["concept_hint"] = child["concept_hint"]
                    cells.append(cell)
            return {"sheet": sheet, "cells": cells,
                    "max_row": len(sections) + 1, "max_col": max_col,
                    "col_widths": {str(c): 120 for c in range(max_col)},
                    "row_heights": {str(r): 24 for r in range(len(sections) + 1)},
                    "images": [], "gridlines": True, "source": "tree_node"}

    def _doc_path(document_id: str) -> Path:
        with lock:
            row = store.conn.execute(
                "SELECT filepath, filename FROM document WHERE document_id=?",
                (document_id,)).fetchone()
        if row is None:
            raise HTTPException(404, "unknown document")
        p = Path(row["filepath"] or "").resolve()
        if not p.exists() or p.suffix.lower() != ".xlsx":
            # 경로가 이사했으면 워크스페이스 raw에서 같은 이름을 찾는다
            fallback = root / "data" / "raw" / row["filename"]
            if not fallback.exists():
                raise HTTPException(404, f"file missing: {row['filename']}")
            p = fallback
        return p

    def _structure(path: Path):
        key = (str(path), path.stat().st_mtime)
        if key not in struct_cache:
            struct_cache.clear()                      # 파일당 1개만 유지 (경량)
            struct_cache[key] = inspector.inspect(path)
        return struct_cache[key]

    # ------------------------------------------------------------- reads ----
    @app.get("/api/documents")
    def documents():
        with lock:
            rows = store.conn.execute(
                """SELECT d.document_id, d.filename,
                     (SELECT drm_status FROM viewer_document_version v
                       WHERE v.document_id=d.document_id
                         AND v.document_version=d.current_version) drm_status,
                     (SELECT render_status FROM viewer_document_version v
                       WHERE v.document_id=d.document_id
                         AND v.document_version=d.current_version) render_status,
                     (SELECT a.status FROM document_template_assignment a
                       WHERE a.document_id=d.document_id
                         AND a.document_version=d.current_version) parsing_status,
                          count(DISTINCT n.node_id) nodes
                   FROM document d
                   LEFT JOIN tree_node n ON n.document_id=d.document_id
                        AND n.status='ACTIVE' AND n.node_type='HEADER'
                   GROUP BY d.document_id ORDER BY d.filename""").fetchall()
        return [dict(r) for r in rows]

    @app.get("/api/concepts")
    def concepts():
        with lock:
            rows = store.conn.execute(
                """SELECT c.concept_id, c.canonical_name, c.domain_level,
                          count(m.mapping_id) sources
                   FROM domain_concept c
                   LEFT JOIN semantic_mapping m ON m.concept_id=c.concept_id
                        AND m.is_active=1 AND m.status IN ('AUTO_APPROVED','APPROVED')
                   WHERE c.status='ACTIVE'
                   GROUP BY c.concept_id ORDER BY sources DESC, c.concept_id""").fetchall()
        return [dict(r) for r in rows]

    # ----------------------------------------------- Parsing Templates ----
    @app.get("/api/parsing/templates")
    def parsing_templates():
        from kg.parsing import template_detail
        with lock:
            ids = [r[0] for r in store.conn.execute(
                "SELECT template_id FROM parsing_template ORDER BY name")]
            return [template_detail(store, template_id) for template_id in ids]

    @app.post("/api/parsing/templates", status_code=201)
    def parsing_template_create(req: ParsingTemplateReq):
        from kg.parsing import ParsingError, create_template
        if not _CID_RE.fullmatch(req.template_id):
            raise HTTPException(422, "invalid template_id")
        try:
            with wlock():
                result = create_template(store, req.template_id, req.name,
                                         req.target_document_kg, req.lifecycle)
                store.commit()
            return result
        except ParsingError as exc:
            raise HTTPException(422, str(exc)) from exc
        except Exception as exc:
            if "UNIQUE constraint" in str(exc):
                raise HTTPException(409, "template already exists") from exc
            raise

    @app.get("/api/parsing/templates/{template_id}")
    def parsing_template_get(template_id: str):
        from kg.parsing import ParsingError, template_detail
        try:
            with lock:
                return template_detail(store, template_id)
        except ParsingError as exc:
            raise HTTPException(404, str(exc)) from exc

    @app.post("/api/parsing/templates/{template_id}/versions", status_code=201)
    def parsing_version_create(template_id: str, req: ParsingVersionReq):
        from kg.parsing import ParsingError, add_version
        try:
            with wlock():
                result = add_version(store, template_id, req.spec, req.created_by)
                store.commit()
            return result
        except ParsingError as exc:
            code = 404 if str(exc) == "unknown template" else 422
            raise HTTPException(code, str(exc)) from exc

    @app.get("/api/parsing/templates/{template_id}/versions/{version}")
    def parsing_version_get(template_id: str, version: int):
        from kg.parsing import ParsingError, version_detail
        try:
            with lock:
                return version_detail(store, template_id, version)
        except ParsingError as exc:
            raise HTTPException(404, str(exc)) from exc

    @app.get("/api/parsing/templates/{template_id}/versions/{version}/documents")
    def parsing_version_documents(template_id: str, version: int):
        with lock:
            rows = store.conn.execute(
                """SELECT a.*,d.filename FROM document_template_assignment a
                     JOIN document d ON d.document_id=a.document_id
                    WHERE a.template_id=? AND a.template_version=? ORDER BY d.filename""",
                (template_id, version)).fetchall()
        return [dict(r) for r in rows]

    @app.post("/api/parsing/documents/{document_id}/assign")
    def parsing_assign(document_id: str, req: ParsingAssignReq):
        from kg.parsing import ParsingError, assign
        try:
            with wlock():
                result = assign(store, document_id, req.document_version,
                                req.template_id, req.template_version)
                store.commit()
            return result
        except ParsingError as exc:
            raise HTTPException(422, str(exc)) from exc

    @app.get("/api/parsing/documents/{document_id}/overrides")
    def parsing_overrides(document_id: str, document_version: str | None = None):
        query = "SELECT * FROM document_override WHERE document_id=?"
        params: list = [document_id]
        if document_version:
            query += " AND document_version=?"
            params.append(document_version)
        with lock:
            rows = store.conn.execute(query + " ORDER BY created_at", params).fetchall()
        result = []
        for row in rows:
            item = dict(row)
            item["override_source"] = json.loads(item.pop("override_source_json"))
            result.append(item)
        return result

    @app.post("/api/parsing/documents/{document_id}/overrides", status_code=201)
    def parsing_override_create(document_id: str, req: ParsingOverrideReq):
        from kg.parsing import ParsingError, save_override
        try:
            with wlock():
                result = save_override(store, document_id, req.document_version,
                                       req.template_mapping_id, req.override_source,
                                       req.reason, req.created_by)
                store.commit()
            return result
        except ParsingError as exc:
            raise HTTPException(422, str(exc)) from exc

    @app.patch("/api/parsing/documents/{document_id}/overrides/{override_id}")
    def parsing_override_patch(document_id: str, override_id: str,
                               req: ParsingOverridePatchReq):
        from kg.parsing import ParsingError, override_detail
        with wlock():
            old = store.conn.execute(
                "SELECT * FROM document_override WHERE override_id=? AND document_id=?",
                (override_id, document_id)).fetchone()
            if old is None:
                raise HTTPException(404, "unknown override")
            if req.status is not None and req.status not in {"APPROVED", "PENDING", "REJECTED"}:
                raise HTTPException(422, "invalid override status")
            source = req.override_source or json.loads(old["override_source_json"])
            if not source.get("range"):
                raise HTTPException(422, "override_source.range is required")
            store.conn.execute(
                """UPDATE document_override SET override_source_json=?,reason=?,status=?,updated_at=?
                     WHERE override_id=?""",
                (json.dumps(source, ensure_ascii=False),
                 req.reason if req.reason is not None else old["reason"],
                 req.status or old["status"], now_iso(), override_id))
            store.commit()
            try:
                return override_detail(store, override_id)
            except ParsingError as exc:  # defensive: row existed in same transaction
                raise HTTPException(404, str(exc)) from exc

    @app.delete("/api/parsing/documents/{document_id}/overrides/{override_id}")
    def parsing_override_delete(document_id: str, override_id: str):
        with wlock():
            cur = store.conn.execute(
                "DELETE FROM document_override WHERE override_id=? AND document_id=?",
                (override_id, document_id))
            if cur.rowcount == 0:
                raise HTTPException(404, "unknown override")
            store.commit()
        return {"ok": True}

    @app.get("/api/parsing/documents/{document_id}/effective-mappings")
    def parsing_effective(document_id: str, document_version: str):
        from kg.parsing import ParsingError, effective_mappings
        try:
            with lock:
                row = store.conn.execute(
                    "SELECT document_id FROM document_version WHERE version_id=?",
                    (document_version,)).fetchone()
                if row is None or row["document_id"] != document_id:
                    raise HTTPException(404, "unknown document version")
                return effective_mappings(store, document_version)
        except ParsingError as exc:
            raise HTTPException(422, str(exc)) from exc

    @app.post("/api/parsing/documents/{document_id}/parse")
    def parsing_parse(document_id: str, req: ParsingParseReq):
        from kg.parsing import (ParsingError, extract_workbook, prepare_parse,
                                save_parse_run)
        path = _doc_path(document_id)
        try:
            with lock:
                assignment, mappings = prepare_parse(store, document_id,
                                                     req.document_version)
            # Workbook IO and extraction may be expensive; never hold the
            # process-wide DB lock while openpyxl walks the workbook.
            extracted = extract_workbook(path, mappings)
            with wlock():
                current = store.conn.execute(
                    "SELECT template_id,template_version FROM document_template_assignment WHERE document_version=?",
                    (req.document_version,)).fetchone()
                if current is None or (current["template_id"], current["template_version"]) != \
                        (assignment["template_id"], assignment["template_version"]):
                    raise HTTPException(409, "template assignment changed during parsing")
                result = save_parse_run(store, document_id, req.document_version,
                                        assignment, extracted)
                store.commit()
            return result
        except ParsingError as exc:
            raise HTTPException(422, str(exc)) from exc

    @app.get("/api/parsing/documents/{document_id}/result")
    def parsing_result(document_id: str, parse_run_id: str | None = None):
        from kg.parsing import ParsingError, parse_result
        with lock:
            rid = parse_run_id
            if rid is None:
                row = store.conn.execute(
                    "SELECT parse_run_id FROM parse_run WHERE document_id=? ORDER BY started_at DESC LIMIT 1",
                    (document_id,)).fetchone()
                rid = row[0] if row else None
            if rid is None:
                raise HTTPException(404, "no parse result")
            try:
                result = parse_result(store, rid)
            except ParsingError as exc:
                raise HTTPException(404, str(exc)) from exc
        if result["document_id"] != document_id:
            raise HTTPException(404, "unknown parse run")
        return result

    @app.get("/api/parsing/document-kg/{dkg_id}/groups")
    def parsing_document_kg_groups(dkg_id: str):
        from kg.parsing import grouped_documents
        with lock:
            return grouped_documents(store, dkg_id)

    # --------------------------------------------------- Read-only Viewer ----
    # Filesystem paths are deliberately never included in these responses.
    @app.get("/api/documents/{document_id}/drm-status")
    def viewer_drm_status(document_id: str, document_version: str | None = None):
        with lock:
            version = document_version
            if version is None:
                doc = store.conn.execute(
                    "SELECT current_version,filename FROM document WHERE document_id=?",
                    (document_id,)).fetchone()
                if doc is None:
                    raise HTTPException(404, "unknown document")
                version = doc["current_version"]
            row = store.conn.execute(
                """SELECT drm_status,drm_error FROM viewer_document_version
                   WHERE document_id=? AND document_version=?""",
                (document_id, version)).fetchone()
            if row:
                return {"document_id": document_id, "document_version": version,
                        "drm_status": row["drm_status"], "error": row["drm_error"]}
            doc = store.conn.execute(
                "SELECT filename FROM document WHERE document_id=?", (document_id,)).fetchone()
            request = request_row(store, doc["filename"]) if doc else None
        if doc is None:
            raise HTTPException(404, "unknown document")
        state = "UNLOCKING" if request and request["status"] == "REQUESTED" else "PROTECTED"
        return {"document_id": document_id, "document_version": version,
                "drm_status": state, "error": None}

    @app.post("/api/documents/{document_id}/drm-unlock", status_code=202)
    def viewer_request_unlock(document_id: str, req: ViewerUnlockReq):
        """Submit to the existing authorized unlock-request workflow only."""
        with wlock():
            doc = store.conn.execute(
                """SELECT d.filename FROM document d JOIN document_version v
                     ON v.document_id=d.document_id
                    WHERE d.document_id=? AND v.version_id=?""",
                (document_id, req.document_version)).fetchone()
            if doc is None:
                raise HTTPException(404, "unknown document version")
            try:
                request = create_request(store, root / "data" / "raw",
                                         doc["filename"], req.note)
            except (FileNotFoundError, ValueError) as exc:
                raise HTTPException(422, str(exc)) from exc
            store.conn.execute(
                """INSERT INTO viewer_document_version
                     (document_id,document_version,sha256,unlocked_path,drm_status,
                      drm_error,render_status,render_error,sheet_count,registered_at,rendered_at)
                   VALUES (?,?,NULL,NULL,'UNLOCKING',NULL,'PENDING',NULL,NULL,?,NULL)
                   ON CONFLICT(document_id,document_version) DO UPDATE SET
                     drm_status='UNLOCKING',drm_error=NULL""",
                (document_id, req.document_version, now_iso()))
            store.commit()
        return {"document_id": document_id, "document_version": req.document_version,
                "drm_status": "UNLOCKING", "request_id": request["request_id"]}

    @app.post("/api/viewer/documents/{document_id}/register-unlocked", status_code=201)
    def viewer_register(document_id: str, req: ViewerRegisterReq):
        """Register output from an authorized external DRM service.

        This endpoint performs no unlock/decryption. It accepts only a basename
        already delivered into the configured staging directory.
        """
        from kg.viewer import ViewerError, mark_validation_failed, register_unlocked
        if Path(req.staging_name).name != req.staging_name or not req.staging_name:
            raise HTTPException(422, "invalid staging filename")
        try:
            with wlock():
                result = register_unlocked(
                    store, document_id, req.document_version,
                    root / "data" / "unlocked-staging" / req.staging_name,
                    root / "data" / "unlocked-staging", root / "data" / "unlocked")
                store.commit()
            return result
        except ViewerError as exc:
            with wlock():
                mark_validation_failed(store, document_id, req.document_version, str(exc))
                store.commit()
            raise HTTPException(422, str(exc)) from exc

    # ---- Viewer API: DRM 파일은 tree_node + sheet_render 사용 ----

    @app.get("/api/viewer/documents/{document_id}/versions/first")
    def viewer_first_version(document_id: str):
        with lock:
            vdv = store.conn.execute(
                "SELECT document_version FROM viewer_document_version WHERE document_id=? LIMIT 1",
                (document_id,)).fetchone()
            if vdv is None:
                dv = store.conn.execute(
                    "SELECT version_id FROM document_version WHERE document_id=? LIMIT 1",
                    (document_id,)).fetchone()
                if dv is None:
                    raise HTTPException(404, "no version found")
                return {"document_id": document_id, "document_version": dv["version_id"]}
            return {"document_id": document_id, "document_version": vdv["document_version"]}

    @app.get("/api/viewer/documents/{document_id}/versions/{document_version}")
    def viewer_document(document_id: str, document_version: str):
        with lock:
            doc = store.conn.execute(
                "SELECT * FROM document WHERE document_id=?", (document_id,)).fetchone()
            if doc is None:
                raise HTTPException(404, "unknown document")
            vdv = store.conn.execute(
                "SELECT * FROM viewer_document_version WHERE document_id=? AND document_version=?",
                (document_id, document_version)).fetchone()
        result = {"document_id": document_id, "document_version": document_version,
                  "filename": doc["filename"],
                  "drm_status": vdv["drm_status"] if vdv else "PROTECTED",
                  "render_status": vdv["render_status"] if vdv else "PENDING"}
        if vdv:
            result["sheet_count"] = vdv["sheet_count"]
        return result

    @app.get("/api/viewer/documents/{document_id}/sheets")
    def viewer_sheets(document_id: str, document_version: str = "",
                      include_hidden: bool = True):
        with lock:
            # tree_node(SHEET)에서 시트 목록
            rows = store.conn.execute(
                """SELECT node_name, rowid FROM tree_node
                   WHERE document_id=? AND status='ACTIVE' AND node_type='SHEET'
                   ORDER BY tree_path""", (document_id,)).fetchall()
        return [{"sheet_index": i, "sheet_name": r["node_name"],
                 "state": "visible"} for i, r in enumerate(rows)]

    @app.post("/api/viewer/documents/{document_id}/render")
    def viewer_render(document_id: str, req: ViewerRenderReq):
        # DRM 파일은 이미 sheet_render에 캐시됨 → 항상 SUCCESS
        # 비DRM은 render_preview 시도 (LibreOffice) — 스테이징: LibreOffice
        # 서브프로세스(최대 120초)는 절대 전역 DB lock을 쥔 채 돌리지 않는다
        from kg.viewer import (ViewerError, execute_render, finalize_render,
                               mark_render_failed, prepare_render)
        with lock:
            doc_row = store.conn.execute(
                "SELECT filepath FROM document WHERE document_id=?",
                (document_id,)).fetchone()
        if not (doc_row and doc_row['filepath']
                and not _is_drm_file(Path(doc_row['filepath']))):
            return {"status": "SUCCESS", "cached": True}
        cache_root = root / "data" / "viewer-cache"
        try:
            with wlock():
                prep = prepare_render(store, document_id, req.document_version,
                                      cache_root)
                store.commit()
        except ViewerError as exc:
            raise HTTPException(503, str(exc)) from exc
        if prep["cached"]:
            return {"status": "SUCCESS", "cached": True}
        try:
            execute_render(prep)                     # lock 밖 — LibreOffice
        except ViewerError as exc:
            with wlock():
                mark_render_failed(store, document_id, req.document_version, str(exc))
                store.commit()
            raise HTTPException(503, str(exc)) from exc
        with wlock():
            finalize_render(store, document_id, req.document_version, prep)
            store.commit()
        return {"status": "SUCCESS"}

    @app.get("/api/viewer/documents/{document_id}/render-status")
    def viewer_render_status(document_id: str, document_version: str):
        # 등록된 뷰어 소스는 실제 상태를 보고, 미등록(sheet_render 경로)은
        # 기존 동작대로 SUCCESS 스텁을 유지한다
        from kg.viewer import ViewerError, document_metadata
        try:
            with lock:
                metadata = document_metadata(store, document_id, document_version)
            return {key: metadata[key] for key in
                    ("document_id", "document_version", "render_status",
                     "render_error")}
        except ViewerError:
            return {"document_id": document_id, "document_version": document_version,
                    "render_status": "SUCCESS", "render_error": None}

    @app.get("/api/viewer/documents/{document_id}/preview")
    @app.get("/api/viewer/documents/{document_id}/preview")
    def viewer_preview(document_id: str, document_version: str = "", sheet: str | None = None):
        """DRM 파일: sheet_render 캐시에서 JSON 반환. CSV: tree_node에서 테이블 생성."""
        with lock:
            # 시트 목록에서 첫 시트 또는 지정 시트
            if not sheet:
                s = store.conn.execute(
                    """SELECT node_name FROM tree_node
                       WHERE document_id=? AND status='ACTIVE' AND node_type='SHEET'
                       ORDER BY tree_path LIMIT 1""", (document_id,)).fetchone()
                if s is None:
                    raise HTTPException(404, "no sheets")
                sheet = s["node_name"]
            row = store.load_render(document_id, sheet)
        if row and row["render_json"]:
            import json as _json
            return _json.loads(row["render_json"])
        # 캐시 없음: tree_node 기반 테이블
        result = _tree_node_table(document_id, sheet)
        if result:
            return result
        raise HTTPException(404, f"no data for {sheet}")

    @app.get("/api/viewer/documents/{document_id}/source")
    def viewer_source(document_id: str, document_version: str, sheet: str,
                      a1_range: str, concept_id: str | None = None):
        with lock:
            # a1_range로 tree_node 찾기
            node = store.conn.execute(
                """SELECT n.*, m.concept_id, m.status as mapping_status
                   FROM tree_node n
                   LEFT JOIN semantic_mapping m ON m.tree_node_id=n.node_id AND m.is_active=1
                   WHERE n.document_id=? AND n.status='ACTIVE'
                     AND n.locator=?
                   LIMIT 1""", (document_id, f"{sheet}!{a1_range}")).fetchone()
            if node is None:
                # locator가 없으면 node_name으로 폴백
                node = store.conn.execute(
                    """SELECT n.*, m.concept_id, m.status as mapping_status
                       FROM tree_node n
                       LEFT JOIN semantic_mapping m ON m.tree_node_id=n.node_id AND m.is_active=1
                       WHERE n.document_id=? AND n.status='ACTIVE'
                         AND n.node_name=? AND n.tree_path LIKE ?
                       LIMIT 1""", (document_id, a1_range, f"%/{sheet}/%")).fetchone()
        if node is None:
            return {"document_id": document_id, "document_version": document_version,
                    "sheet": sheet, "a1_range": a1_range, "concept_id": concept_id}
        return {"document_id": document_id, "document_version": document_version,
                "sheet": sheet, "a1_range": a1_range,
                "concept_id": node["concept_id"] or concept_id,
                "mapping_source": "CONCEPT_HINT" if node["mapping_status"] == "AUTO_APPROVED" else "TEMPLATE"}

    @app.get("/api/search")
    def search(concept: str):
        with lock:
            cid = concept
            if store.concept(cid) is None:            # 이름/동의어로도 검색
                row = store.conn.execute(
                    """SELECT a.concept_id FROM domain_alias a
                       JOIN domain_concept c ON c.concept_id=a.concept_id
                       WHERE a.alias_norm=? AND c.status='ACTIVE' LIMIT 1""",
                    (normalize_label(concept),)).fetchone()
                if row is None:
                    raise HTTPException(404, f"unknown concept: {concept}")
                cid = row["concept_id"]
            res = reverse_lookup(store, cid, include_review=True)
            # 뷰어 점프/Inspector에 필요한 document_id·node_id 부착
            for s in res["sources"]:
                n = store.node(s["node_id"])
                s.pop("payload_id", None)
                s["document_id"] = n["document_id"] if n else None
            # 지식 그래프 1-hop 이웃 (이웃별 연결 소스 수 포함 — 탐색 단서)
            neighbors = concept_neighbors(store, cid)
            counts = {r["concept_id"]: r["n"] for r in store.conn.execute(
                """SELECT concept_id, count(*) n FROM semantic_mapping
                   WHERE is_active=1 AND status IN ('AUTO_APPROVED','APPROVED')
                   GROUP BY concept_id""")}
            for e in neighbors:
                other = e["target_concept_id"] if e["source_concept_id"] == cid \
                    else e["source_concept_id"]
                e["other_sources"] = counts.get(other, 0)
        res["neighbors"] = neighbors
        res["concept"] = {k: res["concept"][k] for k in
                          ("concept_id", "canonical_name", "description",
                           "canonical_unit") if k in res["concept"].keys()}
        return res

    @app.get("/api/review")
    def review_queue(limit: int = 50, doc: str | None = None):
        """검수 큐 (§9 Warning Badge) — doc 지정 시 그 파일의 순차 검수 목록."""
        where = "m.status='REVIEW_REQUIRED' AND m.is_active=1 AND n.status='ACTIVE'"
        params: list = []
        if doc:
            where += " AND n.document_id=?"
            params.append(doc)
        params.append(limit)
        with lock:
            rows = store.conn.execute(
                f"""SELECT m.mapping_id, m.concept_id, m.confidence, n.node_id,
                           n.node_name, n.locator, n.document_id, d.filename,
                           e.reason
                    FROM semantic_mapping m
                    JOIN tree_node n ON n.node_id=m.tree_node_id
                    JOIN document d ON d.document_id=n.document_id
                    LEFT JOIN mapping_evidence e ON e.mapping_id=m.mapping_id
                    WHERE {where}
                    ORDER BY m.confidence DESC LIMIT ?""", params).fetchall()
        return [dict(r) for r in rows]

    render_cache: dict[tuple, dict] = {}

    @app.get("/api/sheet")
    def sheet(doc: str, name: str | None = None):
        """원본 충실 렌더 우선순위:

        1) 읽을 수 있는 원본 xlsx(비DRM) → _render_sheet (항상 최신·최충실)
        2) DB 렌더 캐시 — Windows 쪽에서 사전 생성한 DRM 고충실 렌더 서빙용
        3) DRM 파일 + Excel COM 가능 → _render_sheet_drm
           (시트 복사→충실 렌더; 복사 차단 환경만 값-전용 + degraded)
        4) tree_node 폴백 (구조 트리 기반 표 — degraded 명시)
        """
        # sheets 목록은 항상 tree_node에서 (render 캐시 유무 무관)
        all_sheets = [r["node_name"] for r in store.conn.execute(
            """SELECT node_name FROM tree_node
               WHERE document_id=? AND status='ACTIVE' AND node_type='SHEET'
               ORDER BY tree_path""", (doc,)).fetchall()]

        def _with_meta(render: dict) -> dict:
            with lock:
                meta = store.conn.execute(
                    """SELECT d.current_version,v.drm_status,v.render_status
                         FROM document d LEFT JOIN viewer_document_version v
                           ON v.document_id=d.document_id
                          AND v.document_version=d.current_version
                        WHERE d.document_id=?""", (doc,)).fetchone()
            return {"document_id": doc,
                    "document_version": meta["current_version"] if meta else None,
                    "viewer": ({"drm_status": meta["drm_status"],
                                "render_status": meta["render_status"]} if meta and
                               meta["drm_status"] else None),
                    **render}

        path = None
        try:
            path = _doc_path(doc)
        except HTTPException:
            path = None                       # 파일 소실 → 캐시/tree 폴백
        is_drm = path is not None and _is_drm_file(path)
        renderable = (path is not None and path.suffix.lower() == ".xlsx"
                      and not is_drm)

        # 1) 읽을 수 있는 원본 xlsx — 원본 충실 렌더가 언제나 최우선
        if renderable:
            key = (str(path), path.stat().st_mtime, name or "")
            if key not in render_cache:
                if len(render_cache) > 24:
                    render_cache.clear()
                render_cache[key] = _render_sheet(path, name)
            return _with_meta(render_cache[key])

        # 2) DB 렌더 캐시 (Windows에서 사전 생성한 DRM 충실 렌더)
        target_sheet = name or (all_sheets[0] if all_sheets else None)
        if target_sheet:
            row = store.load_render(doc, target_sheet)
            if row and row["render_json"]:
                import json as _json
                cached = _json.loads(row["render_json"])
                cached["sheets"] = all_sheets
                return {"document_id": doc, **cached}

        # 3) DRM 파일 — Excel COM으로 충실 렌더 시도 (Windows + Excel 환경)
        if is_drm:
            key = (str(path), path.stat().st_mtime, name or "")
            if key not in render_cache:
                try:
                    if len(render_cache) > 24:
                        render_cache.clear()
                    render_cache[key] = _render_sheet_drm(path, name)
                except HTTPException:
                    raise
                except Exception:
                    key = None                # COM 불가(리눅스 등) → tree 폴백
            if key is not None:
                return _with_meta(render_cache[key])

        # 4) tree_node 폴백 — 구조 트리 기반 표 (저하 렌더임을 명시)
        if target_sheet:
            tn_result = _tree_node_table(doc, target_sheet)
            if tn_result:
                # JS 렌더 계약(cols/rows 배열)으로 정규화
                if "cols" not in tn_result:
                    tn_result["cols"] = [120] * tn_result.get("max_col", 0)
                if "rows" not in tn_result:
                    tn_result["rows"] = [24] * tn_result.get("max_row", 0)
                tn_result.setdefault("images", [])
                tn_result.setdefault("gridlines", True)
                tn_result["sheets"] = all_sheets
                tn_result.setdefault(
                    "degraded",
                    "원본 파일을 열 수 없어 구조 트리 기반 표만 표시합니다 — "
                    "병합/서식 없음"
                    + (" (DRM 잠김 — 해제본 등록 시 정상 표시)" if is_drm else ""))
                return {"document_id": doc, **tn_result}
        raise HTTPException(404, "file missing and no tree fallback")

    # ---------------------------------------------- KG View Models (§8 v3) ----
    # DKG 파생은 kg/groups.py로 이관 — 사람 델타(INCLUDED/EXCLUDED) 반영 포함.
    def _document_kgs():
        return document_kgs(store)

    def _attach_group_meta(g: dict) -> dict:
        """DKG에 레시피, Parsing Template, 최근 실행 상태를 붙인다."""
        from kg.parsing import grouped_documents
        rec = active_recipe(store, g["id"])
        if rec is not None:
            spec = json.loads(rec["spec_json"])
            stale = sum(1 for e in spec["template"]
                        if (store.concept(e["concept_id"]) or
                            {"status": None})["status"] != "ACTIVE")
            g["recipe"] = {
                "recipe_id": rec["recipe_id"], "created_at": rec["created_at"],
                "note": rec["note"] or "",
                "template": len(spec["template"]),
                "conflicts": len(spec.get("conflicts") or []),
                "dropped": len(spec.get("dropped") or []),
                "stale_entries": stale,
            }
        else:
            g["recipe"] = None
        run = store.conn.execute(
            "SELECT run_id, mode, status, started_at, finished_at "
            "FROM recrawl_run WHERE root_concept_id=? "
            "ORDER BY started_at DESC LIMIT 1", (g["id"],)).fetchone()
        g["last_recrawl"] = dict(run) if run else None
        # Parsing Template is an operational layer below Document KG, not a KG
        # node. Keep the established graph view and expose it as grouped detail.
        g["parsing_templates"] = grouped_documents(store, g["id"])
        return g

    @app.get("/api/kg/domain")
    def kg_domain():
        """전체 Domain KG Snapshot (§8): 노드(레벨/그룹/소스수) + IS_A 엣지."""
        with lock:
            roots, levels, parents = isa_roots(store)
            src_counts = {r["concept_id"]: r["n"] for r in store.conn.execute(
                """SELECT concept_id, count(*) n FROM semantic_mapping
                   WHERE is_active=1 AND status IN ('AUTO_APPROVED','APPROVED')
                   GROUP BY concept_id""")}
            nodes = [{
                "id": r["concept_id"], "name": r["canonical_name"],
                "level": r["domain_level"], "root": roots.get(r["concept_id"]),
                "parent": parents.get(r["concept_id"]),
                "sources": src_counts.get(r["concept_id"], 0),
            } for r in store.concepts()]
        return {"nodes": nodes, "domain": root.name,
                "edges": [{"s": s, "t": t} for s, t in parents.items()]}

    @app.get("/api/kg/document")
    def kg_document_list():
        with lock:
            kgs = [_attach_group_meta(g) for g in _document_kgs()]
        for g in kgs:                      # 목록은 요약만 (§7.3 데이터 많은 경우)
            g["member_document_ids"] = [d["document_id"] for d in g["member_documents"]]
            g["member_documents"] = g["member_documents"][:4]
        return kgs

    @app.get("/api/kg/document/{dkg_id}")
    def kg_document_detail(dkg_id: str):
        with lock:
            kgs = {g["id"]: g for g in _document_kgs()}
            if dkg_id in kgs:
                _attach_group_meta(kgs[dkg_id])
                overrides = {d: s for (rt, d), s in member_overrides(store).items()
                             if rt == dkg_id}
        if dkg_id not in kgs:
            raise HTTPException(404, "unknown document kg")
        g = kgs[dkg_id]
        for d in g["member_documents"]:
            d["override"] = overrides.get(d["document_id"])
        return g

    @app.get("/api/files")
    def files():
        """S01 파일 분석: 파일별 Ready/Review 상태·매핑률·검토 건수 (§3, §6.1)."""
        with lock:
            rows = store.conn.execute(
                """SELECT d.document_id, d.filename,
                     (SELECT count(DISTINCT substr(n.tree_path,
                          instr(n.tree_path,'/')+1,
                          CASE WHEN instr(substr(n.tree_path, instr(n.tree_path,'/')+1),'/')=0
                               THEN length(n.tree_path)
                               ELSE instr(substr(n.tree_path, instr(n.tree_path,'/')+1),'/')-1 END))
                      FROM tree_node n WHERE n.document_id=d.document_id
                        AND n.status='ACTIVE' AND n.node_type='SHEET') sheets,
                     count(h.node_id) headers,
                     sum(CASE WHEN m.status IN ('AUTO_APPROVED','APPROVED') THEN 1 ELSE 0 END) mapped,
                     sum(CASE WHEN m.status='REVIEW_REQUIRED' THEN 1 ELSE 0 END) review,
                     v.drm_status, v.render_status, a.status parsing_status,
                     t.name template_name, a.template_version
                   FROM document d
                   LEFT JOIN tree_node h ON h.document_id=d.document_id
                        AND h.status='ACTIVE' AND h.node_type='HEADER'
                   LEFT JOIN semantic_mapping m ON m.tree_node_id=h.node_id AND m.is_active=1
                   LEFT JOIN viewer_document_version v ON v.document_id=d.document_id
                        AND v.document_version=d.current_version
                   LEFT JOIN document_template_assignment a ON a.document_id=d.document_id
                        AND a.document_version=d.current_version
                   LEFT JOIN parsing_template t ON t.template_id=a.template_id
                   GROUP BY d.document_id ORDER BY d.filename""").fetchall()
        out = []
        for r in rows:
            headers = r["headers"] or 0
            mapped = r["mapped"] or 0
            review = r["review"] or 0
            status = "ERROR" if headers == 0 else \
                ("REVIEW_REQUIRED" if review > 0 else "READY")
            out.append({
                "document_id": r["document_id"], "filename": r["filename"],
                "sheets": r["sheets"] or 0, "headers": headers,
                "coverage_pct": round(100 * mapped / headers, 1) if headers else 0,
                "review": review, "status": status})
            out[-1].update({"drm_status": r["drm_status"] or "PROTECTED",
                            "render_status": r["render_status"],
                            "parsing_status": r["parsing_status"],
                            "template_name": r["template_name"],
                            "template_version": r["template_version"]})
            # 작성자/작성일 — 검색·필터·정렬용 (원본 파일의 core 속성)
            out[-1].update(_doc_props(root / "data" / "raw" / r["filename"]))
        return out

    @app.get("/api/overlay")
    def overlay(doc: str, name: str):
        """활성 시트의 Semantic Overlay (§4.2): 매핑된 영역의 role/개념/범위."""
        with lock:
            rows = store.conn.execute(
                """SELECT n.node_id, n.node_name, n.locator, n.tree_path,
                          n.data_type, n.metadata, n.node_type,
                          m.concept_id, m.confidence, m.status,
                          c.canonical_name, c.concept_type
                   FROM tree_node n
                   JOIN semantic_mapping m ON m.tree_node_id=n.node_id AND m.is_active=1
                   LEFT JOIN domain_concept c ON c.concept_id=m.concept_id
                   WHERE n.document_id=? AND n.status='ACTIVE'
                     AND n.node_type IN ('HEADER','VALUE')
                     AND m.status IN ('AUTO_APPROVED','APPROVED','REVIEW_REQUIRED')""", (doc,)).fetchall()
        out = []
        for r in rows:
            loc = r["locator"] or ""
            # 시트 필터링: locator 또는 tree_path에서 시트명 추출
            loc_sheet = loc.rsplit("!", 1)[0] if "!" in loc else ""
            if not loc_sheet and r["tree_path"]:
                parts = r["tree_path"].split("/")
                if len(parts) >= 3:
                    loc_sheet = parts[2]
            if loc_sheet != name:
                continue
            rng = loc.rsplit("!", 1)[-1] if "!" in loc else ""
            meta = json.loads(r["metadata"] or "{}")
            role = "IGNORE" if r["status"] == "UNMAPPED" else \
                _node_role(meta, r["data_type"], r["concept_type"])
            out.append({
                "node_id": r["node_id"], "header": r["node_name"], "range": rng,
                "role": role,
                "concept_id": r["concept_id"], "concept_name": r["canonical_name"],
                "confidence": round(r["confidence"], 2), "status": r["status"]})
        return out

    @app.get("/api/source/{node_id}")
    def source_detail(node_id: str):
        """S03 Inspector (§4.1 우측): 영역/역할/개념/값 Preview/Row Context."""
        with lock:
            n = store.node(node_id)
            if n is None:
                raise HTTPException(404, "unknown node")
            meta = json.loads(n["metadata"] or "{}")
            m = store.active_mapping(node_id)
            concept = store.concept(m["concept_id"]) if m and m["concept_id"] else None
            ev = store.conn.execute(
                "SELECT candidates_json, reason FROM mapping_evidence "
                "WHERE mapping_id=?",
                (m["mapping_id"],)).fetchone() if m else None
            pv = store.conn.execute(
                """SELECT pv.row_key, pv.value_num, pv.value_text, pv.cell_address
                   FROM payload_value pv JOIN data_payload p ON p.payload_id=pv.payload_id
                   WHERE p.tree_node_id=? AND p.is_current=1
                   ORDER BY pv.row_idx LIMIT 8""", (node_id,)).fetchall()
            doc = store.conn.execute(
                "SELECT filename,current_version FROM document WHERE document_id=?",
                (n["document_id"],)).fetchone()
            viewer = store.conn.execute(
                """SELECT drm_status,render_status,render_error,sha256
                     FROM viewer_document_version
                    WHERE document_id=? AND document_version=?""",
                (n["document_id"], doc["current_version"] if doc else None)).fetchone()
            assignment = store.conn.execute(
                """SELECT a.template_id,a.template_version,a.status,t.name template_name
                     FROM document_template_assignment a JOIN parsing_template t
                       ON t.template_id=a.template_id
                    WHERE a.document_id=? AND a.document_version=?""",
                (n["document_id"], doc["current_version"] if doc else None)).fetchone()
            parsing_source = None
            if assignment:
                from kg.parsing import effective_mappings
                node_range = (n["locator"] or "").rsplit("!", 1)[-1]
                node_sheet = (n["locator"] or "").rsplit("!", 1)[0]
                candidates = effective_mappings(store, doc["current_version"])
                match = next((x for x in candidates
                              if (not m or not m["concept_id"] or
                                  x["concept_id"] == m["concept_id"])
                              and x["effective_source"].get("range") == node_range
                              and x["effective_source"].get("sheet", node_sheet) == node_sheet), None)
                if match:
                    parsing_source = {
                        "mapping_key": match["mapping_key"],
                        "mapping_source": match["mapping_source"],
                        "template_source": match["template_source"],
                        "effective_source": match["effective_source"],
                        "override_status": match.get("override_status"),
                        "override_reason": match.get("override_reason"),
                    }
        parts = (n["tree_path"] or "").split("/")
        return {
            "node_id": node_id, "header": n["node_name"],
            "document_id": n["document_id"],
            "document_version": doc["current_version"] if doc else None,
            "document": doc["filename"] if doc else "",
            "sheet": parts[1] if len(parts) > 1 else "",
            "range": (n["locator"] or "").rsplit("!", 1)[-1],
            "unit": n["unit"], "data_type": n["data_type"],
            "role": _node_role(meta, n["data_type"],
                               concept["concept_type"] if concept else None),
            "mapping": {
                "mapping_id": m["mapping_id"], "concept_id": m["concept_id"],
                "confidence": round(m["confidence"], 2), "status": m["status"],
                "method": m["method"],
                "reason": (ev["reason"] if ev else None),
            } if m else None,
            "concept_name": concept["canonical_name"] if concept else None,
            "viewer": dict(viewer) if viewer else None,
            "parsing_template": dict(assignment) if assignment else None,
            "parsing_source": parsing_source,
            "candidates": json.loads(ev["candidates_json"])[:5] if ev else [],
            "row_context": {
                "keys": [h for h in (meta.get("adjacent_headers") or [])][:4],
                "header_path": meta.get("header_path") or [],
            },
            "values": [{
                "key": p["row_key"],
                "value": p["value_num"] if p["value_num"] is not None else p["value_text"],
                "cell": p["cell_address"]} for p in pv],
        }

    @app.post("/api/proposal")
    def proposal(body: ProposalReq):
        """S04: 선택 노드 묶음 → Row Context 기반 통합 스키마 제안 (§5.1)."""
        with lock:
            by_concept: dict[str, dict] = {}
            docs: set[str] = set()
            stale: list[str] = []
            for nid in body.node_ids[:500]:
                n = store.node(nid)
                m = store.active_mapping(nid) if n else None
                if n is None or n["status"] != "ACTIVE" or m is None \
                        or not m["concept_id"]:
                    stale.append(nid)   # 재적재로 사라졌거나 매핑 해제된 위치
                    continue
                c = store.concept(m["concept_id"])
                if c is None:
                    stale.append(nid)
                    continue
                usable = m["status"] in ("AUTO_APPROVED", "APPROVED")
                g = by_concept.setdefault(m["concept_id"], {
                    "concept_id": c["concept_id"], "concept_name": c["canonical_name"],
                    "field_name": re.sub(r"[^A-Za-z0-9_]", "_",
                                         (c["canonical_name_en"] or c["concept_id"])
                                         .strip().lower().replace(" ", "_")) or c["concept_id"],
                    "target_unit": c["canonical_unit"], "type": c["data_type"],
                    "units": set(), "sources": 0, "review": 0, "node_ids": [],
                    "role": None})
                if usable:
                    # build와 같은 자격 기준 — REVIEW_REQUIRED는 승인 전까지 제외
                    g["sources"] += 1
                    g["node_ids"].append(nid)
                else:
                    g["review"] += 1
                if n["unit"]:
                    g["units"].add(n["unit"])
                meta = json.loads(n["metadata"] or "{}")
                g["role"] = g["role"] or _node_role(meta, n["data_type"], c["concept_type"])
                doc = store.conn.execute(
                    "SELECT filename FROM document WHERE document_id=?",
                    (n["document_id"],)).fetchone()
                if doc:
                    docs.add(doc["filename"])
        fields = []
        for g in by_concept.values():
            units = sorted(g.pop("units"))
            tgt = g["target_unit"]
            note = (f"{'/'.join(units)} → {tgt}" if tgt and units and
                    (len(units) > 1 or units[0] != tgt) else
                    (f"{tgt} 통일" if tgt else "타입 정규화"))
            if g["review"] and not g["sources"]:
                status, note = "검토", f"검토 대기 {g['review']}건 — 승인 후 포함됩니다"
            elif g["review"]:
                status = "검토"
                note += f" · 검토 대기 {g['review']}건 제외"
            elif len(units) > 1 and not tgt:
                status = "검토"
            else:
                status = "정상"
            fields.append({**g, "units": units, "note": note, "status": status})
        return {"fields": fields, "documents": sorted(docs), "stale_node_ids": stale}

    # ------------------------------------------------------------- write ----
    @app.post("/api/review")
    def review_act(body: ReviewAction):
        if body.action not in ("approve", "reject"):
            raise HTTPException(400, "action must be approve|reject")
        if not re.match(r"^MAP-[0-9a-f]{12}$", body.mapping_id):
            raise HTTPException(400, "bad mapping id")
        with wlock():
            row = store.conn.execute(
                "SELECT 1 FROM semantic_mapping WHERE mapping_id=? AND is_active=1",
                (body.mapping_id,)).fetchone()
            if row is None:
                raise HTTPException(404, "unknown mapping")
            store.review(body.mapping_id, body.action.upper(), "web")
            store.commit()
        return JSONResponse({"ok": True})

    @app.post("/api/remap")
    def remap(body: RemapAction):
        """Inspector '매핑 수정' (§7.4): 사람이 개념을 확정 — APPROVED 매핑 생성."""
        with wlock():
            if store.node(body.node_id) is None:
                raise HTTPException(404, "unknown node")
            if store.concept(body.concept_id) is None:
                raise HTTPException(404, "unknown concept")
            m = store.active_mapping(body.node_id)
            if m is not None:
                store.deactivate_mapping(m["mapping_id"], action="REMAP",
                                         note=f"web → {body.concept_id}")
            store.save_mapping(body.node_id, body.concept_id, 1.0, "human", "APPROVED",
                               context={}, candidates=[], reason="web remap")
            store.commit()
        return JSONResponse({"ok": True})

    @app.post("/api/build")
    def build_db(body: BuildReq):
        """S04→S05: 선택 묶음으로 통합 DB 생성, Schema/Lineage/Report 반환 (§5.3, §9)."""
        from src.units.converter import UnitRegistry

        from kg.integration.builder import build as run_build, define_project
        units_path = root / "config" / "units.yaml"
        units = UnitRegistry.load(units_path) if units_path.exists() else None
        config = {
            "name": body.name,
            "fields": [{"name": f.name, "concept": f.concept, "unit": f.unit,
                        "type": f.type} for f in body.fields],
            "sources": {"include_nodes": body.include_nodes or {}},
            "transform": [
                {"op": "unit_convert",
                 "config": {"skip_nodes": body.raw_node_ids or []}},
                {"op": "union"},
                {"op": "deduplicate"},
            ],
        }
        with wlock():
            from kg.integration.builder import delete_project
            iid = None
            try:
                iid = define_project(store, config)
                result = run_build(store, iid, root / "data" / "kg" / "builds",
                                   units=units)
            except (ValueError, KeyError) as e:
                if iid is not None:
                    delete_project(store, iid)   # 유령 프로젝트 버전 방지 (보상 삭제)
                raise HTTPException(400, str(e))
            # 결과 미리보기 + Schema/Lineage manifest (§9)
            import sqlite3 as _sq
            con = _sq.connect(result["output_db"])
            con.row_factory = _sq.Row
            try:
                actual_cols = {r[1] for r in con.execute(
                    f'PRAGMA table_info("{body.name}")')}
                preview = [dict(r) for r in con.execute(
                    f'SELECT * FROM "{body.name}" LIMIT 5')]
                lineage_docs = con.execute(
                    f'SELECT count(DISTINCT _source_document_id) FROM "{body.name}"'
                ).fetchone()[0]
            finally:
                con.close()
        # Schema Manifest는 산출물의 진실을 말한다 — 소스가 없어 탈락한 필드는
        # included=false + warning으로 드러낸다 (§9.2 Warning은 차단과 분리)
        warnings = list(result.get("warnings") or [])
        schema = []
        for f in body.fields:
            included = f.name in actual_cols
            schema.append({"field": f.name, "concept": f.concept, "unit": f.unit,
                           "type": f.type, "included": included})
            if not included:
                warnings.append({"op": "source_select", "field": f.name,
                                 "reason": "사용 가능한 소스가 없어 결과에서 제외됨 "
                                           "(검토 대기/매핑 해제 여부 확인)"})
        status = "COMPLETED_WITH_WARNINGS" if warnings else "COMPLETED"
        return {
            "status": status, "build_id": result["build_id"],
            "artifact": result["output_db"], "table": result["table"],
            "row_count": result["rows"],
            "schema": schema,
            "lineage": {"edges": result["lineage_edges"], "documents": lineage_docs},
            "build_report": {"frames": result["frames"], "warnings": warnings},
            "preview": preview,
        }

    # ==================================================== KG2: DKG/레시피 ----
    def _require_l1(root_id: str) -> None:
        """lock 안에서 호출 — DKG 식별자는 L1 개념이어야 한다."""
        if not _CID_RE.match(root_id):
            raise HTTPException(400, "bad concept id")
        if not is_l1_concept(store, root_id):
            raise HTTPException(400, "DKG는 L1 개념이어야 합니다")

    @app.get("/api/raw-files")
    def raw_files():
        """data/raw의 미등록 xlsx — 등록 후보 + 잠금(DRM/암호화)·해제 요청 상태."""
        raw_dir = root / "data" / "raw"
        paths = sorted(p for p in raw_dir.glob("*.xlsx")
                       if not p.name.startswith("~$")) if raw_dir.exists() else []
        out = []
        with wlock():
            refresh_release_states(store, raw_dir)   # 해제본 도착 자동 감지
            known = {r["document_id"] for r in store.conn.execute(
                "SELECT document_id FROM document")}
            for p in paths:
                if document_id_for(root, p) in known:
                    continue
                sniff = sniff_container(p)
                row = request_row(store, p.name)
                out.append({
                    "filename": p.name,
                    "document_id": document_id_for(root, p),
                    "locked": sniff["locked"],
                    "com_readable": sniff.get("com_readable", False),
                    "container": sniff["container"],
                    "container_detail": sniff.get("detail"),
                    "drm": ({"request_id": row["request_id"],
                             "status": row["status"],
                             "requested_at": row["requested_at"],
                             "released_at": row["released_at"],
                             "note": row["note"] or ""}
                            if row is not None else None),
                })
        return out

    @app.post("/api/drm/request")
    def drm_request(body: DrmReq):
        """잠긴 파일의 정식 해제 요청 기록 + 결재용 요청서 텍스트 생성."""
        fn = body.filename
        if "/" in fn or "\\" in fn or ".." in fn or not fn.lower().endswith(".xlsx"):
            raise HTTPException(400, "bad filename")
        with wlock():
            try:
                res = create_request(store, root / "data" / "raw", fn,
                                     note=(body.note or "")[:2000])
            except FileNotFoundError:
                raise HTTPException(404, f"file not found: {fn}")
            except ValueError as e:
                raise HTTPException(400, str(e))
        return {"ok": True, **res}

    @app.get("/api/drm")
    def drm_list():
        with wlock():
            refresh_release_states(store, root / "data" / "raw")
            return list_requests(store)

    @app.post("/api/ingest")
    def ingest_document(body: IngestReq):
        """단건 등록: parse(lock 밖) → apply → (그룹 지정 시) INCLUDED+레시피
        → 잔여 judge. 그룹 미지정 시 '같은 형식' DKG 후보를 제안한다."""
        from src.inspect.inspector import PARSER_VERSION
        fn = body.filename
        if ("/" in fn or "\\" in fn or ".." in fn or fn.startswith("~$")
                or not fn.lower().endswith(".xlsx")):
            raise HTTPException(400, "bad filename")
        raw_dir = (root / "data" / "raw").resolve()
        path = (raw_dir / fn).resolve()
        if path.parent != raw_dir:
            raise HTTPException(400, "bad filename")
        if not path.exists():
            raise HTTPException(404, f"file not found: {fn}")
        sniff = sniff_container(path)
        if sniff["locked"] and not sniff.get("com_readable"):
            raise HTTPException(
                400, f"파일이 잠겨 있습니다({sniff.get('detail') or sniff['container']}) "
                     "— DRM 해제 요청 후 해제본이 도착하면 등록할 수 있습니다")
        with lock:
            if body.group_id is not None:
                _require_l1(body.group_id)
            if not store.concepts():
                raise HTTPException(503, "Domain KG가 비어 있습니다 — 먼저 seed를 실행하세요")
        try:
            doc_id, drafts, file_hash = parse_workbook(       # lock 밖 — 파싱
                store, root, path, ws.parser_rules, ws.units, ws.registry)
        except Exception as e:
            raise HTTPException(400, f"파싱 실패: {e}")
        rec_stats = None
        with wlock():
            try:
                ing = apply_parsed(store, doc_id, path, file_hash,
                                   PARSER_VERSION, drafts, force=body.force)
                if body.group_id:
                    set_member_override(store, body.group_id, doc_id, "INCLUDED")
                    store.commit()
                    if body.map:
                        rec = active_recipe(store, body.group_id)
                        if rec is not None:
                            rec_stats = apply_recipe(store, rec, doc_id)
            except HTTPException:
                raise
            except Exception as e:
                store.conn.rollback()
                raise HTTPException(400, str(e))
            retriever = _fresh_retriever() if body.map else None
        map_stats = (map_nodes_staged(store, lock, retriever, _judge(), doc_id)
                     if body.map else None)
        with wlock():
            suggestions = (None if body.group_id
                           else suggest_groups(store, doc_id))
            mark_ingested(store, fn)         # 해제 요청 이력이 있으면 완결로
            store.commit()
            struct_cache.clear()
        return {"ok": True, "document_id": doc_id, "ingest": ing,
                "recipe": rec_stats, "map": map_stats,
                "suggestions": suggestions}

    @app.post("/api/group/{root_id}/member")
    def group_member(root_id: str, body: MemberReq):
        """멤버십 오버라이드 — INCLUDED(핀 고정)/EXCLUDED(파생 부활 차단)."""
        if body.state not in ("INCLUDED", "EXCLUDED"):
            raise HTTPException(400, "state must be INCLUDED|EXCLUDED")
        if not _DOCID_RE.match(body.document_id):
            raise HTTPException(400, "bad document id")
        with wlock():
            _require_l1(root_id)
            if store.conn.execute("SELECT 1 FROM document WHERE document_id=?",
                                  (body.document_id,)).fetchone() is None:
                raise HTTPException(404, "unknown document")
            set_member_override(store, root_id, body.document_id, body.state)
            store.commit()
        return JSONResponse({"ok": True})

    @app.delete("/api/group/{root_id}/member/{document_id}")
    def group_member_clear(root_id: str, document_id: str):
        with wlock():
            _require_l1(root_id)
            n = clear_member_override(store, root_id, document_id)
            store.commit()
        if not n:
            raise HTTPException(404, "no override")
        return JSONResponse({"ok": True})

    @app.post("/api/group/{root_id}/recipe")
    def recipe_snapshot(root_id: str, body: RecipeReq):
        with wlock():
            _require_l1(root_id)
            try:
                res = snapshot_recipe(store, root_id, note=(body.note or "")[:2000],
                                      created_by="web")
            except ValueError as e:
                raise HTTPException(400, str(e))
        return {"ok": True, **res}

    @app.get("/api/group/{root_id}/recipe")
    def recipe_get(root_id: str):
        with lock:
            _require_l1(root_id)
            rec = active_recipe(store, root_id)
            history = [dict(r) for r in store.conn.execute(
                """SELECT recipe_id, status, note, created_at, created_by
                   FROM extraction_recipe WHERE root_concept_id=?
                   ORDER BY created_at DESC LIMIT 20""", (root_id,))]
            spec = json.loads(rec["spec_json"]) if rec is not None else None
            stale = []
            if spec:
                for e in spec["template"]:
                    c = store.concept(e["concept_id"])
                    if c is None or c["status"] != "ACTIVE":
                        stale.append(e["concept_id"])
        if rec is None:
            raise HTTPException(404, "레시피가 없습니다 — 먼저 스냅샷을 생성하세요")
        return {"recipe_id": rec["recipe_id"], "created_at": rec["created_at"],
                "note": rec["note"] or "", "spec": spec,
                "stale_entries": sorted(set(stale)), "history": history}

    @app.get("/api/group/{root_id}/recipe/preview")
    def recipe_preview(root_id: str, document_id: str):
        """dry-run: 이 문서에 활성 레시피가 어떻게 매칭되는지 (쓰기 없음)."""
        if not _DOCID_RE.match(document_id):
            raise HTTPException(400, "bad document id")
        with lock:
            _require_l1(root_id)
            rec = active_recipe(store, root_id)
            if rec is None:
                raise HTTPException(404, "레시피가 없습니다")
            rows = preview_recipe(store, rec, document_id)
        return {"document_id": document_id, "recipe_id": rec["recipe_id"],
                "nodes": rows}

    @app.post("/api/group/{root_id}/recipe/{recipe_id}/rollback")
    def recipe_roll(root_id: str, recipe_id: str):
        if not _RCP_RE.match(recipe_id):
            raise HTTPException(400, "bad recipe id")
        with wlock():
            _require_l1(root_id)
            try:
                rid = rollback_recipe(store, root_id, recipe_id)
            except KeyError:
                raise HTTPException(404, "unknown recipe")
        return {"ok": True, "recipe_id": rid}

    @app.post("/api/group/{root_id}/recrawl")
    def recrawl_group(root_id: str, body: RecrawlReq):
        """멤버 문서 재수집+재매핑 — 백그라운드 실행, run_id로 폴링.
        전역 직렬화: 한 문서가 여러 그룹에 속할 수 있어 그룹별 가드로는 부족."""
        if body.mode not in MODES:
            raise HTTPException(400, "mode must be fill|reset_auto")
        if body.document_ids is not None:
            if not body.document_ids or len(body.document_ids) > 200:
                raise HTTPException(400, "document_ids: 1~200개")
            for d in body.document_ids:
                if not _DOCID_RE.match(d):
                    raise HTTPException(400, "bad document id")
        judge = _judge()                 # 실패 가능 준비물은 busy 설정 전에
        with wlock():
            _require_l1(root_id)
            if recrawl_state["busy"]:
                raise HTTPException(
                    409, f"재크롤링이 이미 실행 중입니다 ({recrawl_state['run_id']})")
            if not store.concepts():
                raise HTTPException(503, "Domain KG가 비어 있습니다")
            members = group_documents(store, root_id)
            docs = body.document_ids or members
            if not docs:
                raise HTTPException(400, "그룹에 문서가 없습니다")
            if set(docs) - set(members):
                raise HTTPException(400, "그룹 멤버가 아닌 문서가 포함됐습니다")
            rec = active_recipe(store, root_id)
            retriever = _fresh_retriever()   # units YAML 파싱 실패 등도 이 앞에서
            run_id = start_run(store, root_id,
                               rec["recipe_id"] if rec else None, body.mode)
            recrawl_state["busy"], recrawl_state["run_id"] = True, run_id

        def _worker():
            try:
                run_recrawl(store, lock, ws, root_id, body.mode, docs, run_id,
                            retriever, judge)
            except Exception as e:                     # 러너 자체 실패 — run 마감
                with lock:
                    store.conn.rollback()
                    store.conn.execute(
                        "UPDATE recrawl_run SET status='FAILED', finished_at=?, "
                        "summary_json=COALESCE(summary_json, ?) WHERE run_id=?",
                        (now_iso(), json.dumps([{"error": repr(e)}]), run_id))
                    store.commit()
            finally:
                with lock:
                    recrawl_state["busy"], recrawl_state["run_id"] = False, None
                    struct_cache.clear()
                    render_cache.clear()

        try:
            threading.Thread(target=_worker, daemon=True,
                             name=f"recrawl-{run_id}").start()
        except Exception:
            # 워커 기동 실패 시 busy/RUNNING이 영구 고착되지 않게 보상한다
            with wlock():
                recrawl_state["busy"], recrawl_state["run_id"] = False, None
                store.conn.execute(
                    "UPDATE recrawl_run SET status='FAILED', finished_at=? "
                    "WHERE run_id=?", (now_iso(), run_id))
                store.commit()
            raise HTTPException(500, "재크롤링 워커 기동 실패")
        return {"ok": True, "run_id": run_id, "documents": len(docs)}

    @app.get("/api/recrawl/{run_id}")
    def recrawl_status(run_id: str):
        if not _RCL_RE.match(run_id):
            raise HTTPException(400, "bad run id")
        with lock:
            row = store.conn.execute(
                "SELECT * FROM recrawl_run WHERE run_id=?", (run_id,)).fetchone()
        if row is None:
            raise HTTPException(404, "unknown run")
        out = dict(row)
        out["summary"] = json.loads(out.pop("summary_json") or "[]")
        return out

    # ==================================================== KG2: Domain 편집 ----
    _CONCEPT_FIELDS = ("canonical_name", "canonical_name_en", "description",
                       "concept_type", "data_type", "domain_level",
                       "canonical_unit", "unit_dimension")

    @app.post("/api/kg/concept")
    def concept_write(body: ConceptReq):
        """생성 또는 부분 수정 — 수정은 읽기-병합-쓰기(미전달 필드 보존)."""
        cid = body.concept_id
        if cid is not None and not _CID_RE.match(cid):
            raise HTTPException(400, "bad concept id")
        if body.domain_level is not None and body.domain_level not in ("L1", "L2", "L3"):
            raise HTTPException(400, "domain_level must be L1|L2|L3")
        with wlock():
            row = store.concept(cid) if cid else None
            if row is None:                                       # 생성
                name = (body.canonical_name or "").strip()
                if not name:
                    raise HTTPException(400, "canonical_name 필수")
                cid = cid or new_id("CONCEPT")
                store.upsert_concept({
                    "concept_id": cid, "canonical_name": name,
                    "canonical_name_en": body.canonical_name_en,
                    "description": body.description,
                    "concept_type": body.concept_type,
                    "data_type": body.data_type,
                    "domain_level": body.domain_level,
                    "canonical_unit": body.canonical_unit,
                    "unit_dimension": body.unit_dimension,
                    "status": "ACTIVE"})
                for a in (name, body.canonical_name_en or ""):
                    if a:                          # loader 관례: 이름은 alias로도
                        store.add_alias(cid, a, normalize_label(a))
                created = True
            else:                                                 # 부분 수정
                merged = dict(row)
                for k in _CONCEPT_FIELDS:
                    v = getattr(body, k)
                    if v is not None:
                        merged[k] = v
                if not (merged.get("canonical_name") or "").strip():
                    raise HTTPException(400, "canonical_name은 비울 수 없습니다")
                store.upsert_concept(
                    {k: merged[k] for k in
                     ("concept_id", "status", *_CONCEPT_FIELDS)})
                # 실제 개명일 때만 새 이름 alias 추가 — 이름 그대로 저장할 때마다
                # 사용자가 지운 동명 alias가 부활하는 결함의 수정
                if body.canonical_name and body.canonical_name != row["canonical_name"]:
                    store.add_alias(cid, body.canonical_name,
                                    normalize_label(body.canonical_name))
                created = False
            store.commit()
        return {"ok": True, "concept_id": cid, "created": created}

    @app.get("/api/kg/concept/{cid}")
    def concept_get(cid: str):
        """편집기용 상세 — 개념 전체 필드 + alias + 관계 (DEPRECATED 포함)."""
        with lock:
            row = store.concept(cid)
            if row is None:
                raise HTTPException(404, "unknown concept")
            aliases = [r["alias_text"] for r in store.conn.execute(
                "SELECT alias_text FROM domain_alias WHERE concept_id=? "
                "ORDER BY alias_norm", (cid,))]
            rels = [dict(r) for r in store.conn.execute(
                "SELECT * FROM domain_relation WHERE source_concept_id=? "
                "OR target_concept_id=? ORDER BY relation_type", (cid, cid))]
            n_map = store.conn.execute(
                """SELECT count(*) FROM semantic_mapping m
                   JOIN tree_node t ON t.node_id=m.tree_node_id AND t.status='ACTIVE'
                   WHERE m.concept_id=? AND m.is_active=1
                     AND m.status IN ('AUTO_APPROVED','APPROVED','REVIEW_REQUIRED')""",
                (cid,)).fetchone()[0]
        return {"concept": dict(row), "aliases": aliases, "relations": rels,
                "active_mappings": n_map}

    @app.post("/api/kg/concept/{cid}/deprecate")
    def concept_deprecate(cid: str):
        """소프트 삭제 — 활성 매핑이 참조 중이면 409 (먼저 재매핑 유도)."""
        with wlock():
            if store.concept(cid) is None:
                raise HTTPException(404, "unknown concept")
            n = store.conn.execute(
                """SELECT count(*) FROM semantic_mapping m
                   JOIN tree_node t ON t.node_id=m.tree_node_id AND t.status='ACTIVE'
                   WHERE m.concept_id=? AND m.is_active=1
                     AND m.status IN ('AUTO_APPROVED','APPROVED','REVIEW_REQUIRED')""",
                (cid,)).fetchone()[0]
            if n:
                raise HTTPException(
                    409, f"활성 매핑 {n}건이 이 개념을 참조합니다 — 먼저 재매핑하세요")
            store.conn.execute(
                "UPDATE domain_concept SET status='DEPRECATED' WHERE concept_id=?",
                (cid,))
            store.commit()
        return JSONResponse({"ok": True})

    @app.post("/api/kg/concept/{cid}/restore")
    def concept_restore(cid: str):
        with wlock():
            if store.concept(cid) is None:
                raise HTTPException(404, "unknown concept")
            store.conn.execute(
                "UPDATE domain_concept SET status='ACTIVE' WHERE concept_id=?", (cid,))
            store.commit()
        return JSONResponse({"ok": True})

    @app.post("/api/kg/alias")
    def alias_add(body: AliasReq):
        alias = (body.alias or "").strip()
        if not alias:
            raise HTTPException(400, "alias 필수")
        with wlock():
            if store.concept(body.concept_id) is None:
                raise HTTPException(404, "unknown concept")
            store.add_alias(body.concept_id, alias, normalize_label(alias))
            store.commit()
        return JSONResponse({"ok": True})

    @app.delete("/api/kg/alias")
    def alias_delete(concept_id: str, alias: str):
        with wlock():
            cur = store.conn.execute(
                "DELETE FROM domain_alias WHERE concept_id=? AND alias_norm=?",
                (concept_id, normalize_label(alias)))
            store.commit()
        if not cur.rowcount:
            raise HTTPException(404, "unknown alias")
        return JSONResponse({"ok": True})

    def _isa_would_cycle(src: str, dst: str) -> bool:
        """src IS_A dst 추가 시 사이클 여부 — dst에서 위로 가는 **모든** 부모
        경로를 BFS로 탐색한다 (다중 부모가 있는 기존 데이터에서 마지막 행만
        보던 가드가 우회되는 결함의 수정)."""
        parents: dict[str, list[str]] = {}
        for r in store.conn.execute(
                "SELECT source_concept_id s, target_concept_id t "
                "FROM domain_relation WHERE relation_type='IS_A'"):
            parents.setdefault(r["s"], []).append(r["t"])
        seen, frontier = {dst}, [dst]
        for _ in range(16):
            if src in seen:
                return True
            nxt = [p for cur in frontier for p in parents.get(cur, [])
                   if p not in seen]
            if not nxt:
                return False
            seen.update(nxt)
            frontier = nxt
        return True                                 # 깊이 초과 = 비정상 체인 취급

    @app.post("/api/kg/relation")
    def relation_add(body: RelationReq):
        if body.type not in VALID_RELATIONS:
            raise HTTPException(400, f"type must be one of {sorted(VALID_RELATIONS)}")
        if body.source == body.target:
            raise HTTPException(400, "self-loop 불가")
        with wlock():
            for c in (body.source, body.target):
                if store.concept(c) is None:
                    raise HTTPException(404, f"unknown concept: {c}")
            if body.type == "IS_A":
                if _isa_would_cycle(body.source, body.target):
                    raise HTTPException(400, "IS_A 사이클이 생깁니다")
                # IS_A 부모는 1개 — 루트 도출(isa_roots)이 단일 부모를 전제한다
                prev = store.conn.execute(
                    "SELECT target_concept_id FROM domain_relation "
                    "WHERE source_concept_id=? AND relation_type='IS_A' "
                    "AND target_concept_id != ?",
                    (body.source, body.target)).fetchone()
                if prev is not None:
                    raise HTTPException(
                        400, f"IS_A 부모는 1개입니다 — 기존 부모"
                             f"({prev['target_concept_id']})를 먼저 삭제하세요")
            store.add_relation(body.source, body.target, body.type)
            store.commit()
        warning = ("IS_A 변경은 Document KG(문서군) 재편성에 영향을 줍니다"
                   if body.type == "IS_A" else None)
        return {"ok": True, "warning": warning}

    @app.delete("/api/kg/relation")
    def relation_delete(source: str, target: str, type: str):
        with wlock():
            cur = store.conn.execute(
                "DELETE FROM domain_relation WHERE source_concept_id=? "
                "AND target_concept_id=? AND relation_type=?",
                (source, target, type))
            store.commit()
        if not cur.rowcount:
            raise HTTPException(404, "unknown relation")
        warning = ("IS_A 변경은 Document KG(문서군) 재편성에 영향을 줍니다"
                   if type == "IS_A" else None)
        return {"ok": True, "warning": warning}

    @app.get("/api/kg/export")
    def kg_export():
        """현재 DB를 domain_kg.yaml 형식으로 — 웹 편집의 YAML 역반영 통로."""
        import yaml as _yaml
        with lock:
            crows = store.conn.execute(
                "SELECT * FROM domain_concept ORDER BY concept_id").fetchall()
            arows = store.conn.execute(
                "SELECT concept_id, alias_text FROM domain_alias "
                "ORDER BY concept_id, alias_norm").fetchall()
            rrows = store.conn.execute(
                "SELECT * FROM domain_relation ORDER BY 1,2,3").fetchall()
        amap: dict[str, list[str]] = {}
        for r in arows:
            amap.setdefault(r["concept_id"], []).append(r["alias_text"])
        concepts = []
        for r in crows:
            c: dict = {"concept_id": r["concept_id"],
                       "canonical_name": r["canonical_name"]}
            for k in ("canonical_name_en", "description", "concept_type",
                      "data_type", "domain_level", "canonical_unit",
                      "unit_dimension"):
                if r[k]:
                    c[k] = r[k]
            if r["status"] != "ACTIVE":
                c["status"] = r["status"]
            als = [a for a in amap.get(r["concept_id"], [])
                   if a not in (r["canonical_name"], r["canonical_name_en"])]
            if als:
                c["aliases"] = als
            concepts.append(c)
        data = {"version": f"db-export-{now_iso()}", "concepts": concepts,
                "relations": [[r["source_concept_id"], r["target_concept_id"],
                               r["relation_type"]] for r in rrows]}
        return PlainTextResponse(
            _yaml.safe_dump(data, allow_unicode=True, sort_keys=False),
            media_type="text/yaml; charset=utf-8")

    # React 포트(frontend/) 빌드가 있으면 /app 에 함께 서빙한다.
    # web_kg는 완전 대체 전까지 / 에서 그대로 유지 (이중 유지 의도).
    react_dist = Path(__file__).resolve().parent.parent / "frontend" / "dist"
    if react_dist.is_dir():
        app.mount("/app", StaticFiles(directory=react_dist, html=True), name="react")
    app.mount("/", StaticFiles(directory=WEB_DIR, html=True), name="static")
    return app


def main() -> int:
    import argparse

    import uvicorn
    p = argparse.ArgumentParser(prog="kg.webapp")
    p.add_argument("--ws", default=".", type=Path)
    p.add_argument("--port", default=8010, type=int)
    p.add_argument("--host", default="127.0.0.1")
    args = p.parse_args()
    uvicorn.run(create_app(args.ws), host=args.host, port=args.port)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
