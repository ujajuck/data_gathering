"""P1/P4/P7 합격 기준: 버전 이력, 증분 갱신, idempotency, rollback (§8, §9, §14)."""
from __future__ import annotations

import time

from src.pipeline import Pipeline

from tests.conftest import F_BATCH, F_INSPECTION, F_QUALITY, stage_fixture
from tests.mutations import change_cell_value, delete_block_rows


def test_initial_load_and_idempotent_reload(tmp_repo):
    stage_fixture(tmp_repo, F_INSPECTION)
    pipe = Pipeline(tmp_repo)
    r1 = pipe.process_dir(tmp_repo / "data" / "raw")
    assert r1[0]["status"] == "SUCCESS" and r1[0]["stats"]["inserted"] == 4

    # 같은 파일 재처리 → semantic cache hit, DB 미변경 (§5.4, §8.6)
    r2 = pipe.process_dir(tmp_repo / "data" / "raw")
    assert r2[0]["cache_hit"] is True

    # cache 우회해도 loader는 idempotent (§15)
    r3 = pipe.process_file(tmp_repo / "data" / "raw" / F_INSPECTION.name, force=True)
    assert r3["stats"] == {"inserted": 0, "updated": 0, "unchanged": 4, "tombstoned": 0}


def test_value_change_versions_only_affected_rows(tmp_repo):
    """측정값 71→73: 해당 observation만 새 버전, record는 version 2 (§9, §14)."""
    target = stage_fixture(tmp_repo, F_INSPECTION)
    pipe = Pipeline(tmp_repo)
    pipe.process_file(target)

    before_ts = None
    time.sleep(0.01)

    # 베어링온도 실측값 C10: 71 → 73 (XML patch — 수식 캐시 등 모두 보존)
    mutated = tmp_repo / "mut.xlsx"
    change_cell_value(target, mutated, "C10", "71", "73")
    mutated.replace(target)

    import datetime
    before_ts = datetime.datetime.now(datetime.timezone.utc).isoformat()
    time.sleep(0.01)

    r = pipe.process_file(target)
    assert r["stats"]["updated"] == 1 and r["stats"]["unchanged"] == 3

    key = "설비 점검표|P-101|2026-08-18"
    obs = pipe.loader.current_observations(key)
    bearing = [o for o in obs if o["observation_key"] == "베어링온도>실측값.실측값"]
    assert len(bearing) == 1 and bearing[0]["raw_value_num"] == 73.0

    # 영향받지 않은 observation은 이전 버전 그대로 (§8.6 영향 row만 갱신)
    vib = [o for o in obs if o["observation_key"].startswith("진동>실측값")]
    assert vib[0]["valid_from"] < bearing[0]["valid_from"]

    # record는 종료+INSERT로 version 2 (§9.1)
    rec = [r for r in pipe.loader.current_records() if r["record_key"] == key][0]
    assert rec["version"] == 2
    history = pipe.loader.conn.execute(
        "SELECT count(*) n FROM record WHERE record_key=?", (key,)).fetchone()["n"]
    assert history == 2

    # 다른 record 3개는 새 버전이 만들어지지 않았다
    others = pipe.loader.conn.execute(
        "SELECT record_key, count(*) n FROM record WHERE record_key != ? GROUP BY record_key",
        (key,)).fetchall()
    assert all(row["n"] == 1 for row in others)

    # Rollback: 변경 전 시점의 current view 재구성 (§14 Rollback)
    past = pipe.loader.observations_as_of(before_ts, key)
    old_bearing = [o for o in past if o["observation_key"] == "베어링온도>실측값.실측값"]
    assert old_bearing[0]["raw_value_num"] == 71.0


def test_document_version_history(tmp_repo):
    """파일 변경이 document_version history에 남는다 (§14, P1 완료 조건)."""
    target = stage_fixture(tmp_repo, F_INSPECTION)
    pipe = Pipeline(tmp_repo)
    pipe.process_file(target)
    mutated = tmp_repo / "mut.xlsx"
    change_cell_value(target, mutated, "C10", "71", "72")
    mutated.replace(target)
    pipe.process_file(target)

    rows = pipe.loader.conn.execute(
        """SELECT dv.is_current, dv.supersedes_version_id FROM document_version dv
           ORDER BY dv.detected_at""").fetchall()
    assert len(rows) == 2
    assert rows[0]["is_current"] == 0 and rows[1]["is_current"] == 1
    assert rows[1]["supersedes_version_id"] is not None


def test_block_deletion_tombstones_record(tmp_repo):
    """원본에서 사라진 Block은 즉시 delete가 아니라 tombstone (§15)."""
    target = stage_fixture(tmp_repo, F_BATCH)
    pipe = Pipeline(tmp_repo)
    pipe.process_file(target)
    assert len(pipe.loader.current_records()) == 4

    mutated = tmp_repo / "mut.xlsx"
    delete_block_rows(target, mutated, 43, 54)  # BATCH #4 삭제
    mutated.replace(target)
    r = pipe.process_file(target)
    assert r["stats"]["tombstoned"] == 1
    assert len(pipe.loader.current_records()) == 3
    dead = pipe.loader.conn.execute(
        "SELECT * FROM record WHERE is_tombstone=1").fetchall()
    assert len(dead) == 1 and dead[0]["business_key"] == "B-260820-01"


def test_multi_document_ingest_produces_unified_output(tmp_repo):
    """서로 다른 양식 3종이 동일한 출력 계약으로 통합된다 (P5)."""
    for f in (F_INSPECTION, F_QUALITY, F_BATCH):
        stage_fixture(tmp_repo, f)
    pipe = Pipeline(tmp_repo)
    results = pipe.process_dir(tmp_repo / "data" / "raw")
    assert [r["records"] for r in results] == [4, 3, 4]

    from src.export.workbook import SHEETS, CanonicalWorkbookExporter
    exporter = CanonicalWorkbookExporter(pipe.loader)
    out = exporter.export(tmp_repo / "canonical.xlsx")

    import openpyxl
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == list(SHEETS)
    assert wb["01_Record_Index"].max_row == 12  # header + 11 records
    assert wb["04_Attachments"].max_row == 12   # header + 11 images

    # 분석용 wide view pivot (§6.1)
    wide = exporter.build_wide_view(["yield_rate", "average_temperature"])
    assert len(wide) == 4
    row = next(w for w in wide if w["record_key"].endswith("B-260818-02|2026-08-18"))
    assert row["yield_rate[값]"] < 92
