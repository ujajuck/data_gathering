"""텍스트박스/이미지가 든 xlsx 생성 헬퍼.

openpyxl은 도형(xdr:sp)을 쓰지 못하므로, 통상 경로로 워크북을 만든 뒤 zip을
다시 열어 드로잉 파트를 직접 주입한다. 앵커가 알려진 두 개의 텍스트박스:

- twoCellAnchor C3 → E6 (from col=2,row=2 / to col=5,row=6, off 0)
- oneCellAnchor H2 + ext 1828800×731520 EMU (192×76.8 px)

with_images=True면 앵커가 알려진 이미지 2장도 넣는다 (openpyxl 네이티브 —
이 경우 openpyxl이 만든 드로잉 XML에 텍스트박스 앵커를 병합한다):

- twoCellAnchor C12 → E16 (192×100 px 영역)
- oneCellAnchor G12 + ext 952500×571500 EMU (100×60 px)
"""
from __future__ import annotations

import io
import shutil
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

DRAWING_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
 <xdr:twoCellAnchor>
  <xdr:from><xdr:col>2</xdr:col><xdr:colOff>0</xdr:colOff>
   <xdr:row>2</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
  <xdr:to><xdr:col>5</xdr:col><xdr:colOff>0</xdr:colOff>
   <xdr:row>6</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>
  <xdr:sp macro="" textlink="">
   <xdr:nvSpPr><xdr:cNvPr id="2" name="TextBox 1"/><xdr:cNvSpPr txBox="1"/></xdr:nvSpPr>
   <xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>
    <a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr>
   <xdr:txBody><a:bodyPr/><a:lstStyle/>
    <a:p><a:r><a:t>주의: 결과값은 C3:E6 영역 기준</a:t></a:r></a:p>
    <a:p><a:r><a:t>검수 전 반출 금지</a:t></a:r></a:p></xdr:txBody>
  </xdr:sp>
  <xdr:clientData/>
 </xdr:twoCellAnchor>
 <xdr:oneCellAnchor>
  <xdr:from><xdr:col>7</xdr:col><xdr:colOff>47625</xdr:colOff>
   <xdr:row>1</xdr:row><xdr:rowOff>19050</xdr:rowOff></xdr:from>
  <xdr:ext cx="1828800" cy="731520"/>
  <xdr:sp macro="" textlink="">
   <xdr:nvSpPr><xdr:cNvPr id="3" name="TextBox 2"/><xdr:cNvSpPr txBox="1"/></xdr:nvSpPr>
   <xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>
    <a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr>
   <xdr:txBody><a:bodyPr/><a:lstStyle/>
    <a:p><a:r><a:t>검사자 메모 — H2 앵커</a:t></a:r></a:p></xdr:txBody>
  </xdr:sp>
  <xdr:clientData/>
 </xdr:oneCellAnchor>
</xdr:wsDr>
"""

_DRAWING_CT = ('<Override PartName="/xl/drawings/drawing1.xml" '
               'ContentType="application/vnd.openxmlformats-officedocument.'
               'drawing+xml"/>')
_SHEET_RELS = ("""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
 <Relationship Id="rId99" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"
  Target="../drawings/drawing1.xml"/>
</Relationships>
""")
_R_NS = 'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'


def _solid_png(w: int, h: int, color: str) -> io.BytesIO:
    from PIL import Image as PILImage
    buf = io.BytesIO()
    PILImage.new("RGB", (w, h), color).save(buf, format="PNG")
    buf.seek(0)
    return buf


def build_textbox_xlsx(path: Path, with_images: bool = False) -> Path:
    """앵커가 알려진 텍스트박스 2개(+옵션 이미지 2장)가 든 xlsx를 만든다."""
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "텍스트박스 렌더 검증 시트"
    ws["A1"].font = Font(bold=True)
    for r in range(3, 7):                      # C3:E6 — 박스가 덮어야 할 영역
        for c in range(3, 6):
            cell = ws.cell(row=r, column=c, value=f"{chr(64 + c)}{r}")
            cell.fill = PatternFill("solid", fgColor="FFE8F0FE")
    ws["H2"] = "H2"
    ws["C9"] = "온도"
    ws["D9"] = 180
    ws["C10"] = "부풀기"
    ws["D10"] = 9.5

    if with_images:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.drawing.spreadsheet_drawing import (AnchorMarker,
                                                          OneCellAnchor,
                                                          TwoCellAnchor)
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        ws["C11"] = "사진 1 (C12:E16)"
        ws["G11"] = "사진 2 (G12)"
        img1 = XLImage(_solid_png(120, 80, "#c0392b"))
        img1.anchor = TwoCellAnchor(
            editAs="twoCell",
            _from=AnchorMarker(col=2, colOff=0, row=11, rowOff=0),
            to=AnchorMarker(col=5, colOff=0, row=16, rowOff=0))
        ws.add_image(img1)
        img2 = XLImage(_solid_png(100, 60, "#2980b9"))
        img2.anchor = OneCellAnchor(
            _from=AnchorMarker(col=6, colOff=0, row=11, rowOff=0),
            ext=XDRPositiveSize2D(cx=952500, cy=571500))
        ws.add_image(img2)
    wb.save(path)

    tmp = path.with_suffix(".tmp.xlsx")
    has_drawing = False
    with zipfile.ZipFile(path) as zin:
        has_drawing = "xl/drawings/drawing1.xml" in zin.namelist()
    with zipfile.ZipFile(path) as zin, \
            zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            data = zin.read(item)
            if item == "xl/drawings/drawing1.xml":
                # openpyxl이 만든 드로잉(이미지)에 텍스트박스 앵커 병합
                droot = ET.fromstring(data)
                for anchor in ET.fromstring(DRAWING_XML):
                    droot.append(anchor)
                data = ET.tostring(droot, xml_declaration=True,
                                   encoding="UTF-8")
            elif not has_drawing and item == "[Content_Types].xml":
                data = data.replace(b"</Types>",
                                    _DRAWING_CT.encode() + b"</Types>")
            elif not has_drawing and item == "xl/worksheets/sheet1.xml":
                text = data.decode("utf-8")
                if "xmlns:r=" not in text:
                    text = text.replace("<worksheet ",
                                        f"<worksheet {_R_NS} ", 1)
                text = text.replace("</worksheet>",
                                    '<drawing r:id="rId99"/></worksheet>')
                data = text.encode("utf-8")
            zout.writestr(item, data)
        if not has_drawing:
            zout.writestr("xl/drawings/drawing1.xml", DRAWING_XML)
            zout.writestr("xl/worksheets/_rels/sheet1.xml.rels", _SHEET_RELS)
    shutil.move(tmp, path)
    return path
