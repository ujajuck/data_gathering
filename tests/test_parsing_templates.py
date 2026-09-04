"""Parsing Template versioning, delta override, runtime, and provenance tests."""
from __future__ import annotations

from hashlib import sha256

import pytest
from openpyxl import Workbook

from kg.parsing import (ParsingError, add_version, assign, create_template,
                        effective_mappings, grouped_documents, run_parse,
                        save_override, template_detail, unassign)
from kg.store import KgStore


def _workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "190도"
    ws["A1"] = "온도"
    ws["B1"] = 190
    ws["F7"] = 42.1
    ws["G7"] = 43.2
    wb.save(path)


def _document(store, path):
    document_id = "doc-financier"
    store.upsert_document(document_id, path.name, str(path))
    version_id = store.add_version(document_id, sha256(path.read_bytes()).hexdigest(), "test")
    return document_id, version_id


def _spec(weight_range="F7:F7"):
    return {"sheet_templates": [{
        "name": "oven_test",
        "match": {"name_regex": "^(180|185|190|195|200)"},
        "mappings": [
            {"key": "temperature", "concept_id": "oven_temperature",
             "source": {"key_search": ["온도"], "offset": {"row": 0, "col": 1}},
             "type": "number", "unit": "C"},
            {"key": "weight", "concept_id": "weight",
             "source": {"range": weight_range}, "type": "number", "unit": "g"},
        ],
    }]}


def test_versioned_template_assignment_and_partial_override(tmp_path):
    path = tmp_path / "coffee.xlsx"
    _workbook(path)
    store = KgStore(tmp_path / "kg.db")
    document_id, version_id = _document(store, path)

    create_template(store, "financier_recipe", "Financier Recipe", "financier")
    v1 = add_version(store, "financier_recipe", _spec(), "tester")
    v2 = add_version(store, "financier_recipe", _spec("H7:H7"), "tester")
    assert (v1["version"], v2["version"]) == (1, 2)
    assert template_detail(store, "financier_recipe")["versions"] == [1, 2]

    assignment = assign(store, document_id, version_id, "financier_recipe", 1)
    assert assignment["template_version"] == 1
    mappings = effective_mappings(store, version_id)
    weight = next(m for m in mappings if m["mapping_key"] == "weight")
    save_override(store, document_id, version_id, weight["mapping_id"],
                  {"range": "G7:G7"}, "column moved", "tester")

    effective = effective_mappings(store, version_id)
    assert len(effective) == 2  # only the changed mapping is stored as a delta
    weight = next(m for m in effective if m["mapping_key"] == "weight")
    assert weight["template_source"] == {"range": "F7:F7"}
    assert weight["effective_source"] == {"range": "G7:G7"}
    assert weight["mapping_source"] == "MANUAL"
    assert store.conn.execute("SELECT count(*) FROM document_override").fetchone()[0] == 1

    result = run_parse(store, document_id, version_id, path)
    assert result["status"] == "SUCCESS"
    assert result["mapping_count"] == 2 and result["override_count"] == 1
    sources = {s["concept_id"]: s for s in result["sources"]}
    assert sources["weight"]["value"] == 43.2
    assert sources["weight"]["mapping_source"] == "MANUAL"
    assert sources["weight"]["template_source"] == {"range": "F7:F7"}
    assert sources["weight"]["effective_source"] == {"range": "G7:G7"}
    assert sources["oven_temperature"]["value"] == 190

    groups = grouped_documents(store, "financier")
    assert groups[0]["version"] == 1
    assert groups[0]["override_documents"] == 1
    assert groups[0]["documents"][0]["filename"] == "coffee.xlsx"
    store.close()


def test_partial_failure_becomes_review_required(tmp_path):
    path = tmp_path / "coffee.xlsx"
    _workbook(path)
    store = KgStore(tmp_path / "kg.db")
    document_id, version_id = _document(store, path)
    create_template(store, "wrong_layout", "Wrong Layout", None)
    spec = _spec()
    spec["sheet_templates"][0]["match"] = {"names": ["missing-sheet"]}
    add_version(store, "wrong_layout", spec)
    assign(store, document_id, version_id, "wrong_layout", 1)

    result = run_parse(store, document_id, version_id, path)
    assert result["status"] == "FAILED"
    assert result["warning_count"] == 2
    assert {s["status"] for s in result["sources"]} == {"MISSING"}
    store.close()


def test_regex_sheet_template_extracts_every_repeated_sheet(tmp_path):
    path = tmp_path / "experiments.xlsx"
    wb = Workbook()
    first = wb.active
    first.title = "180도"
    for index, name in enumerate(("180도", "185도", "190도", "195도")):
        ws = first if index == 0 else wb.create_sheet(name)
        ws["A1"] = "온도"
        ws["B1"] = 180 + index * 5
        ws["F7"] = 40 + index
    wb.save(path)
    store = KgStore(tmp_path / "kg.db")
    document_id, version_id = _document(store, path)
    create_template(store, "oven_repeated", "Repeated Oven", "financier")
    add_version(store, "oven_repeated", _spec())
    assign(store, document_id, version_id, "oven_repeated", 1)

    result = run_parse(store, document_id, version_id, path)
    assert result["status"] == "SUCCESS"
    assert result["mapping_count"] == 8  # two mappings on each of four sheets
    assert {source["sheet_name"] for source in result["sources"]} == \
        {"180도", "185도", "190도", "195도"}
    temperatures = sorted(source["value"] for source in result["sources"]
                          if source["concept_id"] == "oven_temperature")
    assert temperatures == [180, 185, 190, 195]
    store.close()


def test_template_upgrade_keeps_conflicting_override_until_review(tmp_path):
    path = tmp_path / "coffee.xlsx"
    _workbook(path)
    store = KgStore(tmp_path / "kg.db")
    document_id, version_id = _document(store, path)
    create_template(store, "recipe", "Recipe", "financier")
    add_version(store, "recipe", _spec("F7:F7"))
    add_version(store, "recipe", _spec("H7:H7"))
    add_version(store, "recipe", _spec("G7:G7"))
    assign(store, document_id, version_id, "recipe", 1)
    weight = next(m for m in effective_mappings(store, version_id)
                  if m["mapping_key"] == "weight")
    override = save_override(store, document_id, version_id, weight["mapping_id"],
                             {"range": "G7:G7"}, "document exception", "tester")

    upgraded = assign(store, document_id, version_id, "recipe", 2)
    assert upgraded["status"] == "REVIEW_REQUIRED"
    row = store.conn.execute(
        "SELECT status FROM document_override WHERE override_id=?",
        (override["override_id"],)).fetchone()
    assert row["status"] == "CONFLICT"
    weight_v2 = next(m for m in effective_mappings(store, version_id)
                     if m["mapping_key"] == "weight")
    assert weight_v2["template_source"] == {"range": "H7:H7"}
    assert weight_v2["effective_source"] == {"range": "G7:G7"}
    assert weight_v2["mapping_source"] == "MANUAL"

    redundant = assign(store, document_id, version_id, "recipe", 3)
    assert redundant["status"] == "ASSIGNED"
    assert store.conn.execute(
        "SELECT status FROM document_override WHERE override_id=?",
        (override["override_id"],)).fetchone()[0] == "REDUNDANT"
    weight_v3 = next(m for m in effective_mappings(store, version_id)
                     if m["mapping_key"] == "weight")
    assert weight_v3["mapping_source"] == "TEMPLATE"
    store.close()


def test_declared_type_and_unit_normalization_are_applied(tmp_path):
    path = tmp_path / "kelvin.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "190도"
    ws["B1"] = "348.15"
    wb.save(path)
    store = KgStore(tmp_path / "kg.db")
    document_id, version_id = _document(store, path)
    create_template(store, "normalized", "Normalized", None)
    add_version(store, "normalized", {"sheet_templates": [{
        "name": "oven", "match": {"names": ["190도"]}, "mappings": [{
            "key": "temperature", "concept_id": "oven_temperature",
            "source": {"range": "B1"}, "type": "number", "unit": "K",
            "normalization": {"target_unit": "C"},
        }]}]})
    assign(store, document_id, version_id, "normalized", 1)
    result = run_parse(store, document_id, version_id, path)
    assert result["status"] == "SUCCESS"
    assert result["sources"][0]["value"] == pytest.approx(75.0)
    store.close()


def test_document_can_carry_multiple_templates(tmp_path):
    """N:M — 템플릿마다 파싱하려는 정보가 달라도 한 문서에 같이 배정된다."""
    path = tmp_path / "coffee.xlsx"
    _workbook(path)
    store = KgStore(tmp_path / "kg.db")
    document_id, version_id = _document(store, path)
    create_template(store, "exp_view", "Experiment View", "financier")
    add_version(store, "exp_view", _spec())
    create_template(store, "cost_view", "Cost View", "financier")
    add_version(store, "cost_view", {"sheet_templates": [{
        "name": "cost", "match": {"names": ["190도"]}, "mappings": [{
            "key": "unit_cost", "concept_id": "unit_cost",
            "source": {"range": "G7:G7"}, "type": "number"}]}]})

    assign(store, document_id, version_id, "exp_view", 1)
    assign(store, document_id, version_id, "cost_view", 1)   # 교체가 아니라 추가
    assigned = {r[0] for r in store.conn.execute(
        "SELECT template_id FROM document_template_assignment WHERE document_version=?",
        (version_id,))}
    assert assigned == {"exp_view", "cost_view"}

    merged = effective_mappings(store, version_id)
    assert {(m["template_id"], m["mapping_key"]) for m in merged} == {
        ("exp_view", "temperature"), ("exp_view", "weight"),
        ("cost_view", "unit_cost")}
    assert [m["mapping_key"] for m in
            effective_mappings(store, version_id, "cost_view")] == ["unit_cost"]

    with pytest.raises(ParsingError):        # 파싱은 템플릿 단위
        run_parse(store, document_id, version_id, path)
    exp = run_parse(store, document_id, version_id, path, "exp_view")
    cost = run_parse(store, document_id, version_id, path, "cost_view")
    assert exp["status"] == "SUCCESS" and cost["status"] == "SUCCESS"
    assert {s["concept_id"] for s in cost["sources"]} == {"unit_cost"}
    assert {s["concept_id"] for s in exp["sources"]} == {"oven_temperature", "weight"}

    unassign(store, document_id, version_id, "exp_view")     # 해제는 그 템플릿만
    assigned = {r[0] for r in store.conn.execute(
        "SELECT template_id FROM document_template_assignment WHERE document_version=?",
        (version_id,))}
    assert assigned == {"cost_view"}
    with pytest.raises(ParsingError):
        unassign(store, document_id, version_id, "exp_view")
    store.close()


def test_override_and_version_audit_stay_per_template(tmp_path):
    """같은 문서의 다른 템플릿 버전 교체가 서로의 override를 건드리지 않는다."""
    path = tmp_path / "coffee.xlsx"
    _workbook(path)
    store = KgStore(tmp_path / "kg.db")
    document_id, version_id = _document(store, path)
    create_template(store, "a_view", "A", "financier")
    add_version(store, "a_view", _spec("F7:F7"))
    add_version(store, "a_view", _spec("H7:H7"))
    create_template(store, "b_view", "B", "financier")
    add_version(store, "b_view", _spec("F7:F7"))
    assign(store, document_id, version_id, "a_view", 1)
    assign(store, document_id, version_id, "b_view", 1)

    b_weight = next(m for m in effective_mappings(store, version_id, "b_view")
                    if m["mapping_key"] == "weight")
    save_override(store, document_id, version_id, b_weight["mapping_id"],
                  {"range": "G7:G7"}, "b만 예외", "tester")

    upgraded = assign(store, document_id, version_id, "a_view", 2)  # a만 버전 교체
    assert upgraded["status"] == "ASSIGNED"      # b의 override는 a와 무관하다
    b_status = store.conn.execute(
        """SELECT a.status FROM document_template_assignment a
            WHERE a.document_version=? AND a.template_id='b_view'""",
        (version_id,)).fetchone()[0]
    assert b_status == "OVERRIDDEN"
    b_weight2 = next(m for m in effective_mappings(store, version_id, "b_view")
                     if m["mapping_key"] == "weight")
    assert b_weight2["mapping_source"] == "MANUAL"
    a_weight = next(m for m in effective_mappings(store, version_id, "a_view")
                    if m["mapping_key"] == "weight")
    assert a_weight["mapping_source"] == "TEMPLATE"   # 옆 템플릿 override 미전염
    store.close()


def test_legacy_one_to_one_assignment_db_migrates(tmp_path):
    """옛 PK(document_id,document_version) DB를 열면 N:M PK로 재구축된다."""
    db = tmp_path / "kg.db"
    store = KgStore(db)
    path = tmp_path / "coffee.xlsx"
    _workbook(path)
    document_id, version_id = _document(store, path)
    create_template(store, "legacy", "Legacy", None)
    add_version(store, "legacy", _spec())
    assign(store, document_id, version_id, "legacy", 1)
    # 옛(1:1) 형태로 강제 다운그레이드
    store.conn.executescript("""
        PRAGMA foreign_keys = OFF;
        ALTER TABLE document_template_assignment RENAME TO _new;
        CREATE TABLE document_template_assignment (
            document_id TEXT NOT NULL, document_version TEXT NOT NULL,
            template_id TEXT NOT NULL, template_version INTEGER NOT NULL,
            status TEXT NOT NULL DEFAULT 'ASSIGNED', assigned_at TEXT NOT NULL,
            PRIMARY KEY (document_id, document_version));
        INSERT INTO document_template_assignment SELECT * FROM _new;
        DROP TABLE _new;
    """)
    store.conn.commit()
    store.close()

    reopened = KgStore(db)                       # 여기서 마이그레이션
    pk = [r[1] for r in reopened.conn.execute(
        "PRAGMA table_info(document_template_assignment)") if r[5] > 0]
    assert "template_id" in pk
    row = reopened.conn.execute(
        "SELECT template_id,template_version FROM document_template_assignment").fetchone()
    assert (row[0], row[1]) == ("legacy", 1)     # 기존 배정 보존
    create_template(reopened, "second", "Second", None)
    add_version(reopened, "second", _spec())
    assign(reopened, document_id, version_id, "second", 1)   # 이제 추가 배정 가능
    assert reopened.conn.execute(
        "SELECT count(*) FROM document_template_assignment").fetchone()[0] == 2
    reopened.close()
