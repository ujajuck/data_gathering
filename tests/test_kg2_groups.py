"""KG2 회귀 — DKG 그룹 델타 / Extraction Recipe / 재크롤링 / seed 3-모드 / 편집 API.

사용자 요구 4종의 완료 기준:
  1. 같은 형식 새 문서 → DKG 선택 배정(레시피로 매핑 이식)
  2. KG/DKG 편집 가능
  3. 편집 후 재크롤링(fill/reset_auto, 사람 결정 보존)
  4. 크롤링 '코드'가 아닌 선언적 레시피 저장
"""
from __future__ import annotations

import json
import shutil
import threading
import time

import openpyxl
import pytest
from openpyxl.styles import Font, PatternFill

from tests.conftest import FIXTURES

from kg.domain.loader import load_domain_kg
from kg.groups import (clear_member_override, group_documents, isa_roots,
                       set_member_override)
from kg.mapping.judge import RuleJudge
from kg.mapping.mapper import (map_document, map_nodes_staged, remap_reviewed,
                               reset_document_mappings)
from kg.mapping.recipe import (active_recipe, apply_recipe, snapshot_recipe,
                               suggest_groups)
from kg.mapping.retriever import DomainRetriever
from kg.recrawl import run_recrawl, start_run
from kg.store import KgStore
from kg.tree.builder import load_workbook_tree
from kg.tree.diff import apply_tree

KG_YAML = FIXTURES.parent.parent / "domains" / "financier" / "config" / "domain_kg.yaml"
UNITS_YAML = FIXTURES.parent.parent / "domains" / "financier" / "config" / "units.yaml"


@pytest.fixture()
def ws(tmp_path):
    store = KgStore(tmp_path / "kg.db")
    info = load_domain_kg(store, KG_YAML, UNITS_YAML)

    class W:
        pass

    w = W()
    w.root = tmp_path
    w.store = store
    w.units = info["units"]
    w.parser_rules = {}
    w.registry = None

    def ingest(path, parser_version="t"):
        doc_id, drafts, h = load_workbook_tree(store, tmp_path, path, {},
                                               w.units, None)
        return doc_id, apply_tree(store, doc_id, path.name, str(path), h,
                                  parser_version, drafts)

    w.ingest = ingest
    w.retriever = lambda: DomainRetriever(store, units=w.units)
    yield w
    store.close()


def _mini(path, header="심부온도", values=(25.0, 45.0, 65.0), extra_col=None):
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "TEST"
    fill = PatternFill("solid", fgColor="FFD9E1F2")
    headers = ["시점", header] + ([extra_col] if extra_col else [])
    for col, name in enumerate(headers, start=1):
        c = s.cell(row=1, column=col, value=name)
        c.font = Font(bold=True)
        c.fill = fill
    for i, v in enumerate(values):
        s.cell(row=2 + i, column=1, value=f"t{i}")
        s.cell(row=2 + i, column=2, value=v)
        if extra_col:
            s.cell(row=2 + i, column=3, value=v * 2)
    wb.save(path)
    return path


def _map(w, doc_id):
    return map_document(w.store, w.retriever(), RuleJudge(), doc_id)


def _core_node(w, doc_id, name="심부온도"):
    return next(r for r in w.store.active_nodes(doc_id).values()
                if r["node_name"] == name)


def _root_of(w, concept="core_temperature"):
    roots, _, _ = isa_roots(w.store)
    root = roots[concept]
    assert root is not None
    return root


# ------------------------------------------------------- 그룹 멤버십 델타 ----
def test_group_overrides_include_exclude(ws, tmp_path):
    f = _mini(tmp_path / "mini.xlsx")
    doc_id, _ = ws.ingest(f)
    _map(ws, doc_id)
    root = _root_of(ws)
    assert doc_id in group_documents(ws.store, root)      # 파생 멤버

    set_member_override(ws.store, root, doc_id, "EXCLUDED")
    ws.store.commit()
    assert doc_id not in group_documents(ws.store, root)  # tombstone이 파생 차단

    clear_member_override(ws.store, root, doc_id)
    ws.store.commit()
    assert doc_id in group_documents(ws.store, root)      # 오버라이드 해제 → 파생 복귀

    other = "domain_cost"                                  # 매핑 없는 그룹에 핀 고정
    set_member_override(ws.store, other, doc_id, "INCLUDED")
    ws.store.commit()
    assert doc_id in group_documents(ws.store, other)


# ------------------------------------------------------------- 레시피 ----
def test_recipe_snapshot_human_tier_wins(ws, tmp_path):
    """2단 다수결: APPROVED 표가 있으면 AUTO 다수도 못 뒤집는다."""
    d1, _ = ws.ingest(_mini(tmp_path / "a.xlsx"))
    d2, _ = ws.ingest(_mini(tmp_path / "b.xlsx", values=(1.0, 2.0, 3.0)))
    _map(ws, d1)
    _map(ws, d2)
    root = _root_of(ws)
    # 사람 remap은 파생 멤버십도 바꿀 수 있으므로 멤버십을 핀 고정한다
    for d in (d1, d2):
        set_member_override(ws.store, root, d, "INCLUDED")
    ws.store.commit()
    # 사람이 d1의 심부온도를 core_temp_max로 확정 — AUTO(core) 2표를 이긴다
    m = ws.store.active_mapping(_core_node(ws, d1)["node_id"])
    remap_reviewed(ws.store, m["mapping_id"], "core_temp_max", "tester")
    res = snapshot_recipe(ws.store, root, note="t")
    assert res["template"] >= 1
    spec = json.loads(active_recipe(ws.store, root)["spec_json"])
    entry = next(e for e in spec["template"] if e["node_name"] == "심부온도")
    assert entry["concept_id"] == "core_temp_max"
    assert entry["tier"] == "APPROVED"


def test_recipe_snapshot_tie_dropped(ws, tmp_path):
    d1, _ = ws.ingest(_mini(tmp_path / "a.xlsx", extra_col="오븐온도"))
    d2, _ = ws.ingest(_mini(tmp_path / "b.xlsx", values=(1.0, 2.0, 3.0),
                            extra_col="오븐온도"))
    _map(ws, d1)
    _map(ws, d2)
    root = _root_of(ws)
    for d in (d1, d2):
        set_member_override(ws.store, root, d, "INCLUDED")
    ws.store.commit()
    for doc, cid in ((d1, "core_temp_max"), (d2, "ambient_temperature")):
        m = ws.store.active_mapping(_core_node(ws, doc)["node_id"])
        remap_reviewed(ws.store, m["mapping_id"], cid, "tester")
    snapshot_recipe(ws.store, root)                       # 심부온도는 1:1 동률
    spec = json.loads(active_recipe(ws.store, root)["spec_json"])
    assert all(e["node_name"] != "심부온도" for e in spec["template"])
    assert any(d["node_name"] == "심부온도" for d in spec["dropped"])


def test_recipe_applies_to_same_format_new_doc(ws, tmp_path):
    """요구 1: 같은 형식 새 문서 → 레시피가 매핑을 이식 (judge 미개입)."""
    d1, _ = ws.ingest(_mini(tmp_path / "a.xlsx"))
    _map(ws, d1)
    root = _root_of(ws)
    snapshot_recipe(ws.store, root)

    d2, _ = ws.ingest(_mini(tmp_path / "new_doc.xlsx", values=(9.0, 8.0, 7.0)))
    stats = apply_recipe(ws.store, active_recipe(ws.store, root), d2)
    assert stats["applied"] >= 1 and stats["skipped_stale"] == 0
    m = ws.store.active_mapping(_core_node(ws, d2)["node_id"])
    assert m["concept_id"] == "core_temperature"
    assert m["method"] == "recipe"
    assert m["status"] == "AUTO_APPROVED"
    # 멱등: 재적용해도 이미 매핑된 노드는 건드리지 않는다
    again = apply_recipe(ws.store, active_recipe(ws.store, root), d2)
    assert again["applied"] == 0 and again["skipped_mapped"] >= 1


def test_recipe_relaxed_match_demotes_to_review(ws, tmp_path):
    """열 추가로 layout fingerprint가 바뀌면 완화 매칭 → REVIEW 강등."""
    d1, _ = ws.ingest(_mini(tmp_path / "a.xlsx"))
    _map(ws, d1)
    root = _root_of(ws)
    snapshot_recipe(ws.store, root)

    d2, _ = ws.ingest(_mini(tmp_path / "wider.xlsx", extra_col="메모지표"))
    apply_recipe(ws.store, active_recipe(ws.store, root), d2)
    m = ws.store.active_mapping(_core_node(ws, d2)["node_id"])
    assert m is not None and m["method"] == "recipe"
    assert m["status"] == "REVIEW_REQUIRED"               # 정확 일치 실패 → 완화
    ev = ws.store.conn.execute(
        "SELECT context_json FROM mapping_evidence WHERE mapping_id=?",
        (m["mapping_id"],)).fetchone()
    assert json.loads(ev["context_json"])["match"] == "relaxed"


def test_recipe_stale_concept_skipped(ws, tmp_path):
    d1, _ = ws.ingest(_mini(tmp_path / "a.xlsx"))
    _map(ws, d1)
    root = _root_of(ws)
    snapshot_recipe(ws.store, root)
    ws.store.conn.execute(                                 # 개념 소멸 시나리오
        "UPDATE domain_concept SET status='DEPRECATED' WHERE concept_id=?",
        ("core_temperature",))
    ws.store.commit()
    d2, _ = ws.ingest(_mini(tmp_path / "b.xlsx"))
    stats = apply_recipe(ws.store, active_recipe(ws.store, root), d2)
    assert stats["skipped_stale"] >= 1
    m = ws.store.active_mapping(_core_node(ws, d2)["node_id"])
    assert m is None                                       # judge로 위임 (미기록)


def test_suggest_groups_ranks_same_format(ws, tmp_path):
    d1, _ = ws.ingest(_mini(tmp_path / "a.xlsx"))
    _map(ws, d1)
    root = _root_of(ws)
    snapshot_recipe(ws.store, root)
    d2, _ = ws.ingest(_mini(tmp_path / "b.xlsx", values=(4.0, 5.0, 6.0)))
    sugg = suggest_groups(ws.store, d2)
    assert sugg and sugg[0]["root_concept_id"] == root
    assert sugg[0]["match_pct"] >= 50


# ----------------------------------------------------------- 재크롤링 ----
def test_recrawl_fill_retries_unmapped_after_kg_edit(ws, tmp_path):
    """요구 3: KG(alias) 보강 → fill 재크롤링이 UNMAPPED를 재평가."""
    f = _mini(tmp_path / "odd.xlsx", header="요상한지표XX")
    doc_id, _ = ws.ingest(f)
    _map(ws, doc_id)
    m0 = ws.store.active_mapping(_core_node(ws, doc_id, "요상한지표XX")["node_id"])
    assert m0["status"] == "UNMAPPED"

    ws.store.add_alias("core_temperature", "요상한지표XX", "요상한지표xx")
    ws.store.commit()
    from src.mapping.concepts import normalize_label
    ws.store.add_alias("core_temperature", "요상한지표XX",
                       normalize_label("요상한지표XX"))
    ws.store.commit()

    root = "domain_cost"                                   # 어느 그룹이든 멤버면 됨
    set_member_override(ws.store, root, doc_id, "INCLUDED")
    ws.store.commit()
    run_id = start_run(ws.store, root, None, "fill")
    res = run_recrawl(ws.store, threading.Lock(), ws, root, "fill",
                      [doc_id], run_id, ws.retriever(), RuleJudge())
    assert res["status"] == "SUCCESS"
    m1 = ws.store.active_mapping(_core_node(ws, doc_id, "요상한지표XX")["node_id"])
    assert m1["status"] == "AUTO_APPROVED"
    assert m1["concept_id"] == "core_temperature"
    run = ws.store.conn.execute(
        "SELECT * FROM recrawl_run WHERE run_id=?", (run_id,)).fetchone()
    assert run["status"] == "SUCCESS"
    assert json.loads(run["summary_json"])[0]["reset"] >= 1


def test_recrawl_reset_auto_preserves_human_decisions(ws, tmp_path):
    f = _mini(tmp_path / "a.xlsx")
    doc_id, _ = ws.ingest(f)
    _map(ws, doc_id)
    node = _core_node(ws, doc_id)
    m = ws.store.active_mapping(node["node_id"])
    remap_reviewed(ws.store, m["mapping_id"], "oven_temperature", "tester")
    human = ws.store.active_mapping(node["node_id"])
    assert human["status"] == "APPROVED"

    root = _root_of(ws, "oven_temperature")
    run_id = start_run(ws.store, root, None, "reset_auto")
    res = run_recrawl(ws.store, threading.Lock(), ws, root, "reset_auto",
                      [doc_id], run_id, ws.retriever(), RuleJudge())
    assert res["status"] == "SUCCESS"
    after = ws.store.active_mapping(node["node_id"])
    assert after["mapping_id"] == human["mapping_id"]      # 사람 결정 불가침
    assert after["status"] == "APPROVED"


def test_recrawl_missing_file_isolated(ws, tmp_path):
    f = _mini(tmp_path / "a.xlsx")
    doc_id, _ = ws.ingest(f)
    _map(ws, doc_id)
    f.unlink()                                             # 원본 소실
    root = _root_of(ws)
    run_id = start_run(ws.store, root, None, "fill")
    res = run_recrawl(ws.store, threading.Lock(), ws, root, "fill",
                      [doc_id], run_id, ws.retriever(), RuleJudge())
    assert res["status"] == "FAILED" and res["errors"] == 1
    run = ws.store.conn.execute(
        "SELECT summary_json FROM recrawl_run WHERE run_id=?", (run_id,)).fetchone()
    assert "찾을 수 없습니다" in json.loads(run["summary_json"])[0]["error"]


def test_reset_document_mappings_status_based(ws, tmp_path):
    doc_id, _ = ws.ingest(_mini(tmp_path / "a.xlsx"))
    _map(ws, doc_id)
    n_fill = reset_document_mappings(ws.store, doc_id, "fill")
    active = ws.store.conn.execute(
        """SELECT count(*) FROM semantic_mapping m JOIN tree_node n
           ON n.node_id=m.tree_node_id
           WHERE m.is_active=1 AND m.status='UNMAPPED' AND n.document_id=?""",
        (doc_id,)).fetchone()[0]
    assert active == 0                                     # fill은 UNMAPPED만
    n_reset = reset_document_mappings(ws.store, doc_id, "reset_auto")
    assert n_reset >= 1                                    # AUTO도 초기화


# ------------------------------------------------------- seed 3-모드 ----
def test_seed_merge_new_does_not_resurrect_deletions(ws):
    """--merge-new: DB에서 지운 alias/관계가 부활하지 않고, 기존 개념 무접촉."""
    ws.store.conn.execute(
        "DELETE FROM domain_alias WHERE concept_id='core_temperature' "
        "AND alias_norm='심부온도'")
    ws.store.conn.execute(
        "DELETE FROM domain_relation WHERE source_concept_id='core_temperature' "
        "AND relation_type='IS_A'")
    ws.store.conn.execute(
        "UPDATE domain_concept SET description='사용자 편집' "
        "WHERE concept_id='core_temperature'")
    ws.store.commit()
    info = load_domain_kg(ws.store, KG_YAML, UNITS_YAML, only_new_concepts=True)
    assert info["concepts"] == 0                           # 전부 기존 → 무접촉
    assert ws.store.conn.execute(
        "SELECT count(*) FROM domain_alias WHERE concept_id='core_temperature' "
        "AND alias_norm='심부온도'").fetchone()[0] == 0     # 삭제 유지
    assert ws.store.conn.execute(
        "SELECT count(*) FROM domain_relation WHERE source_concept_id="
        "'core_temperature' AND relation_type='IS_A'").fetchone()[0] == 0
    assert ws.store.concept("core_temperature")["description"] == "사용자 편집"


def test_seed_cli_guard_and_merge(tmp_path):
    """cmd_seed: 기본은 부트스트랩 1회, 재실행은 exit 2, --merge-new는 신규만."""
    from kg.cli import main
    wsdir = tmp_path / "ws"
    (wsdir / "config").mkdir(parents=True)
    shutil.copy(KG_YAML, wsdir / "config" / "domain_kg.yaml")
    shutil.copy(UNITS_YAML, wsdir / "config" / "units.yaml")
    assert main(["--ws", str(wsdir), "seed"]) == 0
    assert main(["--ws", str(wsdir), "seed"]) == 2         # DB가 진실
    # YAML에 신규 개념 추가 → merge-new가 그것만 들여온다
    extra = ("\n- concept_id: brand_new_metric\n  canonical_name: 신규지표\n"
             "  domain_level: L3\n  aliases: [새지표]\n")
    p = wsdir / "config" / "domain_kg.yaml"
    text = p.read_text(encoding="utf-8")
    p.write_text(text.replace("relations:", extra + "\nrelations:", 1),
                 encoding="utf-8")
    assert main(["--ws", str(wsdir), "seed", "--merge-new"]) == 0
    store = KgStore(wsdir / "data" / "kg" / "kg.db")
    assert store.concept("brand_new_metric") is not None
    store.close()


# ------------------------------------------------------- 웹 편집 API E2E ----
@pytest.fixture()
def client(tmp_path):
    wsdir = tmp_path / "ws"
    (wsdir / "config").mkdir(parents=True)
    (wsdir / "data" / "raw").mkdir(parents=True)
    shutil.copy(KG_YAML, wsdir / "config" / "domain_kg.yaml")
    shutil.copy(UNITS_YAML, wsdir / "config" / "units.yaml")
    store = KgStore(wsdir / "data" / "kg" / "kg.db")
    load_domain_kg(store, KG_YAML, UNITS_YAML)
    store.close()
    _mini(wsdir / "data" / "raw" / "mini.xlsx")
    _mini(wsdir / "data" / "raw" / "mini2.xlsx", values=(9.0, 9.5, 9.9))

    from fastapi.testclient import TestClient

    from kg.webapp import create_app
    with TestClient(create_app(wsdir)) as c:
        yield c


def test_webapp_ingest_recipe_recrawl_flow(client):
    """요구 1+3 E2E: 등록(제안) → 배정+레시피 → 재크롤링 폴링."""
    assert {f["filename"] for f in client.get("/api/raw-files").json()} == \
        {"mini.xlsx", "mini2.xlsx"}
    r = client.post("/api/ingest", json={"filename": "mini.xlsx"}).json()
    assert r["map"]["nodes"] >= 1
    root = None
    for g in client.get("/api/kg/document").json():
        if r["document_id"] in g["member_document_ids"]:
            root = g["id"]
            break
    assert root is not None
    assert client.post(f"/api/group/{root}/recipe", json={}).json()["ok"]

    # 새 파일: 분석만(map=false) → 같은 형식 DKG 제안 → 배정 확정
    r2 = client.post("/api/ingest",
                     json={"filename": "mini2.xlsx", "map": False}).json()
    assert r2["map"] is None
    assert r2["suggestions"][0]["root_concept_id"] == root
    r3 = client.post("/api/ingest",
                     json={"filename": "mini2.xlsx", "group_id": root}).json()
    assert r3["recipe"]["applied"] >= 1                    # 레시피가 이식됨

    rc = client.post(f"/api/group/{root}/recrawl", json={"mode": "fill"})
    run_id = rc.json()["run_id"]
    for _ in range(60):
        st = client.get(f"/api/recrawl/{run_id}").json()
        if st["status"] != "RUNNING":
            break
        time.sleep(0.5)
    assert st["status"] == "SUCCESS"
    assert len(st["summary"]) == 2                         # 두 문서 모두

    # 미등록 목록에서 사라짐
    assert client.get("/api/raw-files").json() == []


def test_webapp_kg_edit_endpoints(client):
    # 생성 → 수정(부분) → 폐기/복원
    r = client.post("/api/kg/concept",
                    json={"canonical_name": "테스트지표",
                          "domain_level": "L3"}).json()
    cid = r["concept_id"]
    assert r["created"]
    client.post("/api/kg/concept",
                json={"concept_id": cid, "description": "설명만 수정"})
    row = client.get("/api/kg/domain").json()
    me = next(n for n in row["nodes"] if n["id"] == cid)
    assert me["name"] == "테스트지표"                       # 이름 보존(부분 수정)
    assert client.post(f"/api/kg/concept/{cid}/deprecate").json()["ok"]
    assert client.post(f"/api/kg/concept/{cid}/restore").json()["ok"]

    # 사용 중 개념 폐기는 409
    client.post("/api/ingest", json={"filename": "mini.xlsx"})
    assert client.post("/api/kg/concept/core_temperature/deprecate"
                       ).status_code == 409

    # alias 추가/삭제
    assert client.post("/api/kg/alias",
                       json={"concept_id": cid, "alias": "별칭A"}).json()["ok"]
    assert client.delete("/api/kg/alias",
                         params={"concept_id": cid, "alias": "별칭A"}
                         ).json()["ok"]
    assert client.delete("/api/kg/alias",
                         params={"concept_id": cid, "alias": "별칭A"}
                         ).status_code == 404

    # 관계: IS_A 사이클 거부
    client.post("/api/kg/relation",
                json={"source": cid, "target": "domain_quality", "type": "IS_A"})
    assert client.post("/api/kg/relation",
                       json={"source": "domain_quality", "target": cid,
                             "type": "IS_A"}).status_code == 400

    # export에 편집 결과 반영
    text = client.get("/api/kg/export").text
    assert "테스트지표" in text


def test_webapp_member_override_and_bad_input(client):
    r = client.post("/api/ingest", json={"filename": "mini.xlsx"}).json()
    doc = r["document_id"]
    root = client.get("/api/kg/document").json()[0]["id"]
    assert client.post(f"/api/group/{root}/member",
                       json={"document_id": doc, "state": "EXCLUDED"}
                       ).json()["ok"]
    detail = client.get(f"/api/kg/document/{root}")
    if detail.status_code == 200:                          # 다른 문서가 남은 경우
        assert doc not in [d["document_id"]
                           for d in detail.json()["member_documents"]]
    assert client.delete(f"/api/group/{root}/member/{doc}").json()["ok"]
    # 경로 탈출/비 L1 거부
    assert client.post("/api/ingest",
                       json={"filename": "../evil.xlsx"}).status_code == 400
    assert client.post("/api/group/core_temperature/recrawl",
                       json={"mode": "fill"}).status_code == 400
    assert client.post(f"/api/group/{root}/recrawl",
                       json={"mode": "weird"}).status_code == 400
