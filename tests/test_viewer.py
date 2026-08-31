"""Read-only Viewer registration, security, cache and locator tests."""
from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from kg.store import KgStore
from kg.viewer import (ViewerError, register_unlocked, render_preview,
                       mark_validation_failed, sha256_file, source_locator,
                       validate_unlocked_xlsx)


def _xlsx(path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "190도"
    ws.merge_cells("A1:C1")
    ws["A1"] = "굽기 실험"
    ws["A1"].font = Font(bold=True, size=15)
    ws["A1"].fill = PatternFill("solid", fgColor="FFF2CC")
    ws.row_dimensions[1].height = 24
    ws.column_dimensions["A"].width = 20
    ws["G7"] = 43.2
    hidden = wb.create_sheet("내부계산")
    hidden.sheet_state = "hidden"
    wb.save(path)


def _registered(tmp_path: Path):
    staging = tmp_path / "staging"
    staging.mkdir()
    source = staging / "coffee.xlsx"
    _xlsx(source)
    store = KgStore(tmp_path / "kg.db")
    store.upsert_document("doc-1", "coffee.xlsx", str(source))
    version = store.add_version("doc-1", sha256_file(source), "test")
    metadata = register_unlocked(store, "doc-1", version, source, staging,
                                 tmp_path / "unlocked")
    store.commit()
    return store, source, version, metadata


def test_registers_validated_immutable_source_and_sheet_metadata(tmp_path):
    store, source, version, metadata = _registered(tmp_path)
    assert metadata["drm_status"] == "READY"
    assert metadata["sha256"] == sha256_file(source)
    assert metadata["sheet_count"] == 2
    row = store.conn.execute(
        "SELECT unlocked_path FROM viewer_document_version WHERE document_version=?",
        (version,)).fetchone()
    immutable = Path(row["unlocked_path"])
    assert immutable != source and immutable.is_file()
    assert immutable.stat().st_mode & 0o222 == 0
    sheets = store.conn.execute("SELECT * FROM viewer_sheet ORDER BY sheet_index").fetchall()
    assert [s["sheet_name"] for s in sheets] == ["190도", "내부계산"]
    assert json.loads(sheets[0]["merged_ranges_json"]) == ["A1:C1"]
    # Public metadata has no filesystem path.
    assert "unlocked_path" not in metadata and str(tmp_path) not in json.dumps(metadata)
    store.close()


def test_rejects_unvalidated_and_outside_staging_files(tmp_path):
    bad = tmp_path / "protected.xlsx"
    bad.write_bytes(b"encrypted payload" * 20)
    with pytest.raises(ViewerError, match="DRM_PROTECTED"):
        validate_unlocked_xlsx(bad)
    store = KgStore(tmp_path / "kg.db")
    store.upsert_document("doc-1", "protected.xlsx", str(bad))
    version = store.add_version("doc-1", hashlib.sha256(bad.read_bytes()).hexdigest(), "test")
    with pytest.raises(ViewerError, match="PATH_FORBIDDEN"):
        register_unlocked(store, "doc-1", version, bad, tmp_path / "staging",
                          tmp_path / "unlocked")
    mark_validation_failed(store, "doc-1", version, "DRM_PROTECTED")
    failed = store.conn.execute(
        "SELECT drm_status,drm_error FROM viewer_document_version").fetchone()
    assert dict(failed) == {"drm_status": "FAILED", "drm_error": "DRM_PROTECTED"}
    store.close()


def test_render_cache_is_disposable_and_source_hash_never_changes(tmp_path):
    store, _, version, metadata = _registered(tmp_path)
    fake = tmp_path / "fake-soffice"
    fake.write_text("""#!/usr/bin/env python3
import pathlib,sys
out = pathlib.Path(sys.argv[sys.argv.index('--outdir') + 1])
source = pathlib.Path(sys.argv[-1])
(out / (source.stem + '.pdf')).write_bytes(b'%PDF-1.4\\n%%EOF')
""", encoding="utf-8")
    fake.chmod(0o755)
    cache = tmp_path / "cache"
    before = metadata["sha256"]
    pdf = render_preview(store, "doc-1", version, cache, str(fake))
    assert pdf.read_bytes().startswith(b"%PDF")
    assert store.conn.execute(
        "SELECT sha256 FROM viewer_document_version WHERE document_version=?",
        (version,)).fetchone()[0] == before
    immutable_path = Path(store.conn.execute(
        "SELECT unlocked_path FROM viewer_document_version WHERE document_version=?",
        (version,)).fetchone()[0])
    assert sha256_file(immutable_path) == before
    pdf.unlink()  # cache is disposable and can be rebuilt from the XLSX
    assert render_preview(store, "doc-1", version, cache, str(fake)).is_file()
    store.close()


def test_source_locator_is_logical_and_includes_template_assignment(tmp_path):
    store, _, version, _ = _registered(tmp_path)
    from kg.parsing import add_version, assign, create_template
    create_template(store, "recipe", "Recipe", "financier")
    add_version(store, "recipe", {"sheet_templates": [{"name": "oven", "match": {"names": ["190도"]}, "mappings": []}]})
    assign(store, "doc-1", version, "recipe", 1)
    locator = source_locator(store, "doc-1", version, "190도", "g7:g20", "weight")
    assert locator["a1_range"] == "G7:G20"
    assert locator["range"] == {"start_row": 7, "start_col": 7,
                                "end_row": 20, "end_col": 7}
    assert locator["parsing_template"]["template_id"] == "recipe"
    assert not any(key in locator for key in ("x", "y", "width", "height", "path"))
    with pytest.raises(ViewerError, match="INVALID_RANGE"):
        source_locator(store, "doc-1", version, "190도", "../../A1", None)
    store.close()
