"""Fixture mutation generators (설계문서 §14.1 샘플 기반 회귀 fixture).

값 1셀 수정처럼 '다른 것은 그대로'가 중요한 변형은 openpyxl 왕복 시 수식
cached value가 사라지므로 XLSX 내부 XML을 직접 patch 한다. 구조 변형(행 이동,
시트명 변경 등)은 openpyxl로 수행하고 검증도 구조/키 수준으로 한다.
"""
from __future__ import annotations

import re
import shutil
import zipfile
from pathlib import Path

import openpyxl


def _patch_zip_member(src: Path, dst: Path, member: str, fn) -> None:
    with zipfile.ZipFile(src) as zin:
        items = zin.infolist()
        contents = {i.filename: zin.read(i.filename) for i in items}
    contents[member] = fn(contents[member].decode("utf-8")).encode("utf-8")
    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for i in items:
            zout.writestr(i, contents[i.filename])


def change_cell_value(src: Path, dst: Path, cell: str, old: str, new: str,
                      sheet_xml: str = "xl/worksheets/sheet1.xml") -> None:
    """한 셀의 값만 XML 수준에서 교체 — 수식 캐시/스타일/이미지 전부 보존.

    Excel/타 생성기 모두 지원: 태그의 네임스페이스 접두사(x: 등)를 허용한다.
    """
    pattern = re.compile(
        rf'(<(?:\w+:)?c r="{cell}"[^>]*>.*?<(?:\w+:)?v>){re.escape(old)}(</(?:\w+:)?v>)',
        re.DOTALL,
    )

    def patch(xml: str) -> str:
        out, n = pattern.subn(rf"\g<1>{new}\g<2>", xml)
        if n != 1:
            raise ValueError(f"cell {cell} with value {old!r} not found (n={n})")
        return out

    _patch_zip_member(src, dst, sheet_xml, patch)


def rename_sheet(src: Path, dst: Path, old_name: str, new_name: str) -> None:
    def patch(xml: str) -> str:
        if f'name="{old_name}"' not in xml:
            raise ValueError(f"sheet {old_name} not found")
        return xml.replace(f'name="{old_name}"', f'name="{new_name}"')

    _patch_zip_member(src, dst, "xl/workbook.xml", patch)


def shift_rows_down(src: Path, dst: Path, amount: int = 2, at_row: int = 3) -> None:
    """헤더 시작 위치를 아래로 이동 (§14.1). 수식 캐시는 사라지므로 검증은
    record key/블록 구조 수준으로 한다.

    openpyxl insert_rows는 병합 범위를 이동시키지 않으므로 직접 보정한다.
    """
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb.active
    ws.insert_rows(at_row, amount)
    from openpyxl.worksheet.cell_range import CellRange, MultiCellRange

    shifted = MultiCellRange()
    for rng in ws.merged_cells.ranges:
        cr = CellRange(str(rng))
        if cr.min_row >= at_row:
            cr.shift(0, amount)
        shifted.add(cr)
    ws.merged_cells = shifted
    wb.save(dst)


def replace_label(src: Path, dst: Path, old_label: str, new_label: str,
                  sheet_xml: str = "xl/worksheets/sheet1.xml") -> None:
    """헤더/라벨 문자열 교체 (동의어 mutation).

    문자열이 sheet XML(inline)과 sharedStrings.xml 중 어디에 있든 처리한다.
    """
    old_esc = f">{old_label}<"
    new_esc = f">{new_label}<"

    def patch(xml: str) -> str:
        if old_esc not in xml:
            raise ValueError(f"label {old_label!r} not found")
        return xml.replace(old_esc, new_esc)

    with zipfile.ZipFile(src) as z:
        sheet_has = old_esc in z.read(sheet_xml).decode("utf-8")
    _patch_zip_member(src, dst, sheet_xml if sheet_has else "xl/sharedStrings.xml", patch)


def delete_block_rows(src: Path, dst: Path, first_row: int, last_row: int) -> None:
    """반복 Block 하나 삭제 (§14.1 Block 추가/삭제)."""
    shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb.active
    ws.delete_rows(first_row, last_row - first_row + 1)
    wb.save(dst)


def combine_multi_sheet(sources: list[Path], dst: Path) -> None:
    """여러 fixture의 시트를 하나의 workbook으로 합쳐 다중 시트 입력을 만든다.

    실데이터의 '시트 여러 개' 시나리오 검증용. 셀 값/병합/fill을 복사한다
    (이미지·수식 캐시는 제외 — 다중 시트 세그먼테이션 검증이 목적).
    """
    out = openpyxl.Workbook()
    out.remove(out.active)
    from copy import copy

    for src in sources:
        wb = openpyxl.load_workbook(src)
        for ws in wb.worksheets:
            tgt = out.create_sheet(ws.title)
            for row in ws.iter_rows():
                for c in row:
                    nc = tgt.cell(row=c.row, column=c.column, value=c.value)
                    if c.has_style:
                        nc.font = copy(c.font)
                        nc.fill = copy(c.fill)
                        nc.border = copy(c.border)
                        nc.alignment = copy(c.alignment)
                        nc.number_format = c.number_format
            for m in ws.merged_cells.ranges:
                tgt.merge_cells(str(m))
    out.save(dst)
