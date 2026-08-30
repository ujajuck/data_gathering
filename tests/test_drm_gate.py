"""DRM 획득 게이트 회귀 — 잠금 감지 / 정식 해제 요청 / 해제본 자동 감지 (KG2).

원칙 검증: 시스템은 보호를 우회하지 않는다 — 잠긴 파일은 파싱·등록이 명시적으로
거부되고, 해제본(같은 파일명, 읽기 가능)이 도착해야 등록 흐름이 열린다.
"""
from __future__ import annotations

import shutil

import openpyxl
import pytest
from openpyxl.styles import Font, PatternFill

from tests.conftest import FIXTURES

from kg.acquisition import (create_request, refresh_release_states,
                            sniff_container)
from kg.domain.loader import load_domain_kg
from kg.store import KgStore

KG_YAML = FIXTURES.parent.parent / "domains" / "financier" / "config" / "domain_kg.yaml"
UNITS_YAML = FIXTURES.parent.parent / "domains" / "financier" / "config" / "units.yaml"

_OLE_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _real_xlsx(path):
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "TEST"
    fill = PatternFill("solid", fgColor="FFD9E1F2")
    for col, name in enumerate(["시점", "심부온도"], start=1):
        c = s.cell(row=1, column=col, value=name)
        c.font = Font(bold=True)
        c.fill = fill
    for i, v in enumerate((25.0, 45.0)):
        s.cell(row=2 + i, column=1, value=f"t{i}")
        s.cell(row=2 + i, column=2, value=v)
    wb.save(path)
    return path


def _locked_file(path, magic=_OLE_MAGIC):
    path.write_bytes(magic + b"\x00" * 512)     # DRM/암호화 컨테이너 모사
    return path


def test_sniff_container(tmp_path):
    assert sniff_container(_real_xlsx(tmp_path / "a.xlsx"))["locked"] is False
    ole = sniff_container(_locked_file(tmp_path / "b.xlsx"))
    assert ole["locked"] is True and ole["container"] == "ole_cfb"
    junk = sniff_container(_locked_file(tmp_path / "c.xlsx", b"FASOODRM"))
    assert junk["locked"] is True and junk["container"] == "unknown"


def test_request_lifecycle_release_detection(tmp_path):
    store = KgStore(tmp_path / "kg.db")
    raw = tmp_path / "raw"
    raw.mkdir()
    f = _locked_file(raw / "비밀문서.xlsx")

    res = create_request(store, raw, "비밀문서.xlsx", note="분석용")
    assert res["request_id"].startswith("DRM-")
    assert "비밀문서.xlsx" in res["request_text"]
    assert "SHA-256" in res["request_text"]
    row = store.conn.execute("SELECT * FROM drm_request").fetchone()
    assert row["status"] == "REQUESTED" and row["container"] == "ole_cfb"

    # 아직 잠긴 상태 → 해제 감지 없음
    assert refresh_release_states(store, raw) == 0

    # 해제본이 같은 파일명으로 도착 → RELEASED 자동 전환
    _real_xlsx(f)
    assert refresh_release_states(store, raw) == 1
    row = store.conn.execute("SELECT * FROM drm_request").fetchone()
    assert row["status"] == "RELEASED" and row["released_at"]

    # 읽을 수 있는 파일로는 요청 생성 불가
    with pytest.raises(ValueError):
        create_request(store, raw, "비밀문서.xlsx")
    store.close()


@pytest.fixture()
def client(tmp_path):
    wsdir = tmp_path / "ws"
    (wsdir / "config").mkdir(parents=True)
    (wsdir / "data" / "raw").mkdir(parents=True)
    shutil.copy(KG_YAML, wsdir / "config" / "domain_kg.yaml")
    shutil.copy(UNITS_YAML, wsdir / "config" / "units.yaml")
    store = KgStore(wsdir / "data" / "kg" / "kg.db")
    load_domain_kg(store, KG_YAML, UNITS_YAML)
    store.close()
    _locked_file(wsdir / "data" / "raw" / "drm문서.xlsx")

    from fastapi.testclient import TestClient

    from kg.webapp import create_app
    with TestClient(create_app(wsdir)) as c:
        yield c, wsdir


def test_webapp_drm_flow(client):
    """E2E: 잠김 표시 → 등록 400 → 해제 요청 → 해제본 도착 → 등록 → INGESTED."""
    c, wsdir = client
    rows = c.get("/api/raw-files").json()
    me = next(r for r in rows if r["filename"] == "drm문서.xlsx")
    assert me["locked"] is True and me["drm"] is None

    # 잠긴 파일 등록은 우회 없이 거부
    r = c.post("/api/ingest", json={"filename": "drm문서.xlsx"})
    assert r.status_code == 400 and "잠겨" in r.json()["detail"]

    # 정식 해제 요청 → 요청서 텍스트
    r = c.post("/api/drm/request",
               json={"filename": "drm문서.xlsx", "note": "실험 데이터 통합"}).json()
    assert r["ok"] and "실험 데이터 통합" in r["request_text"]
    me = next(x for x in c.get("/api/raw-files").json()
              if x["filename"] == "drm문서.xlsx")
    assert me["drm"]["status"] == "REQUESTED"

    # 해제본 도착(같은 파일명 교체) → 목록에서 RELEASED로 자동 전환
    _real_xlsx(wsdir / "data" / "raw" / "drm문서.xlsx")
    me = next(x for x in c.get("/api/raw-files").json()
              if x["filename"] == "drm문서.xlsx")
    assert me["locked"] is False and me["drm"]["status"] == "RELEASED"

    # 등록 성공 → 요청 이력 완결(INGESTED), 미등록 목록에서 소멸
    r = c.post("/api/ingest", json={"filename": "drm문서.xlsx"}).json()
    assert r["ok"] and r["map"]["nodes"] >= 1
    reqs = c.get("/api/drm").json()
    assert reqs[0]["status"] == "INGESTED"
    assert all(x["filename"] != "drm문서.xlsx"
               for x in c.get("/api/raw-files").json())
