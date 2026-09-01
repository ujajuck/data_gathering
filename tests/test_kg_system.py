"""Fixed Domain KG 시스템 회귀 (설계서 v0.1) — Phase 1~3 완료 기준.

Phase 1: 서로 다른 양식 문서에서 동일 개념 데이터를 KG 노드로 탐색 가능
Phase 2: 선택한 Concept/Source로 재현 가능한 관계형 결과 생성 (+lineage)
Phase 3: 문서 일부 변경 시 영향 범위만 재평가 (tree diff/fingerprint)
"""
from __future__ import annotations

import shutil
import sqlite3

import openpyxl
import pytest
from openpyxl.styles import Font, PatternFill

from tests.conftest import FIXTURES

from kg.domain.loader import load_domain_kg
from kg.integration.builder import build, define_project
from kg.integration.dag import Frame, run_dag
from kg.mapping.judge import RuleJudge
from kg.mapping.mapper import map_document
from kg.mapping.retriever import DomainRetriever
from kg.search import lineage_of, reverse_lookup
from kg.store import KgStore
from kg.tree.builder import load_workbook_tree
from kg.tree.diff import apply_tree

F_CHEESE = FIXTURES / "financier" / "치즈_휘낭시에_실험데이터_혼돈양식_v3.xlsx"
F_COFFEE = FIXTURES / "financier" / "커피_휘낭시에_실험데이터_혼돈양식_v3.xlsx"
KG_YAML = FIXTURES.parent.parent / "domains" / "financier" / "config" / "domain_kg.yaml"
UNITS_YAML = FIXTURES.parent.parent / "domains" / "financier" / "config" / "units.yaml"


@pytest.fixture()
def ws(tmp_path):
    """임시 워크스페이스: store + units + 파일 적재 헬퍼."""
    store = KgStore(tmp_path / "kg.db")
    info = load_domain_kg(store, KG_YAML, UNITS_YAML)

    class W:
        pass

    w = W()
    w.root = tmp_path
    w.store = store
    w.units = info["units"]

    def ingest(path, parser_version="t"):
        doc_id, drafts, h = load_workbook_tree(store, tmp_path, path, {}, w.units, None)
        return doc_id, apply_tree(store, doc_id, path.name, str(path), h,
                                  parser_version, drafts)

    w.ingest = ingest
    w.retriever = lambda: DomainRetriever(store, units=w.units)
    yield w
    store.close()


def _mini_workbook(path, header="심부온도", values=(25.0, 45.0, 65.0)):
    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "TEST"
    fill = PatternFill("solid", fgColor="FFD9E1F2")
    for col, name in enumerate(["시점", header], start=1):
        c = s.cell(row=1, column=col, value=name)
        c.font = Font(bold=True)
        c.fill = fill
    for i, v in enumerate(values):
        s.cell(row=2 + i, column=1, value=f"t{i}")
        s.cell(row=2 + i, column=2, value=v)
    wb.save(path)
    return path


# ------------------------------------------------------------ tree / diff ---
def test_seed_idempotent(ws):
    n1 = len(ws.store.concepts())
    load_domain_kg(ws.store, KG_YAML, UNITS_YAML)
    assert len(ws.store.concepts()) == n1 >= 50


def test_tree_persisted_and_reingest_unchanged(ws):
    doc_id, diff = ws.ingest(F_CHEESE)
    assert diff.summary()["added"] > 150 and not diff.changed
    types = {r["node_type"] for r in ws.store.active_nodes(doc_id).values()}
    assert {"DOCUMENT", "SHEET", "TABLE", "HEADER"} <= types
    _, diff2 = ws.ingest(F_CHEESE)
    s = diff2.summary()
    assert s["added"] == s["changed"] == s["removed"] == 0


def test_value_change_keeps_mapping(ws, tmp_path):
    """값만 바뀌면 semantic fingerprint 유지 → 매핑 재사용 (§12.2)."""
    f = _mini_workbook(tmp_path / "mini.xlsx")
    doc_id, _ = ws.ingest(f)
    map_document(ws.store, ws.retriever(), RuleJudge(), doc_id)
    node = next(r for r in ws.store.active_nodes(doc_id).values()
                if r["node_name"] == "심부온도")
    m1 = ws.store.active_mapping(node["node_id"])
    assert m1 is not None and m1["concept_id"] == "core_temperature"

    _mini_workbook(f, values=(26.0, 46.0, 66.0))          # 값만 수정
    _, diff = ws.ingest(f)
    s = diff.summary()
    assert s["changed"] == 0 and s["removed"] == 0        # 의미 지문 불변
    m2 = ws.store.active_mapping(node["node_id"])
    assert m2["mapping_id"] == m1["mapping_id"]           # 매핑 승계
    # payload는 새 버전으로 교체됨
    pv = ws.store.conn.execute(
        """SELECT value_num FROM payload_value pv
           JOIN data_payload p ON p.payload_id=pv.payload_id
           WHERE p.tree_node_id=? AND p.is_current=1 ORDER BY pv.row_idx""",
        (node["node_id"],)).fetchall()
    assert [r[0] for r in pv] == [26.0, 46.0, 66.0]


def test_header_rename_deactivates_mapping(ws, tmp_path):
    """헤더 텍스트 변경 = 노드 제거+추가, 기존 매핑 비활성화 (§12.1)."""
    f = _mini_workbook(tmp_path / "mini.xlsx")
    doc_id, _ = ws.ingest(f)
    map_document(ws.store, ws.retriever(), RuleJudge(), doc_id)
    node = next(r for r in ws.store.active_nodes(doc_id).values()
                if r["node_name"] == "심부온도")
    m1 = ws.store.active_mapping(node["node_id"])

    _mini_workbook(f, header="중심온도")                  # 헤더 표기 변경
    _, diff = ws.ingest(f)
    s = diff.summary()
    assert s["removed"] >= 1 and s["added"] >= 1
    old = ws.store.conn.execute(
        "SELECT is_active FROM semantic_mapping WHERE mapping_id=?",
        (m1["mapping_id"],)).fetchone()
    assert old["is_active"] == 0
    # 새 노드는 미매핑 상태 → map이 재평가
    stats = map_document(ws.store, ws.retriever(), RuleJudge(), doc_id)
    assert stats["nodes"] >= 1


# --------------------------------------------------------- mapping / search --
def test_mapping_statuses_and_evidence(ws):
    doc_id, _ = ws.ingest(F_CHEESE)
    stats = map_document(ws.store, ws.retriever(), RuleJudge(), doc_id)
    assert stats["AUTO_APPROVED"] > 100
    assert stats["AUTO_APPROVED"] / stats["nodes"] >= 0.8      # §16.1 정확도 기준
    row = ws.store.conn.execute(
        """SELECT e.context_json, e.candidates_json FROM semantic_mapping m
           JOIN mapping_evidence e ON e.mapping_id=m.mapping_id
           WHERE m.status='AUTO_APPROVED' LIMIT 1""").fetchone()
    assert "node_name" in row["context_json"] and "score" in row["candidates_json"]


def test_reverse_lookup_across_formats(ws):
    """Phase 1 완료 기준: 서로 다른 양식에서 같은 개념을 KG 노드로 탐색."""
    for f in (F_CHEESE, F_COFFEE):
        doc_id, _ = ws.ingest(f)
        map_document(ws.store, ws.retriever(), RuleJudge(), doc_id)
    res = reverse_lookup(ws.store, "core_temperature")
    assert len(res["documents"]) == 2                     # 두 문서 횡단
    headers = {s["header"] for s in res["sources"]}
    assert "core T" in headers                            # 치즈: 후행 라벨 전치
    assert any(s["unit"] == "°F" for s in res["sources"])  # 커피: 화씨 표기 보존
    assert res["total_rows"] > 200


def test_review_workflow(ws):
    doc_id, _ = ws.ingest(F_CHEESE)
    map_document(ws.store, ws.retriever(), RuleJudge(), doc_id)
    row = ws.store.conn.execute(
        "SELECT mapping_id FROM semantic_mapping WHERE status='REVIEW_REQUIRED' "
        "AND is_active=1 LIMIT 1").fetchone()
    if row is None:
        pytest.skip("no review-required mapping in fixture")
    ws.store.review(row["mapping_id"], "APPROVE", "tester")
    ws.store.commit()
    after = ws.store.conn.execute(
        "SELECT status FROM semantic_mapping WHERE mapping_id=?",
        (row["mapping_id"],)).fetchone()
    assert after["status"] == "APPROVED"
    hist = ws.store.conn.execute(
        "SELECT action, reviewer FROM review_history WHERE mapping_id=?",
        (row["mapping_id"],)).fetchall()
    assert [(h["action"], h["reviewer"]) for h in hist] == [("APPROVE", "tester")]


# ------------------------------------------------------------------- DAG ----
def _frame():
    return Frame(["a", "b"],
                 [{"a": 1.0, "b": "x"}, {"a": 2.0, "b": "y"}, {"a": None, "b": "x"}],
                 [{"a": None, "b": None}] * 3, {"document_id": "d"})


def test_dag_blocks():
    env = {}
    out = run_dag([_frame()], [{"op": "filter",
                                "config": {"column": "b", "op": "==", "value": "x"}}], env)
    assert len(out[0].rows) == 2
    out = run_dag([_frame()], [{"op": "null_handling",
                                "config": {"columns": ["a"], "mode": "drop"}}], env)
    assert len(out[0].rows) == 2
    out = run_dag([_frame()], [
        {"op": "value_mapping", "config": {"column": "b", "map": {"x": "X"}}},
        {"op": "derived_column",
         "config": {"name": "a2", "expr": {"op": "mul", "left": "a", "right": 2}}},
        {"op": "aggregate", "config": {"group_by": ["b"], "aggs": {"a": "sum"}}},
    ], env)
    rows = {r["b"]: r["a"] for r in out[0].rows}
    assert rows["X"] == 1.0 and rows["y"] == 2.0
    # union + deduplicate
    out = run_dag([_frame(), _frame()], [{"op": "union"},
                                         {"op": "deduplicate", "config": {"keys": ["a", "b"]}}], env)
    assert len(out) == 1 and len(out[0].rows) == 3
    # join on shared key column
    f1 = Frame(["k", "v1"], [{"k": "r1", "v1": 10}], [{"k": None, "v1": None}])
    f2 = Frame(["k", "v2"], [{"k": "r1", "v2": 20}], [{"k": None, "v2": None}])
    out = run_dag([f1, f2], [{"op": "join", "config": {"on": ["k"]}}], env)
    assert out[0].rows == [{"k": "r1", "v1": 10, "v2": 20}]
    with pytest.raises(Exception):
        run_dag([_frame()], [{"op": "nope"}], env)


# ----------------------------------------------------- integration / build --
def test_build_custom_rdbms_with_lineage(ws, tmp_path):
    """Phase 2 완료 기준: Concept/Source 선택 → 재현 가능한 관계형 결과 + 계보."""
    for f in (F_CHEESE, F_COFFEE):
        doc_id, _ = ws.ingest(f)
        map_document(ws.store, ws.retriever(), RuleJudge(), doc_id)
    iid = define_project(ws.store, {
        "name": "t_result",
        "fields": [
            {"name": "core_temp", "concept": "core_temperature",
             "type": "numeric", "unit": "℃"},
            {"name": "rise", "concept": "rise_height", "type": "numeric", "unit": "mm"},
        ],
        "transform": [
            {"op": "unit_convert"},
            {"op": "union"},
            {"op": "null_handling", "config": {"columns": ["core_temp"], "mode": "drop"}},
        ],
    })
    result = build(ws.store, iid, tmp_path / "builds", units=ws.units)
    assert result["status"] == "SUCCESS" and result["rows"] > 100

    con = sqlite3.connect(result["output_db"])
    try:
        cols = [r[1] for r in con.execute("PRAGMA table_info(t_result)")]
        assert {"core_temp", "rise", "_source_document_id", "_source_locator"} <= set(cols)
        # °F 소스(커피)가 ℃로 정규화: 심부온도는 130℃를 넘을 수 없다
        mx = con.execute("SELECT max(core_temp) FROM t_result").fetchone()[0]
        assert mx is not None and mx < 130
        row = con.execute(
            "SELECT _row_id FROM t_result WHERE core_temp IS NOT NULL").fetchone()
    finally:
        con.close()

    chain = lineage_of(ws.store, result["build_id"], row[0], "core_temp")
    assert chain["source_cell"] and chain["document"] and chain["tree_path"]
    assert "unit_convert" in chain["transform_path"]

    # 재현성: 같은 정의로 다시 빌드하면 같은 행 수
    result2 = build(ws.store, iid, tmp_path / "builds2", units=ws.units)
    assert result2["rows"] == result["rows"]


def test_unmapped_concept_rejected(ws):
    with pytest.raises(KeyError):
        reverse_lookup(ws.store, "no_such_concept")


# --------------------------------------------------- 리뷰 회귀 (확정 결함) ---
F_CHOCO = FIXTURES / "financier" / "초코_휘낭시에_실험데이터_혼돈양식_v3.xlsx"


def test_removed_node_revival_no_crash(ws, tmp_path):
    """헤더 개명 후 원복 — REMOVED 노드가 같은 경로로 부활해도 크래시 없이 ACTIVE 복원."""
    f = _mini_workbook(tmp_path / "mini.xlsx")
    doc_id, _ = ws.ingest(f)
    _mini_workbook(f, header="중심온도")
    ws.ingest(f)
    _mini_workbook(f, header="심부온도")                  # 원복
    _, diff = ws.ingest(f)
    assert diff.summary()["added"] >= 1                   # 부활 = added 취급
    node = next(r for r in ws.store.active_nodes(doc_id).values()
                if r["node_name"] == "심부온도")
    assert node["status"] == "ACTIVE" and node["removed_version_id"] is None


def test_apply_tree_rolls_back_on_failure(ws, tmp_path, monkeypatch):
    """도중 실패 시 유령 version/반쪽 트리가 남지 않는다 (원자성)."""
    import kg.tree.diff as diff_mod
    f = _mini_workbook(tmp_path / "mini.xlsx")

    calls = {"n": 0}
    orig = diff_mod._write_payload

    def boom(store, d, version_id):
        calls["n"] += 1
        if calls["n"] >= 2:
            raise RuntimeError("disk full")
        return orig(store, d, version_id)

    monkeypatch.setattr(diff_mod, "_write_payload", boom)
    with pytest.raises(RuntimeError):
        ws.ingest(f)
    monkeypatch.setattr(diff_mod, "_write_payload", orig)
    assert ws.store.conn.execute(
        "SELECT count(*) FROM document_version").fetchone()[0] == 0
    assert ws.store.conn.execute(
        "SELECT count(*) FROM tree_node").fetchone()[0] == 0
    doc_id, diff = ws.ingest(f)                           # 재시도는 정상 동작
    assert diff.summary()["added"] > 0


def test_project_name_injection_rejected(ws):
    """YAML의 이름은 신뢰 경계 밖 — 식별자 화이트리스트 위반은 정의 단계에서 거부."""
    for bad in ('../../../escaped', 'x" (junk TEXT); --', 'a b'):
        with pytest.raises(ValueError):
            define_project(ws.store, {"name": bad, "fields": [
                {"name": "f", "concept": None}]})
    with pytest.raises(ValueError):
        define_project(ws.store, {"name": "ok", "fields": [
            {"name": 'a" TEXT, "injected', "concept": None}]})
    with pytest.raises(ValueError):
        define_project(ws.store, {"name": "ok", "fields": [{"name": "f"}],
                                  "transform": [{"op": "nope"}]})


def test_variant_scope_joins_with_values(ws):
    """전치(caption) 표: variant 라벨('base')이 값 행(base@0…)에 브로드캐스트된다."""
    from kg.integration.builder import assemble_sources
    doc_id, _ = ws.ingest(F_CHOCO)
    map_document(ws.store, ws.retriever(), RuleJudge(), doc_id)
    iid = define_project(ws.store, {
        "name": "variant_join",
        "fields": [
            {"name": "variant", "concept": "addin", "type": "text"},
            {"name": "core_temp", "concept": "core_temperature", "type": "numeric"},
        ]})
    frames, _ = assemble_sources(ws.store, iid)
    both = [r for fr in frames for r in fr.rows
            if r.get("variant") is not None and r.get("core_temp") is not None]
    assert len(both) >= 30                                 # 이전엔 0 (키 불일치)
    assert {"base", "glaze"} <= {r["variant"] for r in both}


def test_map_before_seed_errors(tmp_path):
    store = KgStore(tmp_path / "empty.db")
    with pytest.raises(RuntimeError):
        map_document(store, DomainRetriever(store), RuleJudge())
    store.close()


def test_unit_convert_no_double_apply():
    """변환 후 lineage 단위가 갱신되어 블록 재적용에도 이중 변환이 없다."""
    from src.units.converter import UnitRegistry
    units = UnitRegistry.load(UNITS_YAML)
    f = Frame(["t"], [{"t": 77.7}],
              [{"t": {"payload_id": "p", "row_idx": 0, "node_id": "n", "unit": "°F"}}])
    env = {"units": units, "field_units": {"t": "℃"}}
    out = run_dag([f], [{"op": "unit_convert"}, {"op": "unit_convert"}], env)
    assert abs(out[0].rows[0]["t"] - 25.39) < 0.01         # 한 번만 변환됨


def test_unit_convert_skip_nodes_keeps_raw_value():
    """양식별 전처리 '원값 유지' — skip_nodes에 든 노드의 셀은 변환하지 않는다."""
    from src.units.converter import UnitRegistry
    units = UnitRegistry.load(UNITS_YAML)
    f = Frame(["t"], [{"t": 77.7}, {"t": 77.7}],
              [{"t": {"payload_id": "p", "row_idx": 0, "node_id": "keep", "unit": "°F"}},
               {"t": {"payload_id": "p", "row_idx": 1, "node_id": "conv", "unit": "°F"}}])
    env = {"units": units, "field_units": {"t": "℃"}}
    out = run_dag([f], [{"op": "unit_convert",
                         "config": {"skip_nodes": ["keep"]}}], env)
    assert out[0].rows[0]["t"] == 77.7                     # 원값 유지
    assert abs(out[0].rows[1]["t"] - 25.39) < 0.01         # 나머지는 정규화


def test_incompatible_unit_warned_not_converted():
    from src.units.converter import UnitRegistry
    units = UnitRegistry.load(UNITS_YAML)
    f = Frame(["x"], [{"x": 100.0}],
              [{"x": {"payload_id": "p", "row_idx": 0, "node_id": "n", "unit": "KRW"}}])
    env = {"units": units, "field_units": {"x": "℃"}}
    out = run_dag([f], [{"op": "unit_convert"}], env)
    assert out[0].rows[0]["x"] == 100.0                    # 원본 보존 (§15)
    assert env["warnings"] and env["warnings"][0]["from"] == "KRW"


def test_source_meta_survives_union_select_aggregate(ws):
    f1 = Frame(["k", "v"], [{"k": "a", "v": 1.0}], [{"k": None, "v": None}],
               meta={"document_id": "D1", "sheet": "S1"})
    out = run_dag([f1], [
        {"op": "union"},
        {"op": "select", "config": {"columns": ["k", "v"]}},
        {"op": "aggregate", "config": {"group_by": ["k"], "aggs": {"v": "sum"}}},
    ], {})
    ln = out[0].lineage[0]
    assert ln.get("__frame_meta__", {}).get("document_id") == "D1"


def test_value_mapping_numeric_keys():
    f = Frame(["c"], [{"c": 180.0}], [{"c": None}])
    out = run_dag([f], [{"op": "value_mapping",
                         "config": {"column": "c", "map": {"180": "mid"}}}], {})
    assert out[0].rows[0]["c"] == "mid"


def test_build_zero_columns_clear_error(ws, tmp_path):
    iid = define_project(ws.store, {"name": "empty_proj", "fields": [
        {"name": "f", "concept": None}]})
    with pytest.raises(ValueError, match="컬럼"):
        build(ws.store, iid, tmp_path / "b", units=ws.units)
